"""
merge.py v2.1 (copia valores y formulas, preserva formato maestro)
------------------------------------------------
- Lee todo el Excel de los tecnicos.
- Copia todas las celdas con datos o formulas desde los tecnicos.
- En el maestro, NO sobrescribe celdas que tengan formula.
- Mantiene orden por # de medidor (columna G), filtros y validaciones.
Hoja por defecto: "ERROR FINAL".
Fila inicial de datos: 9
"""

from dataclasses import dataclass, asdict
from typing import Any, Dict, List, Optional, Tuple, cast, Literal
from pathlib import Path
from functools import lru_cache
import copy
import re
import logging
import os
import posixpath
import time
import json
import hashlib
import tempfile
import zipfile
from datetime import datetime
from copy import copy as _copy
from collections import defaultdict
from itertools import islice
from xml.etree import ElementTree as ET
try:
    import psutil
except ImportError:
    psutil = None
logger = logging.getLogger(__name__)
# openpyxl: lectura/escritura de Excel (XLSX)
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import MergedCell  # celdas fusionadas (no ancla son solo lectura)
from openpyxl.formula.translate import Translator  # para trasladar referencias de formulas al pegar
from openpyxl.styles.borders import Border, Side  # <- canónico, Pylance lo resuelve bien

# =========================
#         CONFIG
# =========================
### Fila desde donde comienzan los datos tanto en tecnicos como en maestro.
START_ROW: int = 9
### Nombres de hojas por defecto (si no existen se usa la activa).
MASTER_SHEET_NAME: Optional[str] = "ERROR FINAL"
TECH_SHEET_NAME: Optional[str] = "ERROR FINAL"
### Columna clave para ordenar y detectar fin (# Medidor, columna G).
KEY_SERIE_COL: str = "G"
# Rango fijo para copiar bordes exactamente del origen
BORDER_SRC_START_COL: str = "A"
BORDER_SRC_END_COL: str = "BL"
# Columna de estado (valor visible distinto de 0 fuerza aceptacion de la fila)
STATE_COL = "I"
# Rango de columnas a evaluar para decidir si una fila esta completamente vacia (si lo necesitas)
EMPTY_CHECK_FROM_COL = "B"
EMPTY_CHECK_TO_COL   = "AS"
### (No usamos corte por clave vacia al leer: seguimos hasta max_row)
HARD_STOP_ON_FIRST_BLANK_IN_KEY: bool = False
### Columnas que deben tener dato para considerar una fila valida
REQUIRED_NONEMPTY_COLS = ["G", "M", "N", "Y", "Z", "AL"]
### Si True, las columnas requeridas pueden tener dato ya sea digitado o por formula.
### Si False, solo se aceptara si fue digitado (sin formula).
REQUIRED_ACCEPTS_FORMULA = True
@dataclass
class CellPayload:
    # value: lo que trae la celda (texto, numero o formula como string)
    value: Any
    # is_formula: flag si 'value' es una formula
    is_formula: bool
    # coord: direccion original (ej. "H12") para poder trasladar formulas al destino
    coord: str
    # display: valor calculado (al abrir con data_only=True); util para ordenar/validar
    display: Optional[Any] = None
class MergeError(Exception):
    """Base para errores de consolidacion."""
    pass
class MergeUserError(MergeError):
    """Errores imputables al contenido de los archivos."""
    pass
class MergeFileReadError(MergeError):
    """Errores al leer un archivo especifico."""
    def __init__(self, path: Path, cause: Exception):
        self.path = Path(path)
        self.cause = cause
        super().__init__(f"{self.path}: {cause}")
@dataclass
class TechnicianRow:
    cells: Dict[int, CellPayload]
    source_path: Path
    source_sheet: str
    source_row: int


@dataclass
class ProvenanceEntry:
    medidor: str
    origen_archivo: str
    origen_hoja: str
    origen_fila: int
    hash_fila: str
    insertado_en_fila: int
    timestamp: str


@dataclass
class ProvenanceArtifacts:
    timestamp_iso: str
    timestamp_slug: str
    consolidated_path: Path
    jsonl_path: Path
    report_path: Path
    sheet_added: bool
    technicians: int
    rows: int
    duplicates: Dict[str, List[ProvenanceEntry]]
    conflicts: Dict[str, List[ProvenanceEntry]]
    runtime_seconds: float
_LAST_PROVENANCE: Optional[ProvenanceArtifacts] = None
TRAZA_SHEET_NAME = "TRAZA"
def _normalize_medidor(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip().upper()
def _row_hash(row: Dict[int, CellPayload]) -> str:
    parts = []
    for col_idx in sorted(row):
        payload = row[col_idx]
        display = payload.display if payload.display is not None else payload.value
        parts.append(f"{col_idx}:{repr(payload.value)}:{repr(display)}:{int(payload.is_formula)}")
    digest = "|".join(parts).encode('utf-8', 'ignore')
    return hashlib.sha1(digest).hexdigest()
def get_last_provenance() -> Optional[ProvenanceArtifacts]:
    return _LAST_PROVENANCE
def clear_last_provenance() -> None:
    global _LAST_PROVENANCE
    _LAST_PROVENANCE = None
# =========================
#     UTILIDADES GENERALES


def _has_any_border(b: object) -> bool:
    if not isinstance(b, Border):
        return False
    return any([
        getattr(b.left, "style", None),
        getattr(b.right, "style", None),
        getattr(b.top, "style", None),
        getattr(b.bottom, "style", None),
        getattr(b.diagonal, "style", None) if getattr(b, "diagonal", None) else None,
    ])


# ---- Estilos válidos de borde para openpyxl ----
SideStyle = Literal[
    "thin", "medium", "thick", "double",
    "dashed", "dotted", "dashDot", "dashDotDot",
    "hair", "mediumDashed", "mediumDashDot",
    "mediumDashDotDot", "slantDashDot", "none"
]

_ALLOWED_SIDE_STYLES: set[str] = {
    "thin", "medium", "thick", "double",
    "dashed", "dotted", "dashDot", "dashDotDot",
    "hair", "mediumDashed", "mediumDashDot",
    "mediumDashDotDot", "slantDashDot", "none"
}

def _coerce_side_style(s: Optional[str]) -> Optional[SideStyle]:
    """Devuelve el estilo si es válido; en otro caso None (para contentar a Pylance)."""
    if s is None:
        return None
    s_norm = str(s)
    if s_norm in _ALLOWED_SIDE_STYLES:
        return cast(SideStyle, s_norm)
    return None


def copy_row_borders_exact(ws_src, ws_dst, src_row, dst_row, start_col_letter, end_col_letter):
    from openpyxl.utils import column_index_from_string
    from copy import copy as _copy
    sc = column_index_from_string(start_col_letter)
    ec = column_index_from_string(end_col_letter)
    for c in range(sc, ec + 1):
        b = getattr(ws_src.cell(row=src_row, column=c), "border", None)
        if isinstance(b, Border) and (
            (b.left and b.left.style) or (b.right and b.right.style) or
            (b.top and b.top.style) or (b.bottom and b.bottom.style)
        ):
            ws_dst.cell(row=dst_row, column=c).border = _copy(b)



def _border_has_visible_style(border: Optional[Border]) -> bool:
    if not isinstance(border, Border):
        return False
    sides: Tuple[Optional[Side], ...] = (border.left, border.right, border.top, border.bottom, border.diagonal)
    for side in sides:
        if isinstance(side, Side) and getattr(side, "style", None):
            return True
    return bool(
        getattr(border, "diagonalUp", False)
        or getattr(border, "diagonalDown", False)
        or getattr(border, "outline", False)
    )


@lru_cache(maxsize=32)
def _open_styles_wb(path_str: str):
    return load_workbook(Path(path_str), data_only=False, read_only=False)

def _scan_row_top_bottom_styles(ws_src, src_row: int, col_a: int, col_bl: int) -> Tuple[Optional[SideStyle], Optional[SideStyle]]:
    """
    Recorre A..BL en la fila de origen y devuelve el primer style válido
    encontrado para top y/o bottom. Si no hay, retorna (None, None).
    """
    top_style: Optional[SideStyle] = None
    bottom_style: Optional[SideStyle] = None

    for c in range(col_a, col_bl + 1):
        b = getattr(ws_src.cell(row=src_row, column=c), "border", None)
        if not isinstance(b, Border):
            continue
        if top_style is None:
            top_style = _coerce_side_style(getattr(b.top, "style", None))
        if bottom_style is None:
            bottom_style = _coerce_side_style(getattr(b.bottom, "style", None))
        if top_style and bottom_style:
            break

    return top_style, bottom_style

def copy_row_styles_exact(
    ws_src,
    ws_dst,
    src_row: int,
    dst_row: int,
    start_col_letter: str = "A",
    end_col_letter: str = "BL",
    *,
    copy_values: bool = False,
) -> None:
    from openpyxl.utils import column_index_from_string
    start_idx = column_index_from_string(start_col_letter)
    end_idx   = column_index_from_string(end_col_letter)

    # altura de fila
    try:
        ws_dst.row_dimensions[dst_row].height = ws_src.row_dimensions[src_row].height
    except Exception:
        pass

    for c in range(start_idx, end_idx + 1):
        s = ws_src.cell(row=src_row, column=c)
        d = ws_dst.cell(row=dst_row, column=c)

        if isinstance(s, MergedCell):
            continue

        if copy_values:
            d.value = s.value  # tal cual (ojo: también copia fórmulas)

        # ¡clave!: copiar estilos con copy() para evitar StyleProxy cross-workbook
        d.font          = _copy(s.font)
        d.fill          = _copy(s.fill)
        d.border        = _copy(s.border)
        d.alignment     = _copy(s.alignment)
        d.number_format = s.number_format  # es string, se puede asignar directo
        d.protection    = _copy(s.protection)

def _apply_separator_row(
    ws: Worksheet,
    row: int,
    start_col_letter: str = "A",
    end_col_letter: str = "BL",
    style: SideStyle = "thick",
) -> None:
    """Dibuja un borde superior en A..BL de la fila dada (separador visual)."""
    c0 = column_index_from_string(start_col_letter)
    c1 = column_index_from_string(end_col_letter)
    top_side = Side(style=style)
    for c in range(c0, c1 + 1):
        cell = ws.cell(row=row, column=c)
        cur = cell.border if isinstance(cell.border, Border) else Border()
        cell.border = Border(
            left=cur.left, right=cur.right,
            top=top_side, bottom=cur.bottom,
            diagonal=cur.diagonal, diagonalUp=cur.diagonalUp,
            diagonalDown=cur.diagonalDown, outline=cur.outline
        )


def apply_styles_from_sources_exact(
    consolidated_path: Path,
    ordered_rows: List["TechnicianRow"],
    *,
    start_col: str = "A",
    end_col: str = "BL",
    copy_values: bool = False,
    # --- nuevo para 1.5.3 ---
    separator_style: Optional[SideStyle] = "thick",  # None para desactivar
    group_by: Literal["file", "sheet"] = "file",     # bloque por archivo o por archivo+hoja
) -> None:
    """
    Abre el consolidado y aplica estilos exactos A..BL desde cada archivo/hoja/fila de origen.
    Si copy_values=True, también sobreescribe el valor de cada celda con el del técnico.
    Además, si separator_style está definido, dibuja un borde superior 'style' cuando
    cambia el bloque de técnico (según group_by).
    """
    try:
        wb_dst = load_workbook(consolidated_path, data_only=False)
    except Exception as exc:
        logger.warning("No se pudieron aplicar estilos (abrir destino): %s", exc)
        return

    ws_dst = _safe_get_sheet(wb_dst, MASTER_SHEET_NAME)

    # normaliza/valida estilo del separador
    sep_style: Optional[SideStyle] = _coerce_side_style(separator_style) if separator_style else None

    def _block_key(tr: "TechnicianRow") -> Tuple[str, ...]:
        if group_by == "sheet":
            return (str(tr.source_path), tr.source_sheet)
        # por defecto, agrupa solo por archivo
        return (str(tr.source_path),)

    last_key: Optional[Tuple[str, ...]] = None

    for i, tech_row in enumerate(ordered_rows):
        dst_r = START_ROW + i
        try:
            wb_src = _open_styles_wb(str(tech_row.source_path))  # cacheado
            if not wb_src:
                continue
            ws_src = _safe_get_sheet(wb_src, tech_row.source_sheet)

            # --- 1.5.3: separador si cambia el bloque de técnico ---
            cur_key = _block_key(tech_row)
            if last_key is not None and cur_key != last_key and sep_style:
                _apply_separator_row(ws_dst, dst_r, start_col, end_col, sep_style)

            # Copia exacta de la fila (estilos + opcional valores)
            copy_row_styles_exact(
                ws_src, ws_dst,
                src_row=tech_row.source_row,
                dst_row=dst_r,
                start_col_letter=start_col,
                end_col_letter=end_col,
                copy_values=copy_values,
            )

            last_key = cur_key

        except Exception as e:
            logger.debug("Fila %s: fallo al copiar estilos A..BL: %s", dst_r, e)

    try:
        wb_dst.save(consolidated_path)
    except Exception as exc:
        logger.warning("No se pudieron guardar los estilos en destino: %s", exc)


def apply_borders_from_sources(consolidated_path: Path, ordered_rows: List["TechnicianRow"]) -> None:
    """
    Copia bordes desde el técnico:
    - 1) Copia exacta celda-a-celda solo donde el origen trae borde.
    - 2) Escaneo por fila: si se detecta top/bottom en la fila origen, se replica a TODO A..BL.
    """
    try:
        wb_dst = load_workbook(consolidated_path, data_only=False)
    except Exception as exc:
        logger.warning("No se pudieron aplicar bordes (abrir destino): %s", exc)
        return

    ws_dst = _safe_get_sheet(wb_dst, MASTER_SHEET_NAME)
    col_a  = column_index_from_string("A")
    col_bl = column_index_from_string("BL")

    for i, tech_row in enumerate(ordered_rows):
        dst_r = START_ROW + i
        try:
            wb_src = _open_styles_wb(str(tech_row.source_path))  # lru_cache
            if not wb_src:
                continue
            ws_src = _safe_get_sheet(wb_src, tech_row.source_sheet)

            # (1) Copia exacta A..BL SOLO donde origen trae borde
            copy_row_borders_exact(ws_src, ws_dst,
                                   src_row=tech_row.source_row,
                                   dst_row=dst_r,
                                   start_col_letter="A",
                                   end_col_letter="BL")

            # (2) Overlay horizontal por scan: si hay top/bottom en alguna celda
            # de la fila del técnico, lo replicamos a TODA la fila A..BL
            top_style, bottom_style = _scan_row_top_bottom_styles(ws_src, tech_row.source_row, col_a, col_bl)
            if top_style or bottom_style:
                top_side = Side(style=top_style) if top_style else None
                bottom_side = Side(style=bottom_style) if bottom_style else None
                for c in range(col_a, col_bl + 1):
                    cell = ws_dst.cell(row=dst_r, column=c)
                    cur = cell.border if isinstance(cell.border, Border) else Border()
                    cell.border = Border(
                        left=cur.left, right=cur.right,
                        top=(top_side or cur.top),
                        bottom=(bottom_side or cur.bottom),
                        diagonal=cur.diagonal, diagonalUp=cur.diagonalUp, diagonalDown=cur.diagonalDown,
                        outline=cur.outline
                    )

        except Exception as e:
            logger.debug("Fila %s: fallo al copiar/replicar bordes A..BL: %s", dst_r, e)

    try:
        wb_dst.save(consolidated_path)
    except Exception as exc:
        logger.warning("No se pudieron guardar los bordes en destino: %s", exc)



# =========================
def _safe_get_sheet(wb, desired_name: Optional[str]) -> Worksheet:
    """Devuelve la hoja por nombre si existe; si no, la activa."""
    if desired_name:
        try:
            return cast(Worksheet, wb[desired_name])  # hoja por nombre
        except KeyError:
            return wb.active   # fallback: hoja activa
    return wb.active
def _is_nonempty_typed_cell(cell) -> bool:
    """
    True solo si la celda tiene contenido digitado por el usuario (no formula).
    Cuenta numeros, fechas, booleanos y strings no vacios. Ignora formulas.
    """
    if cell is None:
        return False
    if _is_formula_cell(cell):  # ignora formulas
        return False
    v = cell.value
    if v is None:
        return False
    if isinstance(v, str):
        return v.strip() != ""
    return True  # numeros, bool, datetime, etc.
def _is_formula_cell(cell) -> bool:
    """True si la celda contiene una formula (requiere abrir workbooks con data_only=False)."""
    v = cell.value  # puede ser None, str (incluye '=...') o tipos simples
    if v is None:
        return False
    if getattr(cell, "data_type", None) == "f":  # flag interno de openpyxl
        return True
    # Excel puede guardar la formula como string que empieza con '='
    return isinstance(v, str) and v.startswith("=")
def _is_nonempty_value(v) -> bool:
    """True si hay valor util (no None/''), sin juzgar si es formula o digitado."""
    if v is None:
        return False
    return str(v).strip() != ""  # ignora strings vacios/espacios
def natural_key(raw: object) -> Tuple[Tuple[int, object], ...]:
    """Clave de orden natural para valores tipo medidor."""
    if raw is None:
        return ((2, ""),)
    s = str(raw).strip()
    if not s:
        return ((2, ""),)
    segments = [segment for segment in re.split(r"(\d+)", s.upper()) if segment]
    key: List[Tuple[int, object]] = []
    for segment in segments:
        if segment.isdigit():
            key.append((1, int(segment)))
        else:
            key.append((0, segment))
    return tuple(key)
def _payload_actual_value(payload: Optional[CellPayload]) -> Any:
    """Retorna 'display' si existe (valor visible), si no, el 'value' crudo."""
    if payload is None:
        return None
    return payload.display if payload.display is not None else payload.value
# =========================
#   LECTURA DE TECNICOS (TODO CONTENIDO)
# =========================
def read_rows_from_technician_values_only(path: Path, required_cols: List[str]) -> List[TechnicianRow]:
    """
    Reglas de aceptacion de fila:
    - Acepta si TODAS las columnas obligatorias (required_cols) tienen dato
      (valor o formula si REQUIRED_ACCEPTS_FORMULA=True; solo digitado si False),
      **o** si el valor visible en Estado (columna I) es distinto de 0.
    - Si no cumple ninguna de las dos, se DESCARTA la fila, pero el recorrido continua hasta max_row.
    Se trabaja en modo read_only para evitar cargar todo el libro en memoria.
    """
    start_time = time.perf_counter()
    source_name = path.name
    wb_formula = load_workbook(path, data_only=False, read_only=True)
    wb_values = load_workbook(path, data_only=True, read_only=True)
    ws_formula = _safe_get_sheet(wb_formula, TECH_SHEET_NAME)
    ws_values = _safe_get_sheet(wb_values, TECH_SHEET_NAME)
    state_idx = column_index_from_string(STATE_COL) - 1
    key_idx = column_index_from_string(KEY_SERIE_COL) - 1
    start_col_idx = column_index_from_string("B") - 1
    required_indexes = [column_index_from_string(col) - 1 for col in required_cols]
    minimum_cols = [state_idx + 1, key_idx + 1, column_index_from_string("B")]
    minimum_cols.extend(idx + 1 for idx in required_indexes)
    fallback_col = max(minimum_cols)
    max_col = ws_formula.max_column or fallback_col
    if max_col < fallback_col:
        max_col = fallback_col
    max_row = ws_formula.max_row or START_ROW
    rows: List[TechnicianRow] = []
    processed = 0
    def _required_ok(row_formula) -> bool:
        for idx in required_indexes:
            if idx >= len(row_formula):
                return False
            cell = row_formula[idx]
            if REQUIRED_ACCEPTS_FORMULA:
                ok = _is_nonempty_value(cell.value)
            else:
                ok = _is_nonempty_typed_cell(cell)
            if not ok:
                return False
        return True
    def _estado_is_blank(row_values) -> bool:
        if state_idx >= len(row_values):
            return True
        v = row_values[state_idx].value
        if v is None:
            return True
        if isinstance(v, str):
            return v.strip() == ""
        return False

    def _estado_is_nonzero(row_values) -> bool:
        if state_idx >= len(row_values):
            return False
        v = row_values[state_idx].value
        if v is None:
            return False
        if isinstance(v, bool):
            return bool(v)
        if isinstance(v, (int, float)):
            try:
                return float(v) != 0.0
            except Exception:
                return True
        s = str(v).strip()
        if s == "":
            return False
        try:
            return float(s.replace(",", ".")) != 0.0
        except Exception:
            return s != "0"
    try:
        formula_iter = ws_formula.iter_rows(min_row=START_ROW, max_row=max_row, min_col=1, max_col=max_col)
        values_iter = ws_values.iter_rows(min_row=START_ROW, max_row=max_row, min_col=1, max_col=max_col)
        source_sheet = ws_formula.title
        for row_idx, (row_formula, row_values) in enumerate(zip(formula_iter, values_iter), start=START_ROW):
            processed += 1
            if HARD_STOP_ON_FIRST_BLANK_IN_KEY:
                if key_idx >= len(row_formula) or not _is_nonempty_value(row_formula[key_idx].value):
                    break
            if _estado_is_blank(row_values):
                continue
            if not (_required_ok(row_formula) or _estado_is_nonzero(row_values)):
                continue
            row_dict: Dict[int, CellPayload] = {}
            for col_idx_zero in range(start_col_idx, max_col):
                if col_idx_zero >= len(row_formula):
                    continue
                cell_formula = row_formula[col_idx_zero]
                value = cell_formula.value
                if not _is_nonempty_value(value):
                    continue
                display = None
                if col_idx_zero < len(row_values):
                    display = row_values[col_idx_zero].value
                row_dict[col_idx_zero + 1] = CellPayload(
                    value=value,
                    is_formula=_is_formula_cell(cell_formula),
                    coord=cell_formula.coordinate,
                    display=display,
                )
            if row_dict:
                rows.append(TechnicianRow(cells=row_dict, source_path=path, source_sheet=source_sheet, source_row=row_idx))
    finally:
        wb_formula.close()
        wb_values.close()
    elapsed = time.perf_counter() - start_time
    logger.info(
        "Lectura tecnico %s: filas aceptadas=%d de %d (%.2fs)",
        source_name,
        len(rows),
        processed,
        elapsed,
    )
    return rows
def sort_row_dicts_by_serie(rows: List[Dict[int, CellPayload]]) -> List[Dict[int, CellPayload]]:
    """Ordena las filas por la clave alfanumerica de la columna G (# medidor)."""
    pos_h = column_index_from_string(KEY_SERIE_COL)  # indice de G
    return sorted(rows, key=lambda d: natural_key(_payload_actual_value(d.get(pos_h))))
def sort_technician_rows(rows: List[TechnicianRow]) -> List[TechnicianRow]:
    """Ordena filas de tecnico conservando su metadata."""
    pos_h = column_index_from_string(KEY_SERIE_COL)
    return sorted(rows, key=lambda r: natural_key(_payload_actual_value(r.cells.get(pos_h))))
# =========================
#  LIMPIAR Y ESCRIBIR EN MAESTRO
# =========================
def _clear_destination_area_values_only(ws: Worksheet) -> None:
    """
    Limpia desde START_ROW hasta la ultima fila, pero **solo** celdas que NO contengan formula.
    Esto garantiza que cualquier formula preexistente (p. ej., B10 = B9) NO se borre.
    """
    last_row = ws.max_row or START_ROW   # ultima fila considerada
    last_col = ws.max_column or column_index_from_string("AN")  # ultima columna considerada
    for r in range(START_ROW, last_row + 1):
        for c in range(column_index_from_string("B"), last_col + 1):
            cell = ws.cell(row=r, column=c)
            if not _is_formula_cell(cell):  # solo borra contenido de celdas sin formula
                cell.value = None           # deja formato intacto

def _normalize_zip_path(path: str) -> str:
    sanitized = path.replace('\\', '/')
    normalized = posixpath.normpath(sanitized)
    if normalized.startswith('./'):
        normalized = normalized[2:]
    return normalized


def _resolve_zip_path(base_path: str, target: Optional[str]) -> Optional[str]:
    if not target:
        return None
    origin = posixpath.dirname(base_path)
    combined = posixpath.normpath(posixpath.join(origin, target))
    if combined.startswith('../'):
        return None
    return combined


def _load_content_types(zip_file: zipfile.ZipFile) -> tuple[dict[str, str], dict[str, str]]:
    try:
        xml_bytes = zip_file.read('[Content_Types].xml')
    except KeyError:
        return {}, {}
    try:
        root = ET.fromstring(xml_bytes)
    except ET.ParseError:
        return {}, {}
    ns = {'ct': 'http://schemas.openxmlformats.org/package/2006/content-types'}
    defaults: dict[str, str] = {}
    overrides: dict[str, str] = {}
    for node in root.findall('ct:Default', ns):
        ext = node.attrib.get('Extension')
        ctype = node.attrib.get('ContentType')
        if ext and ctype:
            defaults[ext] = ctype
    for node in root.findall('ct:Override', ns):
        part = node.attrib.get('PartName')
        ctype = node.attrib.get('ContentType')
        if part and ctype:
            overrides[part] = ctype
    return defaults, overrides


def _restore_master_images(master_path: Path, consolidated_path: Path, sheet_name: str) -> None:
    main_ns = 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'
    rel_office_ns = 'http://schemas.openxmlformats.org/officeDocument/2006/relationships'
    rel_pkg_ns = 'http://schemas.openxmlformats.org/package/2006/relationships'
    drawing_rel_type = f'{rel_office_ns}/drawing'
    image_rel_type = f'{rel_office_ns}/image'

    sheet_path: Optional[str] = None
    sheet_rels_path: Optional[str] = None
    drawing_templates: List[ET.Element] = []
    drawing_relationships: List[Tuple[str, str, str]] = []
    extra_files: Dict[str, bytes] = {}
    ct_overrides_needed: Dict[str, str] = {}
    ct_defaults_needed: Dict[str, str] = {}

    try:
        with zipfile.ZipFile(master_path, 'r') as master_zip:
            try:
                workbook_xml = master_zip.read('xl/workbook.xml')
                workbook_root = ET.fromstring(workbook_xml)
            except (KeyError, ET.ParseError):
                return
            ns = {'m': main_ns}
            sheet_elem = next(
                (
                    candidate
                    for candidate in workbook_root.findall('m:sheets/m:sheet', ns)
                    if candidate.get('name') == sheet_name
                ),
                None,
            )
            if sheet_elem is None:
                return
            sheet_rid = sheet_elem.attrib.get(f'{{{rel_office_ns}}}id')
            if not sheet_rid:
                return
            try:
                workbook_rels_root = ET.fromstring(master_zip.read('xl/_rels/workbook.xml.rels'))
            except (KeyError, ET.ParseError):
                return
            sheet_target = None
            for rel in workbook_rels_root.findall(f'{{{rel_pkg_ns}}}Relationship'):
                if rel.attrib.get('Id') == sheet_rid:
                    sheet_target = rel.attrib.get('Target')
                    break
            if not sheet_target:
                return
            sheet_path = _normalize_zip_path(f'xl/{sheet_target}')
            sheet_rels_path = _normalize_zip_path(
                posixpath.join('xl/worksheets/_rels', f"{posixpath.basename(sheet_path)}.rels")
            )
            try:
                sheet_rels_xml = master_zip.read(sheet_rels_path)
                sheet_rels_root = ET.fromstring(sheet_rels_xml)
            except (KeyError, ET.ParseError):
                return
            for rel in sheet_rels_root.findall(f'{{{rel_pkg_ns}}}Relationship'):
                if rel.attrib.get('Type') != drawing_rel_type:
                    continue
                rel_id = rel.attrib.get('Id')
                target = rel.attrib.get('Target')
                if rel_id and target:
                    drawing_relationships.append((rel_id, target, drawing_rel_type))
            if not drawing_relationships:
                return
            try:
                sheet_master_xml = master_zip.read(sheet_path)
                sheet_master_root = ET.fromstring(sheet_master_xml)
            except (KeyError, ET.ParseError):
                return
            drawing_templates = [copy.deepcopy(node) for node in sheet_master_root.findall(f'{{{main_ns}}}drawing')]
            if not drawing_templates:
                return
            defaults, overrides = _load_content_types(master_zip)
            for rel_id, rel_target, rel_type in drawing_relationships:
                drawing_path = _resolve_zip_path(sheet_path, rel_target)
                if not drawing_path or drawing_path not in master_zip.namelist():
                    continue
                extra_files[drawing_path] = master_zip.read(drawing_path)
                part_name = f'/{drawing_path}'
                ct_overrides_needed[part_name] = overrides.get(
                    part_name, 'application/vnd.openxmlformats-officedocument.drawing+xml'
                )
                drawing_rels_candidate = _normalize_zip_path(
                    posixpath.join('xl/drawings/_rels', f'{posixpath.basename(drawing_path)}.rels')
                )
                if drawing_rels_candidate in master_zip.namelist():
                    drawing_rels_bytes = master_zip.read(drawing_rels_candidate)
                    extra_files[drawing_rels_candidate] = drawing_rels_bytes
                    try:
                        drawing_rels_root = ET.fromstring(drawing_rels_bytes)
                    except ET.ParseError:
                        continue
                    for rel in drawing_rels_root.findall(f'{{{rel_pkg_ns}}}Relationship'):
                        if rel.attrib.get('Type') != image_rel_type:
                            continue
                        media_path = _resolve_zip_path(drawing_path, rel.attrib.get('Target'))
                        if not media_path or media_path not in master_zip.namelist():
                            continue
                        extra_files[media_path] = master_zip.read(media_path)
                        media_part = f'/{media_path}'
                        if media_part in overrides:
                            ct_overrides_needed[media_part] = overrides[media_part]
                        else:
                            ext = Path(media_path).suffix.lstrip('.').lower()
                            if ext:
                                ct_defaults_needed[ext] = defaults.get(ext, f'image/{ext}')
    except (zipfile.BadZipFile, FileNotFoundError):
        return

    if not sheet_path or not sheet_rels_path or not drawing_templates or not drawing_relationships:
        return
    if not extra_files:
        return

    try:
        with zipfile.ZipFile(consolidated_path, 'r') as target_zip:
            infos = list(target_zip.infolist())
            contents = {info.filename: target_zip.read(info.filename) for info in infos}
    except (zipfile.BadZipFile, FileNotFoundError):
        return

    sheet_bytes = contents.get(sheet_path)
    if sheet_bytes is None:
        return
    try:
        ET.register_namespace('', main_ns)
        ET.register_namespace('r', rel_office_ns)
        sheet_root = ET.fromstring(sheet_bytes)
    except ET.ParseError:
        return
    if not sheet_root.findall(f'{{{main_ns}}}drawing'):
        for template in drawing_templates:
            sheet_root.append(copy.deepcopy(template))
        if '{http://www.w3.org/2000/xmlns/}r' not in sheet_root.attrib:
            sheet_root.attrib['{http://www.w3.org/2000/xmlns/}r'] = rel_office_ns
        sheet_bytes_new = ET.tostring(sheet_root, encoding='utf-8', xml_declaration=True)
    else:
        sheet_bytes_new = sheet_bytes

    sheet_rels_bytes = contents.get(sheet_rels_path)
    sheet_rels_text = sheet_rels_bytes.decode('utf-8', errors='ignore') if sheet_rels_bytes else ''
    sheet_rels_modified = False
    for rel_id, rel_target, rel_type in drawing_relationships:
        desired = f'  <Relationship Id="{rel_id}" Type="{rel_type}" Target="{rel_target}"/>'
        pattern = re.compile(rf'<Relationship[^>]*Id="{re.escape(rel_id)}"[^>]*/?>')
        match = pattern.search(sheet_rels_text)
        if match:
            if 'Target="' + rel_target + '"' not in match.group(0) or 'Type="' + rel_type + '"' not in match.group(0):
                sheet_rels_text = sheet_rels_text[:match.start()] + desired + sheet_rels_text[match.end():]
                sheet_rels_modified = True
        else:
            if sheet_rels_text:
                insert_pos = sheet_rels_text.rfind('</Relationships>')
                if insert_pos == -1:
                    sheet_rels_text = (
                        '<?xml version="1.0" encoding="UTF-8"?>\n'
                        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">\n'
                        f'{desired}\n'
                        '</Relationships>\n'
                    )
                else:
                    prefix = sheet_rels_text[:insert_pos]
                    suffix = sheet_rels_text[insert_pos:]
                    if prefix and not prefix.endswith('\n'):
                        prefix += '\n'
                    sheet_rels_text = prefix + desired + '\n' + suffix
            else:
                sheet_rels_text = (
                    '<?xml version="1.0" encoding="UTF-8"?>\n'
                    '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">\n'
                    f'{desired}\n'
                    '</Relationships>\n'
                )
            sheet_rels_modified = True

    sheet_rels_bytes_new = sheet_rels_bytes
    if sheet_rels_modified:
        sheet_rels_bytes_new = sheet_rels_text.encode('utf-8')

    ct_bytes = contents.get('[Content_Types].xml')
    if ct_bytes is None:
        return
    ct_text = ct_bytes.decode('utf-8', errors='ignore')
    ct_modified = False
    for part_name, ctype in ct_overrides_needed.items():
        pattern = re.compile(rf'<Override[^>]*PartName="{re.escape(part_name)}"[^>]*/>')
        match = pattern.search(ct_text)
        if match:
            if f'ContentType="{ctype}"' not in match.group(0):
                new_tag = re.sub(r'ContentType="[^"]+"', f'ContentType="{ctype}"', match.group(0))
                ct_text = ct_text[:match.start()] + new_tag + ct_text[match.end():]
                ct_modified = True
        else:
            insert_pos = ct_text.rfind('</Types>')
            if insert_pos != -1:
                prefix = ct_text[:insert_pos]
                suffix = ct_text[insert_pos:]
                if prefix and not prefix.endswith('\n'):
                    prefix += '\n'
                new_line = f'  <Override PartName="{part_name}" ContentType="{ctype}"/>'
                ct_text = prefix + new_line + '\n' + suffix
                ct_modified = True
    for ext, ctype in ct_defaults_needed.items():
        pattern = re.compile(rf'<Default[^>]*Extension="{re.escape(ext)}"[^>]*/>')
        match = pattern.search(ct_text)
        if match:
            if f'ContentType="{ctype}"' not in match.group(0):
                new_tag = re.sub(r'ContentType="[^"]+"', f'ContentType="{ctype}"', match.group(0))
                ct_text = ct_text[:match.start()] + new_tag + ct_text[match.end():]
                ct_modified = True
        else:
            override_pos = ct_text.find('<Override')
            insert_pos = override_pos if override_pos != -1 else ct_text.rfind('</Types>')
            if insert_pos != -1:
                prefix = ct_text[:insert_pos]
                suffix = ct_text[insert_pos:]
                if prefix and not prefix.endswith('\n'):
                    prefix += '\n'
                new_line = f'  <Default Extension="{ext}" ContentType="{ctype}"/>'
                ct_text = prefix + new_line + '\n' + suffix
                ct_modified = True
    ct_bytes_new = ct_text.encode('utf-8') if ct_modified else ct_bytes

    replacements: Dict[str, bytes] = {}
    modified = False
    if sheet_bytes_new != sheet_bytes:
        replacements[sheet_path] = sheet_bytes_new
        modified = True
    if sheet_rels_bytes_new and sheet_rels_bytes_new != (sheet_rels_bytes or b''):
        replacements[sheet_rels_path] = sheet_rels_bytes_new
        modified = True
    if ct_modified:
        replacements['[Content_Types].xml'] = ct_bytes_new
        modified = True
    for path_key, data in extra_files.items():
        if contents.get(path_key) != data:
            replacements[path_key] = data
            modified = True
    if not modified:
        return

    fd, tmp_name = tempfile.mkstemp(suffix='.xlsx', dir=str(consolidated_path.parent))
    os.close(fd)
    try:
        with zipfile.ZipFile(tmp_name, 'w') as zout:
            written: set[str] = set()
            for info in infos:
                filename = info.filename
                data = replacements.get(filename, contents[filename])
                zout.writestr(info, data)
                written.add(filename)
            for path_key, data in replacements.items():
                if path_key in written:
                    continue
                zinfo = zipfile.ZipInfo(path_key)
                zinfo.compress_type = zipfile.ZIP_DEFLATED
                zout.writestr(zinfo, data)
        os.replace(tmp_name, consolidated_path)
    finally:
        if os.path.exists(tmp_name):
            try:
                os.remove(tmp_name)
            except OSError:
                pass

def write_rows_into_master_values_only(master_path: Path, row_dicts: List[Dict[int, CellPayload]]) -> Path:
    """
    Escribe la matriz de filas (valores y formulas) en el maestro:
    - Primero limpia solo celdas sin formula (para no perder formulas existentes).
    - Luego escribe cada fila respetando que si la celda destino tiene formula, NO se toca.
    - Si la celda del tecnico trae formula, se traduce a la nueva coordenada antes de asignarla.
    """
    wb = load_workbook(master_path, data_only=False)  # data_only=False para distinguir formulas
    ws = _safe_get_sheet(wb, MASTER_SHEET_NAME)       # hoja del maestro
    # 1) Limpiar (sin tocar formulas)
    _clear_destination_area_values_only(ws)
    max_col = ws.max_column or column_index_from_string("AN")
    # 2) Escribir fila a fila
    r = START_ROW   # fila destino incremental
    for row in row_dicts:
        for c_idx, payload in row.items():   # recorre solo columnas presentes en la fila origen
            if c_idx > max_col:
                continue  # evita escribir fuera del rango del maestro
            cell = ws.cell(row=r, column=c_idx)
            if isinstance(cell, MergedCell):
                continue  # celdas combinadas (no ancla) son solo lectura en openpyxl
            # --- NUEVA LOGICA: permitir reemplazar formula o blanco con digitado del tecnico ---
            dest_has_formula = _is_formula_cell(cell)
            dest_is_blank = (cell.value is None) or (str(cell.value).strip() == '')
            # "Fuente digitada" = payload no es formula y tiene valor no vacio
            source_is_typed = (not payload.is_formula) and (_is_nonempty_value(payload.value))
            if (dest_has_formula or dest_is_blank) and source_is_typed:
                # Escribir SIEMPRE el valor digitado del tecnico (sobrescribe formula o blanco)
                cell.value = payload.value
                continue  # saltar resto de reglas para esta celda
            # --- LOGICA EXISTENTE (se mantiene igual) ---
            # Si el destino tiene formula y NO hay fuente digitada, no tocar
            if dest_has_formula:
                continue
            # Si la fuente es formula, traducir y pegar; si falla, pegar valor visible
            if payload.is_formula and isinstance(payload.value, str):
                try:
                    translator = Translator(payload.value, payload.coord)
                    cell.value = translator.translate_formula(cell.coordinate)
                except Exception:
                    cell.value = _payload_actual_value(payload)
            else:
                # Valor literal (numero/texto/fecha) se copia tal cual
                cell.value = payload.value
        r += 1  # siguiente fila destino
    out = master_path.with_name(master_path.stem + "_R.xlsx")  # nombre de salida
    wb.save(out)  # guarda libro resultante
    logger.info("Maestro escrito: %s (filas=%d)", out.name, len(row_dicts))
    return out
def _maybe_add_traza_sheet(consolidated_path: Path, entries: List[ProvenanceEntry]) -> bool:
    if not entries:
        return False
    try:
        wb = load_workbook(consolidated_path, data_only=False)
        if TRAZA_SHEET_NAME in wb.sheetnames:
            del wb[TRAZA_SHEET_NAME]
        ws = wb.create_sheet(TRAZA_SHEET_NAME)
        ws.sheet_state = 'hidden'
        ws.append(["medidor", "origen_archivo", "origen_hoja", "origen_fila", "hash_fila", "insertado_en_fila", "timestamp"])
        for entry in entries:
            ws.append([
                entry.medidor,
                entry.origen_archivo,
                entry.origen_hoja,
                entry.origen_fila,
                entry.hash_fila,
                entry.insertado_en_fila,
                entry.timestamp,
            ])
        wb.save(consolidated_path)
        return True
    except Exception as exc:
        logger.exception("No se pudo crear hoja TRAZA: %s", exc)
        return False
def _render_duplicates_table(data: Dict[str, List[ProvenanceEntry]], limit: int = 20) -> List[str]:
    if not data:
        return ['Sin duplicados detectados.']
    lines = ['| Medidor | Total | Archivos | Hashes |', '|---|---|---|---|']
    for medidor, entries in islice(sorted(data.items(), key=lambda item: item[0] or ''), limit):
        files = ', '.join(sorted({e.origen_archivo for e in entries}))
        hashes = ', '.join(sorted({e.hash_fila for e in entries}))
        label = medidor if medidor else '(vacio)'
        lines.append(f"| {label} | {len(entries)} | {files} | {hashes} |")
    return lines
def _render_conflicts(conflicts: Dict[str, List[ProvenanceEntry]], limit: int = 20) -> List[str]:
    if not conflicts:
        return ['Sin conflictos detectados.']
    lines: List[str] = []
    for medidor, entries in islice(sorted(conflicts.items(), key=lambda item: item[0] or ''), limit):
        label = medidor if medidor else '(vacio)'
        unique_hashes = sorted({e.hash_fila for e in entries})
        lines.append(f"- **{label}** ({len(unique_hashes)} hash(es), {len(entries)} fila(s))")
        for entry in entries:
            lines.append(
                f"  - hash {entry.hash_fila} · archivo {entry.origen_archivo} · hoja {entry.origen_hoja} · fila {entry.origen_fila}"
            )
    return lines
def _build_markdown_report(artifacts: ProvenanceArtifacts, sheet_requested: bool) -> str:
    lines: List[str] = []
    lines.append('# Reporte de trazabilidad 1.3.4')
    lines.append('')
    lines.append(f"- Fecha UTC: {artifacts.timestamp_iso}")
    lines.append(f"- Tiempo total: {artifacts.runtime_seconds:.2f}s")
    lines.append(f"- Tecnicos procesados: {artifacts.technicians}")
    lines.append(f"- Filas insertadas: {artifacts.rows}")
    lines.append(f"- Consolidado: {artifacts.consolidated_path}")
    lines.append(f"- JSONL: {artifacts.jsonl_path}")
    lines.append(f"- Hoja TRAZA solicitada: {'si' if sheet_requested else 'no'}")
    lines.append(f"- Hoja TRAZA creada: {'si' if artifacts.sheet_added else 'no'}")
    lines.append('')
    lines.append('## Duplicados por medidor')
    lines.extend(_render_duplicates_table(artifacts.duplicates))
    lines.append('')
    lines.append('## Conflictos por medidor')
    lines.extend(_render_conflicts(artifacts.conflicts))
    lines.append('')
    lines.append('## Archivos generados')
    lines.append(f"- JSONL: {artifacts.jsonl_path}")
    lines.append(f"- Reporte: {artifacts.report_path}")
    return '\n'.join(lines) + '\n'
def _generate_provenance_artifacts(
    ordered_rows: List[TechnicianRow],
    consolidated_path: Path,
    technicians_count: int,
    runtime_seconds: float,
    provenance_dir: Optional[Path] = None,
) -> ProvenanceArtifacts:
    timestamp_dt = datetime.utcnow().replace(microsecond=0)
    timestamp_iso = timestamp_dt.isoformat()
    timestamp_slug = timestamp_dt.strftime('%Y%m%d_%H%M')

    base_dir = provenance_dir or Path('reports') / 'trazabilidad'
    base_dir.mkdir(parents=True, exist_ok=True)

    entries: List[ProvenanceEntry] = []
    grouped: Dict[str, List[ProvenanceEntry]] = defaultdict(list)
    pos_h = column_index_from_string(KEY_SERIE_COL)

    for offset, tech_row in enumerate(ordered_rows):
        medidor_value = _payload_actual_value(tech_row.cells.get(pos_h))
        medidor_text = '' if medidor_value is None else str(medidor_value).strip()
        normalized = _normalize_medidor(medidor_value)
        hash_value = _row_hash(tech_row.cells)
        entry = ProvenanceEntry(
            medidor=medidor_text,
            origen_archivo=tech_row.source_path.name,
            origen_hoja=tech_row.source_sheet,
            origen_fila=tech_row.source_row,
            hash_fila=hash_value,
            insertado_en_fila=START_ROW + offset,
            timestamp=timestamp_iso,
        )
        entries.append(entry)
        grouped[normalized].append(entry)

    duplicates_display: Dict[str, List[ProvenanceEntry]] = {
        (entries_list[0].medidor or key): entries_list
        for key, entries_list in grouped.items()
        if len(entries_list) > 1
    }

    conflicts_display: Dict[str, List[ProvenanceEntry]] = {
        medidor: entries_list
        for medidor, entries_list in duplicates_display.items()
        if len({e.hash_fila for e in entries_list}) > 1
    }

    jsonl_path = base_dir / f"traza_1_3_4_{timestamp_slug}.jsonl"
    with jsonl_path.open('w', encoding='utf-8') as handle:
        for entry in entries:
            handle.write(json.dumps(asdict(entry), ensure_ascii=False) + '\n')

    sheet_requested = os.environ.get('OI_TRZ_SHEET', '0') == '1'
    sheet_added = False
    if sheet_requested:
        sheet_added = _maybe_add_traza_sheet(consolidated_path, entries)

    report_path = base_dir / f"reporte_1_3_4_{timestamp_slug}.md"
    artifacts = ProvenanceArtifacts(
        timestamp_iso=timestamp_iso,
        timestamp_slug=timestamp_slug,
        consolidated_path=consolidated_path,
        jsonl_path=jsonl_path,
        report_path=report_path,
        sheet_added=sheet_added,
        technicians=technicians_count,
        rows=len(entries),
        duplicates=duplicates_display,
        conflicts=conflicts_display,
        runtime_seconds=runtime_seconds,
    )

    report_content = _build_markdown_report(artifacts, sheet_requested)
    report_path.write_text(report_content, encoding='utf-8')

    return artifacts

# =========================
#  VALIDACIONES (las existentes)
# =========================
### Valores por defecto para CF si aun no existe Parametros.xlsx
DEFAULT_PARAMS = {
    "Q3": {3: (0.5, 9.2), 4: (0.5, 9.2), 5: (0.5, 9.2), 6: (0.5, 9.15)},
    "Q2": {3: (1.05, 1.124), 4: (0.5, 9.2), 5: (1.5, 1.56), 6: (1.45, 1.52)},
    "Q1": {3: (1.05, 1.124), 4: (0.5, 9.2), 5: (1.5, 1.56), 6: (1.45, 1.52)},
}
ParamMap = Dict[str, Dict[int, Tuple[float, float]]]  # alias de tipos para claridad
def load_params_from_excel(params_path: Path) -> ParamMap:
    """
    Lee Parametros.xlsx (hoja BANCOS) con columnas: Banco, Q, Min, Max.
    Devuelve: {"Q3": {3:(min,max),...}, "Q2": {...}, "Q1": {...}}
    """
    if not params_path.exists():
        return DEFAULT_PARAMS # si no existe archivo, usar defaults
    wb = load_workbook(params_path, data_only=True)  # solo valores
    sh = "BANCOS"
    if sh not in wb.sheetnames:
        return DEFAULT_PARAMS  # si no existe la hoja esperada
    ws = wb[sh]
    out: ParamMap = {"Q3": {}, "Q2": {}, "Q1": {}}  # estructura destino
    # Encabezados → mapa nombre→columna
    headers: Dict[str, int] = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=c).value
        if v is not None:
            headers[str(v).strip().lower()] = c
    for need in ("banco", "q", "min", "max"):
        if need not in headers:
            return DEFAULT_PARAMS # faltan columnas requeridas
    def _to_float(x):
        try:
            return float(str(x).replace(",", ".")) # admite comas decimales
        except Exception:
            return None
    r = 2  # primera fila de datos
    while r <= ws.max_row:
        raw_banco = ws.cell(row=r, column=headers["banco"]).value
        raw_q = ws.cell(row=r, column=headers["q"]).value
        raw_min = ws.cell(row=r, column=headers["min"]).value
        raw_max = ws.cell(row=r, column=headers["max"]).value
        # Normalizacion de tipos
        try:
            banco = int(float(str(raw_banco).strip()))
        except Exception:
            banco = None
        q = (str(raw_q).strip().upper() if raw_q is not None else "")
        vmin = _to_float(raw_min)
        vmax = _to_float(raw_max)
        # Guarda si hay consistencia y pertenece al set esperado
        if q in ("Q3", "Q2", "Q1") and banco in (3, 4, 5, 6) and vmin is not None and vmax is not None:
            out[q][banco] = (vmin, vmax)
        r += 1
    # Completar faltantes con defaults para bancos 3..6
    for q in ("Q3", "Q2", "Q1"):
        for b in (3, 4, 5, 6):
            if b not in out[q]:
                out[q][b] = DEFAULT_PARAMS[q][b]
    return out
def _last_data_row(ws: Worksheet, key_col_letter: str = KEY_SERIE_COL, start_row: int = 9) -> int:
    """Ultima fila con algo en la columna clave (G) contando desde start_row."""
    last = ws.max_row or start_row   # ultima fila usada
    last_found = start_row - 1       # si no hay datos, queda en start_row - 1
    col_idx = column_index_from_string(key_col_letter)
    for r in range(start_row, last + 1):
        v = ws.cell(row=r, column=col_idx).value
        if v not in (None, ""):
            last_found = r
    return last_found
def apply_pressure_cf(master_path: Path, params_path: Optional[Path] = None) -> None:
    """
    Validacion de presion deshabilitada; se conserva el archivo sin aplicar formato condicional.
    """
    logger.info("Validacion de presion deshabilitada; no se aplican reglas de formato condicional.")
    return

# =========================
#   FUNCION PRINCIPAL
# =========================
def _format_memory_usage() -> Optional[str]:
    if psutil is None:
        return None
    try:
        process = psutil.Process(os.getpid())
        rss_mb = process.memory_info().rss / (1024 * 1024)
        return f"{rss_mb:.1f} MB RSS"
    except Exception:
        return None
def build_and_write(
    master_path: Path,
    technician_paths: List[Path],
    *,
    order_by_col_g: bool = True,
    collect_provenance: bool = True,
    provenance_dir: Optional[Path] = None,
) -> Path:
    """
    - Lee TODOS los Excels de tecnicos (valores y formulas, conservando el valor evaluado para ordenar).
    - Ordena por # Medidor usando el contenido visible.
    - Escribe solo en celdas sin formula del maestro (conserva formulas existentes).
    - Reaplica las validaciones/CF.
    - Opcionalmente genera artefactos de trazabilidad y hoja oculta TRAZA.
    """
    start_time = time.perf_counter()
    logger.info(
        "Iniciando consolidacion: maestro=%s, tecnicos=%d",
        master_path.name,
        len(technician_paths),
    )
    clear_last_provenance()
    rows_with_meta: List[TechnicianRow] = []
    total_rows = 0
    for p in technician_paths:
        t_start = time.perf_counter()
        try:
            rows = read_rows_from_technician_values_only(p, REQUIRED_NONEMPTY_COLS)
        except Exception as exc:
            raise MergeFileReadError(p, exc) from exc
        if rows:
            rows_with_meta.extend(rows)
            total_rows += len(rows)
            logger.info(
                "Archivo tecnico %s: filas validas=%d (%.2fs)",
                p.name,
                len(rows),
                time.perf_counter() - t_start,
            )
        else:
            logger.warning("Archivo tecnico %s sin filas validas", p.name)
    if not rows_with_meta:
        raise MergeUserError("No se encontraron filas validas en los archivos de tecnicos.")
    if order_by_col_g:
         logger.info("Total de filas combinadas antes de ordenar: %d", total_rows)
         ordered_rows = sort_technician_rows(rows_with_meta)
    else:
         logger.info("Total de filas combinadas (sin reordenar): %d", total_rows)
         ordered_rows = rows_with_meta
    ordered_payloads = [row.cells for row in ordered_rows]
    try:
        out = write_rows_into_master_values_only(master_path, ordered_payloads)
    except Exception as exc:
        raise MergeFileReadError(master_path, exc) from exc
    try:
        apply_pressure_cf(out, Path(__file__).parent / "Parametros.xlsx")
    except Exception as exc:
        logger.warning("No se pudieron aplicar validaciones: %s", exc)
    #try:
    #    apply_borders_from_sources(out, ordered_rows)
    #except Exception as exc:
    #    logger.warning("No se pudieron aplicar bordes exactos: %s", exc)
    try:
    # PRUEBA: copia valores + formatos exactos (incluye bordes) A..BL
        apply_styles_from_sources_exact(
    out,
    ordered_rows,
    start_col="A",
    end_col="BL",
    copy_values=False,
    # separator_style="double",  # <- si quieres probar “double”
    # group_by="sheet",          # <- si prefieres separador por archivo+hoja
)
    except Exception as exc:
        logger.warning("No se pudieron aplicar estilos exactos: %s", exc)

    # Restaurar imágenes DESPUÉS de aplicar bordes (evita errores de namespaces al re-guardar)
    try:
        _restore_master_images(master_path, out, MASTER_SHEET_NAME or 'ERROR FINAL')
    except Exception as exc:
        logger.warning('No se pudieron restaurar imagenes del maestro (post-bordes): %s', exc)
    elapsed = time.perf_counter() - start_time
    mem_usage = _format_memory_usage()
    if mem_usage:
        logger.info(
            "Consolidacion completada en %.2fs (memoria %s, filas=%d)",
            elapsed,
            mem_usage,
            len(ordered_rows),
        )
    else:
        logger.info(
            "Consolidacion completada en %.2fs (filas=%d)",
            elapsed,
            len(ordered_rows),
        )
    if collect_provenance:
        artifacts = _generate_provenance_artifacts(
            ordered_rows,
            out,
            len(technician_paths),
            elapsed,
            provenance_dir=provenance_dir,
        )
        logger.info(
            "Trazabilidad generada: jsonl=%s, reporte=%s, duplicados=%d, conflictos=%d",
            artifacts.jsonl_path,
            artifacts.report_path,
            len(artifacts.duplicates),
            len(artifacts.conflicts),
        )
        global _LAST_PROVENANCE
        _LAST_PROVENANCE = artifacts
    return out  # ruta del maestro generado (_R.xlsx)
