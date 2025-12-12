from __future__ import annotations
from typing import cast
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import Cell as XLCell
from app.services.integrations.vima_to_lista import VimaToListaConfig, map_vima_to_lista

def test_map_vima_to_lista_incremental() -> None:
    wb_v, wb_l = Workbook(), Workbook()
    ws_v: Worksheet = cast(Worksheet, wb_v.active)
    ws_l: Worksheet = cast(Worksheet, wb_l.active)
    ws_v.title, ws_l.title = "VIMA", "LISTA"

    # LISTA ya tiene hasta OI-1106-2025
    ws_l["B11"] = "OI-1104-2025"
    ws_l["B12"] = "OI-1105-2025"
    ws_l["B13"] = "OI-1106-2025"

    # VIMA trae 1104..1109 (debe copiar solo 1107..1109)
    row = 11
    for oi in ["OI-1104-2025", "OI-1105-2025", "OI-1106-2025", 
               "OI-1107-2025","OI-1108-2025", "OI-1109-2025"
               ]:
        # ✅ Casteamos a Cell para que Pylance no se queje de MergedCell/None
        cast(XLCell,ws_v.cell(row=row, column=3)).value = oi     # C = OI
        cast(XLCell,ws_v.cell(row=row, column=7)).value = 1      # G..N llenos mínimamente
        cast(XLCell,ws_v.cell(row=row, column=8)).value = "Cert"
        cast(XLCell,ws_v.cell(row=row, column=9)).value = "Marca"
        cast(XLCell,ws_v.cell(row=row, column=10)).value = "Modelo"
        cast(XLCell,ws_v.cell(row=row, column=11)).value = "DN"
        cast(XLCell,ws_v.cell(row=row, column=12)).value = "L"
        cast(XLCell,ws_v.cell(row=row, column=13)).value = "Rel"
        cast(XLCell,ws_v.cell(row=row, column=14)).value = "Res"
        row += 1
    
    cfg = VimaToListaConfig(vima_sheet="VIMA", lista_sheet="LISTA", incremental=True)
    res = map_vima_to_lista(wb_v, wb_l, cfg)
    # Cierra explícitamente los workbooks creados en el test
    wb_v.close()
    wb_l.close()

    assert res["rows_copied"] == 3  # 1107..1109
    assert ws_l["B14"].value == "OI-1107-2025"
    assert ws_l["B15"].value == "OI-1108-2025"
    assert ws_l["B16"].value == "OI-1109-2025"