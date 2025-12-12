# tests/test_cf_apply.py
from openpyxl import Workbook
from openpyxl.formatting.rule import Rule, DifferentialStyle
from openpyxl.styles import PatternFill
from app.services.updates.update_base_by_model import _apply_cf_from_template_band

def test_cf_apply_contains_text_expands_to_band():
    # Template con una regla 'containsText' justo en la fila semilla (9)
    t_wb = Workbook()
    t_ws = t_wb.active
    t_ws.title = "ERROR FINAL"

    dxf = DifferentialStyle(fill=PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid"))
    rule = Rule(type="containsText", operator="containsText", text="NO CONFORME", dxf=dxf, stopIfTrue=False)
    t_ws.conditional_formatting.add("AU9", rule)

    # Destino vacío
    wb = Workbook()
    ws = wb.active

    # Aplicar CF al bloque A:CQ, filas 9..12
    added = _apply_cf_from_template_band(t_ws, ws, start_row=9, end_row=12,
                                         min_col=1, max_col= column_index_from_string("CQ"), seed_row=9)
    assert added >= 1

    # Debe existir sqref AU9:AU12
    cf_map = getattr(ws.conditional_formatting, "_cf_rules", {})
    sqrefs = [str(k) for k in cf_map.keys()]
    assert "AU9:AU12" in sqrefs
