import pytest
from httpx import AsyncClient, ASGITransport
from app.main import app
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from typing import cast
import gc
from contextlib import closing

@pytest.mark.asyncio
async def test_excel_inspect_openpyxl(tmp_path: Path):
    # preparar un archivo de prueba local
    fp = tmp_path / "t.xlsx"
    from openpyxl import Workbook
    wb = Workbook()
    ws: Worksheet = cast(Worksheet,wb.active)
    ws.title = "SheetA"
    ws["A1"] = "X"
    wb.save(fp)
    wb.close()  # <- evita ZipFile.__del__ en Python 3.13

    transport = ASGITransport(app=app, raise_app_exceptions=True)
    async with AsyncClient(transport=transport, base_url="http://testserver") as ac:
        resp = await ac.post("/excel/inspect", json={"file_path": str(fp)})
        assert resp.status_code == 200
        data = resp.json()
        assert data["engine"] == "openpyxl"
        assert "SheetA" in data["meta"]["sheets"]

@pytest.mark.asyncio
async def test_excel_update_openpyxl(tmp_path: Path):
    fp = tmp_path / "t_update.xlsx"
    from openpyxl import Workbook
    wb = Workbook()
    ws: Worksheet = cast(Worksheet,wb.active)
    ws.title = "Hoja1"
    ws["B3"] = "old"
    wb.save(fp)
    wb.close()

    transport = ASGITransport(app=app, raise_app_exceptions=True)
    async with AsyncClient(transport=transport, base_url="http://testserver") as ac:
        payload = {
            "file_path": str(fp),
            "edits": [{"sheet": "Hoja1", "cell": "B3", "value": "new"}],
            "save_mode": "same_password"
        }
        resp = await ac.post("/excel/update", json=payload)
        assert resp.status_code == 200
        data = resp.json()
        assert data["engine"] == "openpyxl"
        assert data["saved"] is True

    with closing(load_workbook(fp)) as wb2:   # contexto => cierre garantizado
        assert wb2["Hoja1"]["B3"].value == "new"

    # Fuerza liberación de objetos ZIP pendientes en este test
    gc.collect()