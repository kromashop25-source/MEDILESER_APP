# tests/test_vima_to_lista_api.py
from __future__ import annotations
import json
import pytest
from httpx import AsyncClient, ASGITransport
from typing import cast
from pathlib import Path
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from app.main import app

@pytest.mark.asyncio
async def test_dry_run_ok(tmp_path: Path):
    # Prepara VIMA
    vima = tmp_path / "vima.xlsm"
    wb_v = Workbook()
    ws_v: Worksheet = cast(Worksheet, wb_v.active)
    ws_v.title = "VIMA"
    ws_v["C11"] = "OI-1107-2025"
    for col, val in zip("GHIJKLMN", [1,"c","m","mo","dn","l","r","res"]):
        ws_v[f"{col}11"] = val
    wb_v.save(vima); wb_v.close()

    # Prepara LISTA con último OI 1106-2025
    lista = tmp_path / "lista.xlsx"
    wb_l = Workbook()
    ws_l: Worksheet = cast(Worksheet, wb_l.active)
    ws_l.title = "LISTA"
    ws_l["B11"] = "OI-1106-2025"
    wb_l.save(lista); wb_l.close()

    payload = {
        "vima_path": str(vima),
        "lista_path": str(lista),
        "output_path": str(tmp_path / "salida.xlsx"),
        "vima_password": None,
        "vima_sheet": "VIMA",
        "lista_sheet": "LISTA",
        "incremental": True,
        "strict_incremental": True
    }

    transport = ASGITransport(app=app, raise_app_exceptions=True)
    async with AsyncClient(transport=transport, base_url="http://test") as ac:
        r = await ac.post("/integrations/vima-to-lista/dry-run", json=payload)
        assert r.status_code == 200
        data = r.json()
        assert data["ok"] is True
        assert data["would_copy"] == 1
        assert data["first_oi_to_copy"] == "OI-1107-2025"

@pytest.mark.asyncio
async def test_incremental_estricto_400(tmp_path: Path):
    # LISTA con último B11 inválido
    lista = tmp_path / "lista.xlsx"
    wb_l = Workbook(); ws_l = cast(Worksheet, wb_l.active)
    ws_l.title = "LISTA"; ws_l["B11"] = "ULTIMO-INVALIDO"
    wb_l.save(lista); wb_l.close()

    # VIMA con fila válida
    vima = tmp_path / "vima.xlsm"
    wb_v = Workbook(); ws_v = cast(Worksheet, wb_v.active)
    ws_v.title = "VIMA"
    ws_v["C11"] = "OI-1107-2025"
    for col, val in zip("GHIJKLMN", [1,"c","m","mo","dn","l","r","res"]):
        ws_v[f"{col}11"] = val
    wb_v.save(vima); wb_v.close()

    payload = {
        "vima_path": str(vima),
        "lista_path": str(lista),
        "output_path": str(tmp_path / "salida.xlsx"),
        "vima_password": None,
        "vima_sheet": "VIMA",
        "lista_sheet": "LISTA",
        "incremental": True,
        "strict_incremental": True
    }

    transport = ASGITransport(app=app, raise_app_exceptions=True)
    async with AsyncClient(transport=transport, base_url="http://test") as ac:
        r = await ac.post("/integrations/vima-to-lista", json=payload)
        assert r.status_code == 400
        assert "estricto" in r.json()["detail"].lower()
