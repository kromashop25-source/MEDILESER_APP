from __future__ import annotations

from openpyxl import Workbook
from openpyxl.utils import column_index_from_string
from openpyxl.worksheet.worksheet import Worksheet
from typing import cast
from app.services.integrations.vima_to_lista import VimaToListaConfig, map_vima_to_lista


def _has_merge(ws: Worksheet, ref: str) -> bool:
    return any(str(r) == ref for r in ws.merged_cells.ranges)


def test_map_vima_to_lista_basic_strict() -> None:
    wb_v, wb_l = Workbook(), Workbook()
    ws_v: Worksheet = cast(Worksheet, wb_v.active)
    ws_l: Worksheet = cast(Worksheet, wb_l.active)
    ws_v.title = "VIMA"
    ws_l.title = "LISTA"

    ws_v["C11"] = "OI-1001"
    ws_v["G11"] = 1
    ws_v["H11"] = "Cert"
    ws_v["I11"] = "Marca"
    ws_v["J11"] = "Modelo"
    ws_v["K11"] = "DN"
    ws_v["L11"] = "L"
    ws_v["M11"] = "Rel"
    ws_v["N11"] = "Res"
    ws_v["C12"] = "OI-1002"  # invalid (G..N empty) -> skipped in strict mode

    cfg = VimaToListaConfig(vima_sheet="VIMA", lista_sheet="LISTA", require_all_g_to_n=True)
    res = map_vima_to_lista(wb_v, wb_l, cfg)

    assert res["rows_copied"] == 1
    ws_l = wb_l["LISTA"]
    assert ws_l["B11"].value == "OI-1001"
    assert ws_l["G11"].value == "Cert"
    assert ws_l["M11"].value == "Res"


def test_map_vima_to_lista_lax_mode() -> None:
    wb_v, wb_l = Workbook(), Workbook()
    ws_v: Worksheet = cast(Worksheet, wb_v.active)
    ws_l: Worksheet = cast(Worksheet, wb_l.active)
    ws_v.title = "VIMA"
    ws_l.title = "LISTA"

    ws_v["C11"] = "OI-2001"
    ws_v["G11"] = 2
    ws_v["H11"] = "Cert-A"
    ws_v["I11"] = "Marca-A"
    ws_v["J11"] = "Modelo-A"
    ws_v["K11"] = "DN-A"
    ws_v["L11"] = "L-A"
    ws_v["M11"] = "Rel-A"
    ws_v["N11"] = "Res-A"

    ws_v["C12"] = "OI-2002"
    ws_v["H12"] = "Cert-Only"  # only one of G..N -> valid in lax mode

    cfg = VimaToListaConfig(vima_sheet="VIMA", lista_sheet="LISTA", require_all_g_to_n=False)
    res = map_vima_to_lista(wb_v, wb_l, cfg)

    assert res["rows_copied"] == 2
    ws_l = wb_l["LISTA"]
    assert ws_l["B12"].value == "OI-2002"
    assert ws_l["G12"].value == "Cert-Only"


def test_map_vima_to_lista_merged_cells() -> None:
    wb_v, wb_l = Workbook(), Workbook()
    ws_v: Worksheet = cast(Worksheet, wb_v.active)
    ws_l: Worksheet = cast(Worksheet, wb_l.active)
    ws_v.title = "VIMA"
    ws_l.title = "LISTA"

    ws_v["C11"] = "OI-3001"
    ws_v["G11"] = 1
    ws_v["J11"] = "Modelo"
    ws_v["K11"] = "DN"
    ws_v["L11"] = "L"
    ws_v["M11"] = "Rel"
    ws_v["N11"] = "Res"

    h = column_index_from_string("H")
    i = column_index_from_string("I")
    ws_v.merge_cells(start_row=11, start_column=h, end_row=11, end_column=i)
    ws_v.cell(row=11, column=h, value="CertMerge")

    cfg = VimaToListaConfig(vima_sheet="VIMA", lista_sheet="LISTA", require_all_g_to_n=True)
    res = map_vima_to_lista(wb_v, wb_l, cfg)
    assert res["rows_copied"] == 1

    ws_l = wb_l["LISTA"]
    assert _has_merge(ws_l, "G11:H11")
    assert ws_l["G11"].value == "CertMerge"


def test_preserves_hidden_columns() -> None:
    wb_v, wb_l = Workbook(), Workbook()
    ws_v: Worksheet = cast(Worksheet, wb_v.active)
    ws_l: Worksheet = cast(Worksheet, wb_l.active)
    ws_v.title = "VIMA"
    ws_l.title = "LISTA"

    ws_l.column_dimensions["D"].hidden = True

    ws_v["C11"] = "OI-4001"
    ws_v["G11"] = 3
    ws_v["H11"] = "Cert-X"
    ws_v["I11"] = "Marca-X"
    ws_v["J11"] = "Modelo-X"
    ws_v["K11"] = "DN-X"
    ws_v["L11"] = "L-X"
    ws_v["M11"] = "Rel-X"
    ws_v["N11"] = "Res-X"

    cfg = VimaToListaConfig(vima_sheet="VIMA", lista_sheet="LISTA", require_all_g_to_n=True)
    map_vima_to_lista(wb_v, wb_l, cfg)

    ws_l = wb_l["LISTA"]
    assert ws_l.column_dimensions["D"].hidden is True
