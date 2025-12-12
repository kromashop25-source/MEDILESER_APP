from __future__ import annotations

import io
import weakref
from typing import Optional

from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook

# Mantener vivos los streams que usan workbooks en read_only=True
# (la key es débil => al destruirse el Workbook, la entrada desaparece)
_WB_STREAMS: "weakref.WeakKeyDictionary[Workbook, io.BytesIO]" = weakref.WeakKeyDictionary()


def load_workbook_maybe_encrypted(
    path: str,
    password: Optional[str] = None,
    keep_vba: bool = True,
    data_only: bool = True,
) -> Workbook:
    """
    Abre un Excel (.xlsx/.xlsm) que podría estar protegido con contraseña.
    - Si hay contraseña: se descifra en memoria (BytesIO) y se carga con openpyxl.
    - La contraseña no se persiste ni se imprime; se elimina de memoria al final.
    """
    if password:
        try:
            import msoffcrypto  # type: ignore
        except ImportError as e:
            raise RuntimeError(
                "Se requiere 'msoffcrypto-tool' para abrir libros con contraseña. "
                "Instala con: pip install msoffcrypto-tool"
            ) from e

        with open(path, "rb") as f:
            office = msoffcrypto.OfficeFile(f)
            office.load_key(password=password)
            bio = io.BytesIO()
            office.decrypt(bio)

        bio.seek(0)
        wb = load_workbook(bio, data_only=data_only, keep_vba=keep_vba)
        # Mantener vivo el stream (no cerrar 'bio' aquí)
        _WB_STREAMS[wb] = bio

        # Seguridad: eliminar trazas de la clave
        del password
        return wb

    # Sin protección
    return load_workbook(path, data_only=data_only, keep_vba=keep_vba)


def load_workbook_fast_for_scan(path: str, password: Optional[str] = None) -> Workbook:
    """
    Apertura RÁPIDA pensada para DRY-RUN / pre-escaneo:
    - read_only=True: streaming (no crea todos los objetos en memoria)
    - data_only=True: solo valores (no fórmulas)
    - keep_vba=False: no necesitamos macros para analizar
    - keep_links=False: evita resolver vínculos externos
    """
    if password:
        try:
            import msoffcrypto  # type: ignore
        except ImportError as e:
            raise RuntimeError(
                "Se requiere 'msoffcrypto-tool' para abrir libros con contraseña. "
                "Instala con: pip install msoffcrypto-tool"
            ) from e

        with open(path, "rb") as f:
            office = msoffcrypto.OfficeFile(f)
            office.load_key(password=password)
            bio = io.BytesIO()
            office.decrypt(bio)

        bio.seek(0)
        wb = load_workbook(
            bio,
            data_only=True,
            read_only=True,
            keep_vba=False,
            keep_links=False,
        )
        # Mantener vivo el BytesIO mientras viva el Workbook (sin atributos dinámicos)
        _WB_STREAMS[wb] = bio
        del password
        # ¡No cerrar 'bio' aquí! Se liberará cuando el Workbook sea recolectado o
        # si llamas explícitamente a release_workbook_stream(wb).
        return wb

    # Sin protección
    wb = load_workbook(
        path,
        data_only=True,
        read_only=True,
        keep_vba=False,
        keep_links=False,
    )
    # Cuando se abre por ruta, openpyxl mantiene el ZipFile; no es necesario anclar.
    return wb


def release_workbook_stream(wb: Optional[Workbook]) -> None:
    """
    Libera el BytesIO asociado (si lo hay).
    Útil para cerrar recursos explícitamente en pruebas o procesos masivos.
    """
    if wb is None:
        return
    try:
        bio = _WB_STREAMS.pop(wb, None)
        if bio and hasattr(bio, "close"):
            bio.close()
    except Exception:
        pass
