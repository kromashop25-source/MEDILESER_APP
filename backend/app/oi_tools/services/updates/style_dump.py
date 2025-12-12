from __future__ import annotations
from typing import Dict, Any, Optional, cast, Tuple
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.styles.colors import Color
from openpyxl.styles.borders import Side
from openpyxl.worksheet.worksheet import Worksheet
import json
import os
import argparse

# ----------------- utilidades de coerción a tipos JSON -----------------

def _to_int(x) -> Optional[int]:
    try:
        if x is None:
            return None
        return int(x)
    except Exception:
        try:
            return int(float(x))
        except Exception:
            return None

def _to_float(x) -> Optional[float]:
    try:
        if x is None:
            return None
        return float(x)
    except Exception:
        return None

def _to_bool(x) -> Optional[bool]:
    if x is None:
        return None
    return bool(x)

def _to_str(x) -> Optional[str]:
    if x is None:
        return None
    return str(x)

# ----------------- serialización segura de estilos -----------------

def _color_to_json(c: Optional[Color]) -> Any:
    if not c:
        return None
    try:
        # Preferimos RGB ARGB como texto (FFFF0000, etc.)
        rgb = getattr(c, "rgb", None)
        if rgb:
            return str(rgb)
        theme = _to_int(getattr(c, "theme", None))
        tint = _to_float(getattr(c, "tint", None))
        if theme is not None:
            return {"theme": theme, "tint": tint}
        indexed = _to_int(getattr(c, "indexed", None))
        if indexed is not None:
            return {"indexed": indexed}
    except Exception:
        pass
    return None

def _side_to_json(s: Optional[Side]) -> Any:
    if not s:
        return None
    style = _to_str(getattr(s, "style", None))
    col = _color_to_json(getattr(s, "color", None))
    return {"style": style, "color": col}

def _ensure_ws(wb, sheet_name: str) -> Worksheet:
    # Devuelve siempre una Worksheet real (no ChartSheet)
    if sheet_name and sheet_name in wb.sheetnames:
        ws_like = wb[sheet_name]
    else:
        ws_like = wb.active
    if isinstance(ws_like, Worksheet):
        return cast(Worksheet, ws_like)
    # Fallback: primera hoja de cálculo "real"
    if wb.worksheets:
        return wb.worksheets[0]
    raise ValueError("El libro no contiene hojas de cálculo válidas.")

def cell_style_to_dict(ws: Worksheet, row: int, col: int) -> Dict[str, Any]:
    c = ws.cell(row=row, column=col)
    return {
        "numberFormat": _to_str(c.number_format),

        "font": {
            "name": _to_str(getattr(c.font, "name", None)),
            "size": _to_float(getattr(c.font, "size", None)),
            "bold": _to_bool(getattr(c.font, "bold", None)),
            "italic": _to_bool(getattr(c.font, "italic", None)),
            "underline": _to_str(getattr(c.font, "underline", None)),
            "strike": _to_bool(getattr(c.font, "strike", None)),
            "vertAlign": _to_str(getattr(c.font, "vertAlign", None)),
            "color": _color_to_json(getattr(c.font, "color", None)),
        },

        "alignment": {
            "horizontal": _to_str(getattr(c.alignment, "horizontal", None)),
            "vertical": _to_str(getattr(c.alignment, "vertical", None)),
            "wrapText": _to_bool(getattr(c.alignment, "wrap_text", None)),
            "shrinkToFit": _to_bool(getattr(c.alignment, "shrink_to_fit", None)),
            "indent": _to_int(getattr(c.alignment, "indent", None)),
            "textRotation": _to_int(getattr(c.alignment, "text_rotation", None)),
        },

        "border": {
            "left":   _side_to_json(getattr(c.border, "left", None)),
            "right":  _side_to_json(getattr(c.border, "right", None)),
            "top":    _side_to_json(getattr(c.border, "top", None)),
            "bottom": _side_to_json(getattr(c.border, "bottom", None)),
            "diagonal": _side_to_json(getattr(c.border, "diagonal", None)),
            "diagonalUp": _to_bool(getattr(c.border, "diagonalUp", None)),
            "diagonalDown": _to_bool(getattr(c.border, "diagonalDown", None)),
            "outline": _to_bool(getattr(c.border, "outline", None)),
        },

        "fill": {
            "patternType": _to_str(getattr(c.fill, "patternType", None)),
            "fgColor": _color_to_json(getattr(c.fill, "fgColor", None)),
            "bgColor": _color_to_json(getattr(c.fill, "bgColor", None)),
        },

        "protection": {
            "locked": _to_bool(getattr(c.protection, "locked", None)),
            "hidden": _to_bool(getattr(c.protection, "hidden", None)),
        },
    }

def dump_styles_from_workbook(
    xlsx_path: str,
    *,
    sheet_name: str = "ERROR FINAL",
    row: int = 9,
    start_col: str = "AU",
    end_col: str = "CQ",
    out_json_path: str = "app/data/styles_au_cq.json",
) -> str:
    wb = load_workbook(xlsx_path, data_only=False)
    try:
        ws = _ensure_ws(wb, sheet_name)
        c1 = column_index_from_string(start_col)
        c2 = column_index_from_string(end_col)

        styles: Dict[str, Dict[str, Any]] = {}
        for col in range(c1, c2 + 1):
            letter = get_column_letter(col)
            styles[letter] = cell_style_to_dict(ws, row, col)

        out_dir = os.path.dirname(out_json_path)
        if out_dir:
            os.makedirs(out_dir, exist_ok=True)
        with open(out_json_path, "w", encoding="utf-8") as f:
            json.dump(styles, f, indent=2, ensure_ascii=False)
        return out_json_path
    finally:
        wb.close()

# ------ CLI: python -m app.services.updates.style_dump --xlsx "ruta.xlsx" ------
def _main():
    ap = argparse.ArgumentParser(description="Dump estilos AU:CQ de la fila 9 a JSON.")
    ap.add_argument("--xlsx", required=True, help="Ruta del Excel muestra (Base con estilos correctos).")
    ap.add_argument("--sheet", default="ERROR FINAL", help="Nombre de hoja (default: ERROR FINAL)")
    ap.add_argument("--row", type=int, default=9, help="Fila a muestrear (default: 9)")
    ap.add_argument("--range", default="AU:CQ", help="Rango columnas (default: AU:CQ)")
    ap.add_argument("--out", default="app/data/styles_au_cq.json", help="Ruta output JSON")
    args = ap.parse_args()

    start_col, end_col = (s.strip() for s in args.range.split(":"))
    out = dump_styles_from_workbook(
        args.xlsx,
        sheet_name=args.sheet,
        row=args.row,
        start_col=start_col,
        end_col=end_col,
        out_json_path=args.out,
    )
    print(f"Styles guardados en: {out}")

if __name__ == "__main__":
    _main()
