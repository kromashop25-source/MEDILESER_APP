# app/services/updates/cf_dump.py
from __future__ import annotations
import argparse, json, re, os
from typing import Any, Dict, List, Tuple, Optional, cast
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.formatting.rule import Rule
from openpyxl.styles.colors import Color

# ---------- Utilidades básicas ----------

def _to_argb(s: str) -> str:
    """f
    Normaliza un color string a ARGB de 8 dígitos aceptado por openpyxl.
    Acepta '#RRGGBB', 'RRGGBB', 'AARRGGBB'. Si recibe 6 dígitos, antepone 'FF'.
    """
    x = str(s).strip().lstrip("#").upper()
    if not re.fullmatch(r"[0-9A-F]{6,8}", x or ""):
        m = re.findall(r"[0-9A-F]", x)
        x = "".join(m)[-8:] if m else ""
    if len(x) == 6:
        return "FF" + x
    if len(x) == 8:
        return x
    return "FF000000"

def _color_to_json(c: Optional[Color]) -> Optional[str]:
    if not c:
        return None
    # Preferimos rgb si está seteado
    if c.rgb:
        return _to_argb(c.rgb)
    # Si es theme/indexed, guardamos esa referencia (no generamos ARGB fijo)
    if c.type == "theme" and c.theme is not None:
        # dejamos un marcador "theme:<n>[:tint]"
        if c.tint is not None:
            return f"theme:{c.theme}:tint:{c.tint}"
        return f"theme:{c.theme}"
    if c.type == "indexed" and c.indexed is not None:
        return f"indexed:{c.indexed}"
    return None

def _font_to_json(f) -> Dict[str, Any]:
    if f is None:
        return {}
    out: Dict[str, Any] = {}
    for k in ("name", "size", "bold", "italic", "underline", "strike", "vertAlign"):
        v = getattr(f, k, None)
        if v is not None:
            out[k] = v
    cj = _color_to_json(getattr(f, "color", None))
    if cj is not None:
        out["color"] = cj
    return out

def _fill_to_json(fl) -> Dict[str, Any]:
    if fl is None:
        return {}
    out: Dict[str, Any] = {}
    pt = getattr(fl, "patternType", None)
    if pt:
        out["patternType"] = pt
    fg = _color_to_json(getattr(fl, "fgColor", None))
    bg = _color_to_json(getattr(fl, "bgColor", None))
    if fg is not None:
        out["fgColor"] = fg
    if bg is not None:
        out["bgColor"] = bg
    return out

def _side_to_json(s) -> Optional[Dict[str, Any]]:
    if s is None:
        return None
    out: Dict[str, Any] = {}
    if getattr(s, "style", None):
        out["style"] = s.style
    col = _color_to_json(getattr(s, "color", None))
    if col is not None:
        out["color"] = col
    return out or None

def _border_to_json(b) -> Dict[str, Any]:
    if b is None:
        return {}
    out: Dict[str, Any] = {}
    for name in ("left", "right", "top", "bottom", "diagonal"):
        sj = _side_to_json(getattr(b, name, None))
        if sj:
            out[name] = sj
    for flag in ("diagonalUp", "diagonalDown", "outline"):
        v = getattr(b, flag, None)
        if v is not None:
            out[flag] = bool(v)
    return out

def _alignment_to_json(a) -> Dict[str, Any]:
    if a is None:
        return {}
    out: Dict[str, Any] = {}
    mapping = {
        "horizontal": "horizontal",
        "vertical": "vertical",
        "wrapText": "wrap_text",
        "shrinkToFit": "shrink_to_fit",
        "indent": "indent",
        "textRotation": "text_rotation",
    }
    for json_key, attr in mapping.items():
        v = getattr(a, attr, None)
        if v is not None:
            out[json_key] = v
    return out

def _dxf_to_json(dxf) -> Dict[str, Any]:
    """
    Serializa DifferentialStyle (DXF) a JSON (solo lo que suele usarse:
    font/fill/border/alignment/numberFormat/protection).
    """
    if dxf is None:
        return {}
    out: Dict[str, Any] = {}
    f = getattr(dxf, "font", None)
    if f:
        out["font"] = _font_to_json(f)
    fl = getattr(dxf, "fill", None)
    if fl:
        out["fill"] = _fill_to_json(fl)
    b = getattr(dxf, "border", None)
    if b:
        out["border"] = _border_to_json(b)
    a = getattr(dxf, "alignment", None)
    if a:
        out["alignment"] = _alignment_to_json(a)
    nf = getattr(dxf, "number_format", None) or getattr(dxf, "numFmt", None)
    if nf:
        out["numberFormat"] = nf
    pr = getattr(dxf, "protection", None)
    if pr:
        out["protection"] = {k: getattr(pr, k) for k in ("locked", "hidden") if getattr(pr, k, None) is not None}
    return out

# Adaptar referencias de fila en fórmulas a la fila semilla (solo referencias sin $)
_CELLREF = re.compile(r"(?<![A-Z0-9\$])([A-Z]{1,3})(\d+)")
def _adapt_formula_to_seed(formula: str, seed_row: int) -> str:
    def repl(m):
        col, row = m.group(1), m.group(2)
        # si el patrón es $A$1 no entra aquí; sólo referencias sin $
        return f"{col}{seed_row}"
    return _CELLREF.sub(repl, formula)

def _parse_band(band: str) -> Tuple[int, int]:
    band = band.strip().upper()
    if ":" in band:
        a, b = band.split(":", 1)
    else:
        a, b = band, band
    return column_index_from_string(a), column_index_from_string(b)

def _range_to_key(min_col: int, max_col: int) -> str:
    a = get_column_letter(min_col)
    b = get_column_letter(max_col)
    return a if a == b else f"{a}:{b}"

# ---------- DUMP principal ----------

def dump_cf(ws: Worksheet, band: str, seed_row: int = 9) -> Dict[str, List[Dict[str, Any]]]:
    """
    Extrae todas las reglas de formato condicional dentro del rango de columnas `band`
    (p. ej. 'A:CQ') y devuelve un dict JSON: { "COL" o "COL1:COLN": [ reglas... ] }
    Las fórmulas se normalizan para que hagan referencia a la `seed_row`.
    """
    band_min, band_max = _parse_band(band)
    out: Dict[str, List[Dict[str, Any]]] = {}

    # Snapshot seguro del contenedor (tipos amplios para compatibilidad de versiones)
    snapshot: List[Tuple[Any, List[Any]]] = []
    try:
        # Algunas versiones devuelven listas, otras tuplas, y los stubs de tipos no coinciden.
        for cf_obj, rules in ws.conditional_formatting:  # type: ignore[misc]
                rules_list = list(rules) if isinstance(rules, (list, tuple)) else [rules]
                snapshot.append((cf_obj, rules_list))
    except Exception:
        # fallback a atributo interno (puede no existir en futuras versiones)
        try:
            for cf_obj, rules in ws.conditional_formatting._cf_rules.items():  # type: ignore[attr-defined]
                snapshot.append((cf_obj, list(rules)))
        except Exception:
            return out  # sin reglas

    for cf_obj, rules in snapshot:
        # cf_obj.sqref es un MultiCellRange
        try:
            ranges = list(cf_obj.sqref)  # lista de CellRange
        except Exception:
            continue

        for r in ranges:
            # intersección con la banda pedida
            c1 = max(r.min_col, band_min)
            c2 = min(r.max_col, band_max)
            if c2 < c1:
                continue
            key = _range_to_key(c1, c2)

            for rule in rules:
                rj: Dict[str, Any] = {}
                # Tipos típicos: "cellIs", "expression" (openpyxl llama expression a 'formula')
                rtype = getattr(rule, "type", None)
                if rtype == "expression":
                    rj["type"] = "formula"
                else:
                    rj["type"] = rtype or "cellIs"

                op = getattr(rule, "operator", None)
                if op:
                    rj["operator"] = op

                # Fórmulas: lista[str]; adaptamos referencias a seed_row
                flist = getattr(rule, "formula", None) or []
                if flist:
                    rj["formula"] = [_adapt_formula_to_seed(str(f), seed_row) for f in flist]

                stop = getattr(rule, "stopIfTrue", None)
                if stop is not None:
                    rj["stopIfTrue"] = bool(stop)

                dxf = getattr(rule, "dxf", None)
                dj = _dxf_to_json(dxf)
                if dj:
                    rj["dxf"] = dj

                out.setdefault(key, []).append(rj)

    return out

# ---------- CLI ----------

def _main():
    p = argparse.ArgumentParser(description="Dump de Formatos Condicionales a JSON")
    p.add_argument("--xlsx", required=True, help="Ruta del XLSX base")
    p.add_argument("--sheet", default="ERROR FINAL", help="Nombre de hoja (por defecto 'ERROR FINAL')")
    p.add_argument("--band", default="A:CQ", help="Rango de columnas a inspeccionar (p.ej. 'A:CQ')")
    p.add_argument("--seed-row", type=int, default=9, help="Fila semilla para normalizar fórmulas (por defecto 9)")
    p.add_argument("--out", required=True, help="Ruta de salida JSON")
    args = p.parse_args()

    wb = load_workbook(args.xlsx, data_only=False)
    ws_like = wb[args.sheet] if args.sheet in wb.sheetnames else wb.active
    # Asegurar Worksheet (si es Chartsheet u otro tipo, coger la primera Worksheet disponible)
    from openpyxl.worksheet.worksheet import Worksheet
    if isinstance(ws_like, Worksheet):
        ws = ws_like
    else:
        ws = wb.worksheets[0]

    data = dump_cf(ws, band=args.band, seed_row=args.seed_row)

    # Acomodar claves a letras mayúsculas por seguridad
    norm = {}
    for k, v in data.items():
        norm[k.upper()] = v

    os.makedirs(os.path.dirname(args.out) or ".", exist_ok=True)
    with open(args.out, "w", encoding="utf-8") as f:
        json.dump(norm, f, ensure_ascii=False, indent=2)

    print(f"OK -> {args.out}")

if __name__ == "__main__":
    _main()
