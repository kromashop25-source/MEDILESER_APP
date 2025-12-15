from __future__ import annotations
from dataclasses import dataclass
from typing import Dict, List, Tuple, Optional, TypedDict, cast, Any, Set
from functools import lru_cache
from io import BytesIO
import re
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter, range_boundaries
from pathlib import Path
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import Cell, MergedCell
from openpyxl.styles.colors import Color
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, Protection
from openpyxl.formatting.rule import CellIsRule, Rule, FormulaRule
from openpyxl.formatting.formatting import ConditionalFormatting
from openpyxl.styles.differential import DifferentialStyle
from copy import copy as _shallow_copy, deepcopy
from zipfile import BadZipFile, ZipFile
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.worksheet.cell_range import CellRange, MultiCellRange
import json
import os
import posixpath
import xml.etree.ElementTree as ET

# ================= Excepciones controladas =================

class PasswordRequiredError(Exception):
    """El archivo está cifrado y no se proporcionó contraseña."""
    pass

class WrongPasswordError(Exception):
    """La contraseña proporcionada es incorrecta."""
    pass

# ================= Modelos de datos =================

@dataclass
class PasswordBundle:
    default: Optional[str]
    per_file: Dict[str, str]

@dataclass
class UpdateOptions:
    oi_pattern: str = r"^OI-(\d+)-(\d{4})$"
    oi_start_row: int = 9
    base_start_row: int = 9
    target_sheet_name: Optional[str] = None  # usa activa si None
    # Plantilla obligatoria que contiene estilos y CF. Hoja debe tener misma maqueta.
    cf_template_path: Optional[str] = None

@dataclass
class SheetDrawingBundle:
    sheet_path: str
    sheet_rel_path: str
    sheet_rels_xml: Optional[bytes]
    rel_id: str
    drawing_target: str
    drawing_path: str
    drawing_xml: bytes
    drawing_rels_path: Optional[str]
    drawing_rels_xml: Optional[bytes]
    media: Dict[str, bytes]
    drawing_element: bytes
    content_types: Dict[str, str]

class OIFile(TypedDict):
    name: str
    bytes: bytes

# ================= Utilidades de hoja =================

def _pick_worksheet(wb, preferred: Optional[str] = "ERROR FINAL") -> Worksheet:
    """
    Devuelve la Worksheet a usar:
    - Si 'preferred' existe => esa.
    - Si no, la primera Worksheet del libro.
    """
    if preferred and preferred in wb.sheetnames:
        ws = wb[preferred]
        if isinstance(ws, Worksheet):
            return ws  # type: ignore[return-value]
    if wb.worksheets:
        return wb.worksheets[0]
    raise ValueError("El libro no contiene hojas de cálculo.")

# Nombre de la hoja por defecto (ambos documentos)

DEFAULT_SHEET_NAME = "ERROR FINAL"
# Límite de filas de Excel (xlsx)
EXCEL_MAX_ROWS = 1_048_576
def _pick_ws_for_read(wb, preferred_name: Optional[str] = DEFAULT_SHEET_NAME) -> Worksheet:
    """
    Selecciona la hoja de lecura para OIs:
    - Si existe 'preferred_name' => esa hoja.
    - Si no, la primera hoja 'Worksheet'.
    - Último recurso: activa (cast).
    """
    try:
        if preferred_name and preferred_name in wb.sheetnames:
            ws = wb[preferred_name]
            if isinstance(ws, Worksheet):
                return cast(Worksheet, ws)
    except Exception:
        pass
    for ws in getattr(wb, "worksheets", []):
        if isinstance(ws, Worksheet):
            return cast(Worksheet, ws)
    return cast(Worksheet, wb.active)

def _pick_ws_for_base(wb, target_name: Optional[str]) -> Worksheet:
    """
    Selecciona la hoja destino en la Base:
    - Si 'target_name' existe => esa hoja.
    - Si no, intenta DEFAULT_SHEET_NAME.
    - Si no, la primera hoja 'Worksheet'.
    """

    try:
        if target_name and target_name in wb.sheetnames:
            ws = wb[target_name]
            if isinstance(ws, Worksheet):
                return cast(Worksheet, ws)
    except Exception:
        pass
    try:
        if DEFAULT_SHEET_NAME in wb.sheetnames:
            ws = wb[DEFAULT_SHEET_NAME]
            if isinstance(ws, Worksheet):
                return cast(Worksheet, ws)
    except Exception:
        pass
    for ws in getattr(wb, "worksheets", []):
        if isinstance(ws, Worksheet):
            return cast(Worksheet, ws)
    return cast(Worksheet, wb.active)

# ================= Helpers de apertura =================

def _try_open_workbook(file_bytes: bytes, password: Optional[str], *, data_only: bool = True):
    """
    Intenta abrir un Excel. Si está cifrado:
      - sin password -> PasswordRequiredError
      - password errónea -> WrongPasswordError
    """

    bio = BytesIO(file_bytes)

    try:
        return load_workbook(bio, data_only=data_only)
    except (BadZipFile, InvalidFileException):
        # Puede ser cifrado. Si no hay password, pedirla.
        if not password:
            raise PasswordRequiredError("Archivo cifrado: se requiere contraseña.")
        # Intentar descifrado con msoffcrypto
        try:
            import msoffcrypto
        except Exception:
            raise PasswordRequiredError("No se puede abrir archivo cifrado sin msoffcrypto.")
        src = BytesIO(file_bytes)
        office = msoffcrypto.OfficeFile(src)
        try:
            office.load_key(password=password)
        except Exception as e:
            raise WrongPasswordError("Contraseña incorrecta.") from e
        out = BytesIO()

        try:
            office.decrypt(out)

        except Exception as e:

            raise WrongPasswordError("Contraseña incorrecta.") from e

        out.seek(0)

        try:

            return load_workbook(out, data_only=data_only)

        finally:

            out.close()


# ================= Helpers de imagenes (drawings) =================

def _normalize_rel_path(base_path: str, target: str) -> str:
    target = target.replace("\\", "/")
    if target.startswith("/"):
        return target.lstrip("/")
    base_dir = posixpath.dirname(base_path)
    return posixpath.normpath(posixpath.join(base_dir, target))


def _extract_sheet_drawing_bundle(xlsx_bytes: bytes, sheet_name: str) -> Optional[SheetDrawingBundle]:
    """Extrae el dibujo (logo) de una hoja especifica desde un XLSX."""
    if not xlsx_bytes:
        return None
    try:
        with ZipFile(BytesIO(xlsx_bytes)) as zip_src:
            workbook_xml = zip_src.read("xl/workbook.xml")
            workbook_rels_xml = zip_src.read("xl/_rels/workbook.xml.rels")
            workbook_tree = ET.fromstring(workbook_xml)
            ns_main = {
                "main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
                "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
            }
            rel_id: Optional[str] = None
            for sheet in workbook_tree.findall("main:sheets/main:sheet", ns_main):
                if sheet.get("name") == sheet_name:
                    rel_id = sheet.get("{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id")
                    if rel_id:
                        break
            if not rel_id:
                return None
            workbook_rels_tree = ET.fromstring(workbook_rels_xml)
            rel_ns = "http://schemas.openxmlformats.org/package/2006/relationships"
            sheet_target: Optional[str] = None
            for rel in workbook_rels_tree.findall(f"{{{rel_ns}}}Relationship"):
                if rel.get("Id") == rel_id:
                    sheet_target = rel.get("Target")
                    break
            if not sheet_target:
                return None
            sheet_path = sheet_target
            if not sheet_path.startswith("xl/"):
                sheet_path = posixpath.normpath(posixpath.join("xl", sheet_path))
            sheet_xml = zip_src.read(sheet_path)
            sheet_tree = ET.fromstring(sheet_xml)
            ns_sheet = {"ws": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
            drawing_el = sheet_tree.find("ws:drawing", ns_sheet)
            if drawing_el is None:
                return None
            drawing_rel_id = drawing_el.get("{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id")
            if not drawing_rel_id:
                return None
            sheet_rel_path = posixpath.join(
                posixpath.dirname(sheet_path),
                "_rels",
                posixpath.basename(sheet_path) + ".rels",
            )
            sheet_rels_xml: Optional[bytes] = None
            sheet_rels_tree = None
            drawing_target: Optional[str] = None
            if sheet_rel_path in zip_src.namelist():
                sheet_rels_xml = zip_src.read(sheet_rel_path)
                try:
                    sheet_rels_tree = ET.fromstring(sheet_rels_xml)
                except ET.ParseError:
                    sheet_rels_tree = None
            if sheet_rels_tree is not None:
                for rel in sheet_rels_tree.findall(f"{{{rel_ns}}}Relationship"):
                    if rel.get("Id") == drawing_rel_id:
                        drawing_target = rel.get("Target")
                        break
            if not drawing_target:
                return None
            drawing_path = _normalize_rel_path(sheet_path, drawing_target)
            if not drawing_path.startswith("xl/"):
                drawing_path = posixpath.normpath(posixpath.join("xl", drawing_path))
            drawing_xml = zip_src.read(drawing_path)
            drawing_rels_path = posixpath.join(
                posixpath.dirname(drawing_path),
                "_rels",
                posixpath.basename(drawing_path) + ".rels",
            )
            drawing_rels_xml: Optional[bytes] = None
            drawing_rels_tree = None
            if drawing_rels_path in zip_src.namelist():
                drawing_rels_xml = zip_src.read(drawing_rels_path)
                try:
                    drawing_rels_tree = ET.fromstring(drawing_rels_xml)
                except ET.ParseError:
                    drawing_rels_tree = None
            media: Dict[str, bytes] = {}
            if drawing_rels_tree is not None:
                for rel in drawing_rels_tree.findall(f"{{{rel_ns}}}Relationship"):
                    if rel.get("Type") != "http://schemas.openxmlformats.org/officeDocument/2006/relationships/image":
                        continue
                    target = rel.get("Target")
                    if not target:
                        continue
                    media_path = _normalize_rel_path(drawing_path, target)
                    if not media_path.startswith("xl/"):
                        media_path = posixpath.normpath(posixpath.join("xl", media_path))
                    if media_path in zip_src.namelist():
                        media[media_path] = zip_src.read(media_path)
            content_types: Dict[str, str] = {}
            try:
                content_tree = ET.fromstring(zip_src.read("[Content_Types].xml"))
            except Exception:
                content_tree = None
            parts = ["/" + drawing_path] + ["/" + name for name in media.keys()]
            if content_tree is not None:
                ns_ct = {"ct": "http://schemas.openxmlformats.org/package/2006/content-types"}
                for override in content_tree.findall("ct:Override", ns_ct):
                    part_name = override.get("PartName")
                    if part_name in parts:
                        content_types[part_name] = override.get("ContentType") or ""
            if "/" + drawing_path not in content_types:
                content_types["/" + drawing_path] = "application/vnd.openxmlformats-officedocument.drawing+xml"
            for media_path in media.keys():
                part = "/" + media_path
                if part not in content_types:
                    ext = posixpath.splitext(media_path)[1].lower()
                    if ext == ".png":
                        content_types[part] = "image/png"
                    elif ext in (".jpg", ".jpeg"):
                        content_types[part] = "image/jpeg"
                    elif ext == ".gif":
                        content_types[part] = "image/gif"
                    elif ext == ".bmp":
                        content_types[part] = "image/bmp"
                    else:
                        content_types[part] = "application/octet-stream"
            drawing_element_bytes = ET.tostring(drawing_el, encoding="utf-8")
            return SheetDrawingBundle(
                sheet_path=sheet_path,
                sheet_rel_path=sheet_rel_path,
                sheet_rels_xml=sheet_rels_xml,
                rel_id=drawing_rel_id,
                drawing_target=drawing_target,
                drawing_path=drawing_path,
                drawing_xml=drawing_xml,
                drawing_rels_path=drawing_rels_path if drawing_rels_xml is not None else None,
                drawing_rels_xml=drawing_rels_xml,
                media=media,
                drawing_element=drawing_element_bytes,
                content_types=content_types,
            )
    except Exception:
        return None


def _inject_sheet_drawing(xlsx_bytes: bytes, bundle: Optional[SheetDrawingBundle]) -> bytes:
    """Inserta el dibujo extraido en un XLSX generado."""
    if not bundle:
        return xlsx_bytes
    input_stream = BytesIO(xlsx_bytes)
    output_stream = BytesIO()
    rel_ns = "http://schemas.openxmlformats.org/package/2006/relationships"
    ct_ns = "http://schemas.openxmlformats.org/package/2006/content-types"
    sheet_ns = {"ws": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
    skip_names: Set[str] = {bundle.drawing_path}
    if bundle.drawing_rels_path:
        skip_names.add(bundle.drawing_rels_path)
    skip_names.update(bundle.media.keys())
    with ZipFile(input_stream, "r") as zin, ZipFile(output_stream, "w") as zout:
        seen: Set[str] = set()
        for info in zin.infolist():
            name = info.filename
            seen.add(name)
            if name in skip_names:
                continue
            data = zin.read(name)
            if name == bundle.sheet_path:
                try:
                    tree = ET.fromstring(data)
                except ET.ParseError:
                    zout.writestr(info, data)
                    continue
                if tree.find("ws:drawing", sheet_ns) is None:
                    drawing_el = ET.fromstring(bundle.drawing_element)
                    tree.append(drawing_el)
                updated = ET.tostring(tree, encoding="utf-8", xml_declaration=True)
                zout.writestr(info, updated)
                continue
            if name == bundle.sheet_rel_path:
                try:
                    rels_tree = ET.fromstring(data)
                except ET.ParseError:
                    if bundle.sheet_rels_xml is not None:
                        zout.writestr(info, bundle.sheet_rels_xml)
                    else:
                        zout.writestr(info, data)
                    continue
                exists = any(rel.get("Id") == bundle.rel_id for rel in rels_tree.findall(f"{{{rel_ns}}}Relationship"))
                if not exists:
                    rel = ET.Element(f"{{{rel_ns}}}Relationship", {
                        "Id": bundle.rel_id,
                        "Type": "http://schemas.openxmlformats.org/officeDocument/2006/relationships/drawing",
                        "Target": bundle.drawing_target,
                    })
                    rels_tree.append(rel)
                updated = ET.tostring(rels_tree, encoding="utf-8", xml_declaration=True)
                zout.writestr(info, updated)
                continue
            if name == "[Content_Types].xml":
                try:
                    ct_tree = ET.fromstring(data)
                except ET.ParseError:
                    zout.writestr(info, data)
                    continue
                existing_parts = {ov.get("PartName") for ov in ct_tree.findall(f"{{{ct_ns}}}Override")}
                for part, ctype in bundle.content_types.items():
                    if part not in existing_parts:
                        ET.SubElement(ct_tree, f"{{{ct_ns}}}Override", {"PartName": part, "ContentType": ctype})
                updated = ET.tostring(ct_tree, encoding="utf-8", xml_declaration=True)
                zout.writestr(info, updated)
                continue
            zout.writestr(info, data)
        if bundle.sheet_rel_path not in seen:
            if bundle.sheet_rels_xml is not None:
                zout.writestr(bundle.sheet_rel_path, bundle.sheet_rels_xml)
            else:
                root = ET.Element("Relationships", xmlns=rel_ns)
                ET.SubElement(root, "Relationship", {
                    "Id": bundle.rel_id,
                    "Type": "http://schemas.openxmlformats.org/officeDocument/2006/relationships/drawing",
                    "Target": bundle.drawing_target,
                })
                zout.writestr(bundle.sheet_rel_path, ET.tostring(root, encoding="utf-8", xml_declaration=True))
        if "[Content_Types].xml" not in seen and bundle.content_types:
            ct_root = ET.Element("Types", xmlns=ct_ns)
            for part, ctype in bundle.content_types.items():
                ET.SubElement(ct_root, "Override", {"PartName": part, "ContentType": ctype})
            zout.writestr("[Content_Types].xml", ET.tostring(ct_root, encoding="utf-8", xml_declaration=True))
        zout.writestr(bundle.drawing_path, bundle.drawing_xml)
        if bundle.drawing_rels_path and bundle.drawing_rels_xml is not None:
            zout.writestr(bundle.drawing_rels_path, bundle.drawing_rels_xml)
        for media_name, media_bytes in bundle.media.items():
            zout.writestr(media_name, media_bytes)
    return output_stream.getvalue()
# ================= Lógica de conteo (A:AT) =================

_COL_A  = column_index_from_string("A")

_COL_AT = column_index_from_string("AT")

def _row_is_empty_in_range(ws: Worksheet, row: int, min_col: int = _COL_A, max_col: int = _COL_AT) -> bool:

    for c in range(min_col, max_col + 1):

        if ws.cell(row=row, column=c).value not in (None, ""):

            return False

    return True

def _count_rows_to_copy(ws: Worksheet, start_row: int) -> int:

    """Cuenta filas desde start_row en A:AT hasta encontrar una fila completamente vacía."""

    r = start_row

    count = 0

    while True:

        if _row_is_empty_in_range(ws, r, _COL_A, _COL_AT):

            break

        count += 1

        r += 1

    return count

def _first_free_row(ws: Worksheet, start_row: int) -> int:

    """Primera fila 'libre' (sin valores) desde start_row; revisa A..J para acelerar."""

    r = start_row

    while True:

        any_val = False

        for c in range(1, 11):

            if ws.cell(row=r, column=c).value not in (None, ""):

                any_val = True

                break

        if not any_val:

            return r

        r += 1

# ================= Orden por patrón OI =================

_RE_CACHE: Dict[str, re.Pattern[str]] = {}

def _re_cached(pat: str) -> re.Pattern[str]:

    r = _RE_CACHE.get(pat)

    if r is None:

        r = re.compile(pat)

        _RE_CACHE[pat] = r

    return r

def _parse_oi_key(name: str, pattern: str) -> Tuple[int, int]:

    """Devuelve (num, year) a partir de 'OI-####-YYYY' (según pattern)."""

    base = name.rsplit("/", 1)[-1].rsplit("\\", 1)[-1]

    base = base.rsplit(".", 1)[0]

    m = _re_cached(pattern).match(base)

    if not m:

        raise ValueError(f"Nombre no cumple patrón: {name}")

    num = int(m.group(1))

    year = int(m.group(2))

    return (num, year)

def _sorted_ois(oi_list: List[OIFile], pattern: str) -> List[OIFile]:

    return sorted(oi_list, key=lambda x: _parse_oi_key(x["name"], pattern))

########################

#  Fórmulas desde TXT  #

########################

def _parse_formulas_base(text: str) -> Dict[str, str]:

    out: Dict[str, str] = {}

    for raw in text.splitlines():

        line = raw.strip()

        if not line or line.startswith("#"):

            continue

        if ":=" not in line:

            continue

        col, rhs = line.split(":=", 1)

        col = col.strip().upper()

        rhs = rhs.strip()

        if rhs.startswith("="):

            rhs = rhs[1:]

        if col:

            out[col] = rhs

    return out

def _spanish_to_english_formula(body: str) -> str:

    s = body

    # Funciones ES?EN

    rep_pairs = [

        (r"\bSI\(", "IF("),

        (r"\bY\(", "AND("),

        (r"\bO\(", "OR("),

        (r"\bEXTRAE\(", "MID("),

        (r"\bLARGO\(", "LEN("),

        (r"\bSIGNO\(", "SIGN("),

        (r"\bCONCATENAR\(", "CONCATENATE("),

        (r"\bABS\(", "ABS("),

    ]

    for pat, repl in rep_pairs:

        s = re.sub(pat, repl, s)

    # Booleanos

    s = re.sub(r"\bFALSO\b", "FALSE", s)

    s = re.sub(r"\bVERDADERO\b", "TRUE", s)

    # 0,9 ? 0.9

    s = re.sub(r"(\d+),(\d+)", r"\1.\2", s)

    # ; ? ,

    s = s.replace(";", ",")

    return s

def _adapt_row_refs_from_9(body: str, row: int) -> str:

    # Cambia ...9 por ...<row>, sin tocar absolutas tipo $AH$6

    pattern = r"(?<![A-Z0-9\$])([A-Z]{1,3})9(?!\d)"

    return re.sub(pattern, lambda m: f"{m.group(1)}{row}", body)

def _generate_formula(form_map: Dict[str, str], col_letter: str, row: int) -> Optional[str]:

    base = form_map.get(col_letter.upper())

    if not base:

        return None

    b = _adapt_row_refs_from_9(base, row)

    b = _spanish_to_english_formula(b)

    if not b.startswith("="):

        b = "=" + b

    return b

def _load_formulas_map_or_raise() -> Dict[str, str]:

    """

    Lee app/data/templates/oi_tools/FORMULAS_BASE.txt y valida:

    - Formato COL:=FÓRMULA (semilla en fila 9)

    - Columnas en AU..CQ excluyendo AX

    """

    # Este archivo vive en: app/oi_tools/services/updates/update_base_by_model.py
    # parents[3] apunta a .../backend/app
    fp = Path(__file__).resolve().parents[3] / "data" / "templates" / "oi_tools" / "FORMULAS_BASE.txt"

    if not fp.exists():

        raise ValueError("No se encontró app/data/templates/oi_tools/FORMULAS_BASE.txt.")

    try:

        txt = fp.read_text(encoding="utf-8")

    except Exception as e:

        raise ValueError(f"No se pudo leer FORMULAS_BASE.txt: {e}")

    m = _parse_formulas_base(txt)

    if not m:

        raise ValueError("FORMULAS_BASE.txt vacío o sin líneas válidas (esperado: 'COL:=FÓRMULA').")

    # Validar rango columnas AU..CQ excepto AX

    valid_cols = []

    from_col = column_index_from_string("AU")

    to_col   = column_index_from_string("CQ")

    for dc in range(from_col, to_col + 1):

        letter = get_column_letter(dc)

        if letter != "AX":

            valid_cols.append(letter)

    invalid = [c for c in m.keys() if c not in valid_cols]

    if invalid:

        raise ValueError(f"Columnas no permitidas en FORMULAS_BASE.txt: {', '.join(invalid)}.")

    return m

# ================= API del servicio =================

def probe_open_all_ois(oi_list: List[OIFile], passwords: PasswordBundle, pattern: str) -> None:

    """

    Pre-chequeo: intenta abrir cada OI para disparar 401/403 antes del stream NDJSON.

    """

    for item in oi_list:

        name = item["name"]

        pwd = passwords.per_file.get(name) or passwords.default

        try:

            wb = _try_open_workbook(item["bytes"], pwd, data_only=True)

            wb.close()

        except WrongPasswordError:

            raise WrongPasswordError(f"Contraseña incorrecta para: {name}")

        except PasswordRequiredError:

            raise PasswordRequiredError(f"Se requiere contraseña para: {name}")

def dry_run_update_base_from_ois(base_bytes: bytes,

                                 oi_list: List[OIFile],

                                 passwords: PasswordBundle,

                                 opt: UpdateOptions):

    """

    Generador NDJSON:

      - stage: 'received' => lista de OIs (ya ordenados)

      - stage: 'analysis' => por cada OI, filas contadas (A:AT)

      - stage: 'summary'  => total y first_free_row

    """

    # Base para calcular la primera fila libre

    wb_base = load_workbook(BytesIO(base_bytes), data_only=True)

    # Base: usar target_sheet_name si llega; si no, "ERROR FINAL"; si no, primera hoja

    ws_base = _pick_worksheet(wb_base, opt.target_sheet_name or "ERROR FINAL")

    # Ordenar OIs

    sorted_oi = _sorted_ois(oi_list, opt.oi_pattern)

    yield {"stage": "received", "ois": [x["name"] for x in sorted_oi]}

    total_rows = 0

    for item in sorted_oi:

        name = item["name"]

        pwd = passwords.per_file.get(name) or passwords.default

        wb = _try_open_workbook(item["bytes"], pwd, data_only=True)

        # OI: preferir "ERROR FINAL"; si no, primera hoja

        ws = _pick_worksheet(wb, "ERROR FINAL")

        rows = _count_rows_to_copy(ws, opt.oi_start_row)

        total_rows += rows

        yield {"stage": "analysis", "oi": name, "rows": rows}

        wb.close()

    ff = _first_free_row(ws_base, opt.base_start_row)

    wb_base.close()

    # higiene: limpiar secretos

    passwords.default = None

    passwords.per_file.clear()

    yield {

        "stage": "summary",

        "would_copy_rows": total_rows,

        "first_free_row_in_base": ff,

    }

# ---- Helpers numéricos -------------------------------------------------------

def _coerce_int(value: Any) -> Optional[int]:

    """

    Convierte valores comunes de Excel a int:

    - int -> int

    - float -> int (trunca)

    - str  -> intenta parsear (admite coma o punto decimal)

    - otros/None -> None

    """

    if value is None:

        return None

    if isinstance(value, bool):

        return int(value)

    if isinstance(value, int):

        return value

    if isinstance(value, float):

        try:

            return int(value)

        except Exception:

            return None

    if isinstance(value, str):

        s = value.strip().replace(",", ".")

        try:

            return int(float(s))

        except Exception:

            # fallback: extraer el primer entero que aparezca

            m = re.search(r"-?\d+", s)

            return int(m.group(0)) if m else None

    return None


# ================= Plantilla Excel (fuente única de estilos/CF) ==============

# Ruta: app/data/templates/oi_tools/BASE_TEMPLATE.xlsx

TEMPLATE_XLSX_NAME = "BASE_TEMPLATE.xlsx"

# Este archivo vive en: app/oi_tools/services/updates/update_base_by_model.py
# parents[3] apunta a .../backend/app
APP_DIR = Path(__file__).resolve().parents[3]
TEMPLATE_PATH = APP_DIR / "data" / "templates" / "oi_tools" / TEMPLATE_XLSX_NAME

# Tipos admitidos por openpyxl.Rule (para apaciguar al type checker y validar entrada)

VALID_RULE_TYPES: Set[str] = {

    "cellIs", "expression", "containsText", "notContainsText",

    "beginsWith", "endsWith", "timePeriod",

    "colorScale", "dataBar", "iconSet", "top10",

    "aboveAverage", "duplicateValues", "uniqueValues",

    "blanks", "noBlanks", "errors", "noErrors"

}

VALID_OPERATORS: Set[str] = {

    "lessThan", "lessThanOrEqual", "equal", "notEqual",

    "greaterThan", "greaterThanOrEqual", "between", "notBetween"

}

def _clone_style(obj):

    """

    Devuelve una copia 'hashtable' del estilo para evitar StyleProxy cross-workbook.

    Intenta .copy() (APIs de openpyxl) y cae a copy() superficial si hace falta.

    """

    if obj is None:

        return None

    # Muchos estilos de openpyxl implementan copy.()

    try:

        return obj.copy() # type: ignore[attr-defined]

    except Exception:

        try:

            return _shallow_copy(obj)

        except Exception:

            # Útilo recurso: devolver el mismo objeto (no debería ocurrir)

            return obj

        

def copy_cell_style(src, dst) -> None:

    """

    Clona estilos de src -> dst (sin copiar valor).

    Incluye: font, fill, border, alignment, number_format, protection.

    """

    dst.font = _clone_style(src.font)

    dst.fill = _clone_style(src.fill)

    dst.border = _clone_style(src.border)

    dst.alignment = _clone_style(src.alignment)

    # number_format es string (seguro asignarlo directo)

    dst.number_format = src.number_format

    dst.protection = _clone_style(src.protection)

def _clone_dxf(dxf: Optional[DifferentialStyle]) -> Optional[DifferentialStyle]:

    """

    Clona un DifferentialStyle (para reglas de formato condicional).

    """

    if dxf is None:

        return None

    return DifferentialStyle(

        font=_clone_style(getattr(dxf, "font", None)),

        fill=_clone_style(getattr(dxf, "fill", None)),

        border=_clone_style(getattr(dxf, "border", None)),

        numFmt=getattr(dxf, "numFmt", None),

        protection=_clone_style(getattr(dxf, "protection", None)),

        alignment=_clone_style(getattr(dxf, "alignment", None)),

    )

def _open_template_ws(target_sheet_name: Optional[str]) -> Optional[Worksheet]:

    """

    Abre app/data/templates/oi_tools/BASE_TEMPLATE.xlsx y devuelve la hoja a usar (o None si no existe).

    """

    try:

        if not TEMPLATE_PATH.exists():

            return None

        wb = load_workbook(TEMPLATE_PATH, data_only=False)

        ws = _pick_worksheet(wb, target_sheet_name or DEFAULT_SHEET_NAME)

        # guardamos el workbook en el propio objeto worksheet para evitar GC

        setattr(ws, "_owner_wb", wb)  # evita que se cierre prematuramente

        return ws

    except Exception:

        return None

def _copy_styles_from_template_row(ws_tpl: Worksheet,

                                   tpl_row: int,

                                   ws_dst: Worksheet,

                                   dst_row: int,

                                   col_start: int,

                                   col_end: int) -> None:

    """Copia estilos 1:1 desde la fila `tpl_row` de la plantilla al `dst_row` en [col_start:col_end]."""

    for dc in range(col_start, col_end + 1):

        src = ws_tpl.cell(row=tpl_row, column=dc)

        dst = _ensure_writable_cell(ws_dst, dst_row, dc)

        copy_cell_style(src, dst)

    

def _ensure_writable_cell(ws: Worksheet, row: int, col: int) -> Cell:

    """

    Si (row, col) apunta a una MergedCell, des-mergea el rango que la cubre

    y devuelve la celda normal del mismo (row, col). Así evitamos errores

    de asignación a MergedCell.

    """

    cell = ws.cell(row=row, column=col)

    if isinstance(cell, MergedCell):

        # Buscar el merge que la incluye y deshacerlo

        for rng in list(ws.merged_cells.ranges):

            if rng.min_row <= row <= rng.max_row and rng.min_col <= col <= rng.max_col:

                ws.unmerge_cells(str(rng))

                break

        cell = ws.cell(row=row, column=col)

    # Garantizamos que aquí siempre devolvemos una Cell "normal"

    return cell  # type: ignore[return-value]

def _merge_ranges_from_source_to_dest(ws_src: Worksheet, ws_dst: Worksheet,

                                      src_top: int, src_bottom: int,

                                      dst_top: int, min_col: int, max_col: int) -> None:

    """

    Replica merges que caen dentro del bloque [src_top:src_bottom] y columnas [min_col:max_col],

    desplazándolos para que comiencen en dst_top.

    """

    # Tomamos una snapshot para no iterar mientras modificamos

    merged = list(getattr(ws_src, "merged_cells").ranges)

    for m in merged:

        r1, c1, r2, c2 = m.min_row, m.min_col, m.max_row, m.max_col

        if r2 < src_top or r1 > src_bottom:  # fuera del bloque vertical

            continue

        if c2 < min_col or c1 > max_col:     # fuera de A:AT

            continue

        # intersección con nuestro bloque

        r1i = max(r1, src_top); r2i = min(r2, src_bottom)

        c1i = max(c1, min_col); c2i = min(c2, max_col)

        # desplazar al destino

        row_offset = dst_top - src_top

        new_r1 = r1i + row_offset

        new_r2 = r2i + row_offset

        try:

            ws_dst.merge_cells(start_row=new_r1, start_column=c1i, end_row=new_r2, end_column=c2i)

        except Exception:

            # Ignora solapes/duplicados inocuos

            pass

def _range_bounds(range_like: Any) -> Tuple[int, int, int, int]:

    """

    Devuelve (min_row, max_row, min_col, max_col) para:

    - CellRange / NamedRange-like (con atributos min_* / max_*)

    - MultiCellRange -> no aquí (se expande afuera)

    - str con notación A1 (usa range_boundaries)

    Lanza AttributeError si no puede normalizar.

    """

    attrs = ("min_row", "max_row", "min_col", "max_col")

    if all(hasattr(range_like, a) for a in attrs):

        min_row = getattr(range_like, "min_row", None)

        max_row = getattr(range_like, "max_row", None)

        min_col = getattr(range_like, "min_col", None)

        max_col = getattr(range_like, "max_col", None)

        if not all(isinstance(v, int) for v in (min_row, max_row, min_col, max_col)):

            raise AttributeError("Objeto de rango sin límites compatibles.")

        return cast(Tuple[int, int, int, int], (min_row, max_row, min_col, max_col))

    if isinstance(range_like, str):

        min_c, min_r, max_c, max_r = range_boundaries(range_like)

        return cast(Tuple[int, int, int, int], (min_r, max_r, min_c, max_c))

    raise AttributeError("Objeto de rango sin límites compatibles.")

def _cf_key_to_ranges(range_key: Any) -> List[CellRange]:
    if isinstance(range_key, ConditionalFormatting):
        sqref = getattr(range_key, "sqref", None)
        if isinstance(sqref, MultiCellRange):
            return list(sqref.ranges)
        if isinstance(sqref, CellRange):
            return [sqref]
        if isinstance(sqref, str):
            try:
                return list(MultiCellRange(sqref).ranges)
            except Exception:
                try:
                    return [CellRange(sqref)]
                except Exception:
                    return []
        if sqref is not None:
            sqref_str = str(sqref)
            try:
                return list(MultiCellRange(sqref_str).ranges)
            except Exception:
                try:
                    return [CellRange(sqref_str)]
                except Exception:
                    return []
        return []
    if isinstance(range_key, MultiCellRange):
        return list(range_key.ranges)
    if isinstance(range_key, CellRange):
        return [range_key]
    if isinstance(range_key, str):
        try:
            return list(MultiCellRange(range_key).ranges)
        except Exception:
            try:
                return [CellRange(range_key)]
            except Exception:
                return []
    return []

def _clear_cf_in_band(ws: Worksheet, start_row: int, end_row: int, min_col: int, max_col: int) -> None:
    cf = ws.conditional_formatting
    try:
        items = list(getattr(cf, "_cf_rules", {}).items())
    except Exception:
        ws.conditional_formatting = type(cf)()
        return

    from collections import OrderedDict
    new_map = OrderedDict()
    for range_key, rules in items:
        try:
            subranges = _cf_key_to_ranges(range_key)
        except Exception:
            subranges = []
        if subranges:
            kept: List[CellRange] = []
            for sr in subranges:
                try:
                    r1, r2, c1, c2 = _range_bounds(sr)
                except Exception:
                    continue
                if r2 < start_row or r1 > end_row or c2 < min_col or c1 > max_col:
                    kept.append(sr)
            if kept:
                coords = [sr.coord for sr in kept]
                sqref = MultiCellRange(coords) if len(coords) > 1 else coords[0]
                if isinstance(range_key, ConditionalFormatting):
                    new_key = ConditionalFormatting(sqref=sqref, pivot=getattr(range_key, "pivot", None))
                else:
                    new_key = sqref
                new_map[new_key] = rules
            continue
        new_map[range_key] = rules
    try:
        cf._cf_rules = new_map  # type: ignore[attr-defined]
    except Exception:
        ws.conditional_formatting = type(cf)()

def _adapt_row_refs_from_seed(body: str, seed_row: int, new_start: int) -> str:
    """Adapta formulas de CF relativas a la fila semilla."""
    pattern = re.compile(rf"(?<![A-Za-z0-9_])(\$?[A-Z]{1,3})(\$?){seed_row}(?!\d)")

    def _replace(match: re.Match[str]) -> str:
        col_token = match.group(1)
        row_abs = match.group(2)
        if row_abs:
            return match.group(0)
        return f"{col_token}{new_start}"

    return pattern.sub(_replace, body)

def _clone_cf_rule(rule: Rule, row_mapping: Dict[int, int]) -> Rule:
    rtype_raw = getattr(rule, "type", None)
    rtype_str = str(rtype_raw) if rtype_raw is not None else "expression"
    if rtype_str not in VALID_RULE_TYPES:
        rtype_str = "expression"
    new_rule = Rule(type=cast(Any, rtype_str))

    bool_attrs = ("stopIfTrue", "percent", "bottom", "equalAverage")
    for attr in bool_attrs:
        value = getattr(rule, attr, None)
        if value is not None:
            setattr(new_rule, attr, bool(value))

    op_raw = getattr(rule, "operator", None)
    if op_raw is not None:
        op_str = str(op_raw)
        if op_str in VALID_OPERATORS:
            new_rule.operator = cast(Any, op_str)

    text_value = getattr(rule, "text", None)
    if text_value is not None:
        new_rule.text = str(text_value)

    time_period = getattr(rule, "timePeriod", None)
    if time_period is not None:
        new_rule.timePeriod = cast(Any, str(time_period))

    rank_value = getattr(rule, "rank", None)
    if rank_value is not None:
        try:
            new_rule.rank = int(rank_value)
        except Exception:
            pass

    formulas_in = list(getattr(rule, "formula", []) or [])
    if formulas_in:
        adapted: List[str] = []
        for formula in formulas_in:
            body = str(formula)
            for src_row, dst_row in row_mapping.items():
                if src_row != dst_row:
                    body = _adapt_row_refs_from_seed(body, src_row, dst_row)
            adapted.append(body)
        new_rule.formula = adapted

    dxf = getattr(rule, "dxf", None)
    if dxf is not None:
        new_rule.dxf = _clone_dxf(dxf)

    for attr in ("dataBar", "colorScale", "iconSet"):
        payload = getattr(rule, attr, None)
        if payload is not None:
            setattr(new_rule, attr, deepcopy(payload))

    priority = getattr(rule, "priority", None)
    if isinstance(priority, int):
        try:
            new_rule.priority = priority
        except Exception:
            pass

    return new_rule

def _apply_cf_from_template_band(
    ws_tpl: Worksheet,
    ws_dst: Worksheet,
    start_row: int,
    end_row: int,
    min_col: int,
    max_col: int,
    seed_row: int = 9,
) -> int:
    """Clona reglas de formato condicional desde la plantilla al bloque destino."""
    if end_row < start_row:
        return 0

    added = 0
    cf_tpl = ws_tpl.conditional_formatting
    for range_key, rules in list(getattr(cf_tpl, "_cf_rules", {}).items()):
        try:
            subranges = _cf_key_to_ranges(range_key)
        except Exception:
            subranges = []
        if not subranges:
            continue

        for sr in subranges:
            try:
                src_r1, src_r2, src_c1, src_c2 = _range_bounds(sr)
            except Exception:
                continue

            if src_r2 < seed_row:
                continue

            c1 = max(src_c1, min_col)
            c2 = min(src_c2, max_col)
            if c2 < c1:
                continue

            row_mapping: Dict[int, int] = {seed_row: start_row}
            if src_r2 >= seed_row:
                start_src_row = max(src_r1, seed_row)
                for src_row in range(start_src_row, src_r2 + 1):
                    dst_row = start_row + (src_row - seed_row)
                    row_mapping[src_row] = dst_row

            for dc in range(c1, c2 + 1):
                col_letter = get_column_letter(dc)
                target = f"{col_letter}{start_row}:{col_letter}{end_row}"
                for rule in list(rules):
                    cloned = _clone_cf_rule(rule, row_mapping)
                    ws_dst.conditional_formatting.add(target, cloned)
                    added += 1

    return added

def execute_update_base_from_ois(base_bytes: bytes,

                                 oi_list: List[OIFile],

                                 passwords: PasswordBundle,

                                 opt: UpdateOptions,

                                 replicate_merges: bool = True,

                                 replicate_row_heights: bool = False,

                                 replicate_col_widths: bool = False,

                                 progress_cb = None,

                                 enforce_excel_limit: bool = True):

    """

    Devuelve (xlsx_bytes, resumen_dict)

    - Copia A:AT desde cada OI a la Base conservando estilos de la OI (valores).

    - Columna AX = nombre del OI (sin extensión).

    - Columnas AU:CQ = fórmulas copiadas desde fila plantilla (fila anterior a la primera libre),

      excepto en AX (no sobrescribir el nombre).

    """

    out = BytesIO()

    wb_dst = load_workbook(BytesIO(base_bytes), data_only=False)

    # Base: usar target_sheet_name si llega; si no, "ERROR FINAL"; si no, primera hoja

    ws_dst = _pick_worksheet(wb_dst, opt.target_sheet_name or "ERROR FINAL")

    drawing_bundle = _extract_sheet_drawing_bundle(base_bytes, ws_dst.title)

   # # PLANTILLA para estilos y CF (única fuente). Usa cf_template_path o BASE_TEMPLATE.xlsx por defecto.

    tpl_path = opt.cf_template_path or TEMPLATE_PATH

    if not os.path.exists(tpl_path):

        wb_dst.close()

        raise ValueError(f"No se encontró la plantilla de estilos/CF en: {tpl_path}")

    try:

        tpl_wb = load_workbook(tpl_path, data_only=False)

        tpl_ws = _pick_worksheet(tpl_wb, opt.target_sheet_name or DEFAULT_SHEET_NAME)

    except Exception as e:

        wb_dst.close()

        raise ValueError(f"No se pudo abrir la plantilla CF '{tpl_path}': {e}")

    # Fila libre inicial; la 'fila semilla' de la plantilla será base_start_row (p.ej. 9).

    first_free = _first_free_row(ws_dst, opt.base_start_row)

    template_row = first_free - 1 if first_free > opt.base_start_row else None  # (solo por compatibilidad)


    seed_row_tpl = opt.base_start_row

 

    

    col_A  = _COL_A

    col_AT = _COL_AT

    col_AU = column_index_from_string("AU")

    col_CQ = column_index_from_string("CQ")

    col_AX = column_index_from_string("AX")

    total_rows = 0

    blocks = []

    widths_applied = False

    sorted_oi = _sorted_ois(oi_list, opt.oi_pattern)

    # Usaremos la hoja ya abierta como `tpl_ws`

    ws_tpl = tpl_ws

    tpl_seed_row = opt.base_start_row

    # Cargar fórmulas desde TXT (única verdad)

    form_map = _load_formulas_map_or_raise()

    for item in sorted_oi:

        name = item["name"]

        pwd = passwords.per_file.get(name) or passwords.default

        # Abrimos OI en data_only=True => pegamos valores en A:AT

        wb_src = _try_open_workbook(item["bytes"], pwd, data_only=True)

        # OI: preferir "ERROR FINAL"; si no, primera hoja

        ws_src = _pick_worksheet(wb_src, "ERROR FINAL")

        # Prgoreso

        if progress_cb:

            progress_cb({"stage": "processing", "message": f"Abriendo {name}"})

        rows = _count_rows_to_copy(ws_src, opt.oi_start_row)

        if rows <= 0:

            wb_src.close()

            continue

        # Límite de filas: evitar overflow antes de escribir

        if enforce_excel_limit and (first_free + rows -1) > EXCEL_MAX_ROWS:

            wb_src.close()

            raise ValueError(

                f"Se excede el máximo de filas de Excel. Fila libre={first_free}, "

                f"filas a pegar={rows}, máximo={EXCEL_MAX_ROWS}."

            )

        dst_top = first_free

        # 0) Anchos de columna (A:AT) si se pidió y aun no se aplicó

        if replicate_col_widths and not widths_applied:

            for dc in range(col_A, col_AT + 1):

                letter = get_column_letter(dc)

                cd = ws_src.column_dimensions.get(letter)

                w = cd.width if cd is not None else None

                if w is not None:

                    ws_dst.column_dimensions[letter].width = w

            widths_applied = True

        # 2) Columna A + Copiar A:AT (valores + estilos desde OI)      

        for r_off in range(rows):

            sr = opt.oi_start_row + r_off

            dr = dst_top + r_off

            # Estilo base para todas las columnas A:CQ desde la plantilla (si hay)

            if ws_tpl is not None:

                for dc_all in range(1, column_index_from_string("CQ") + 1):

                    dstc = _ensure_writable_cell(ws_dst, dr, dc_all)

                    srct = ws_tpl.cell(row=tpl_seed_row, column=dc_all)

                    copy_cell_style(srct, dstc)

            # --- Columna A:  estilo de plantilla si existe ---

            c_idx = _ensure_writable_cell(ws_dst, dr, 1)

            # Estilo SIEMPRE desde plantilla (col A)

            copy_cell_style(tpl_ws.cell(row=seed_row_tpl, column=1), c_idx)

                            
            # Alturas de fila si se pidió

            if replicate_row_heights:

                rd = ws_src.row_dimensions.get(sr)

                h = rd.height if rd is not None else None

                if h is not None:

                    ws_dst.row_dimensions[dr].height = h

            for dc in range(col_A, col_AT + 1):

                sc = dc

                c_src = ws_src.cell(row=sr, column=sc)

                c_dst = _ensure_writable_cell(ws_dst, dr, dc)

                c_dst.value = c_src.value  # solo VALOR; estilo ya viene de la plantilla (al inicio del renglón)

        # 3) Replicar merges (opcional) dentro del bloque A:AT

        if replicate_merges:

            _merge_ranges_from_source_to_dest(

                ws_src, ws_dst,

                src_top=opt.oi_start_row, src_bottom=opt.oi_start_row + rows - 1,

                dst_top=dst_top, min_col=col_A, max_col=col_AT

            )

        # 4) AX = nombre del OI (sin extensión), con estilo de plantilla si existe

        name_wo_ext = name.rsplit(".", 1)[0]

        for r_off in range(rows):

            dr = dst_top + r_off

            c_dst = _ensure_writable_cell(ws_dst, dr, col_AX)

            c_dst.value = name_wo_ext  # type: ignore[assignment]

            # Estilo SIEMPRE desde plantilla (col AX)

            copy_cell_style(tpl_ws.cell(row=seed_row_tpl, column=col_AX), c_dst)

        # 5) AU:CQ: fórmulas y ESTILO desde PLANTILLA (si existe).

        for r_off in range(rows):

            dr = dst_top + r_off

            for dc in range(col_AU, col_CQ + 1):

                if dc == col_AX:

                    continue  # no sobrescribir el nombre en AX

                dst_cell = _ensure_writable_cell(ws_dst, dr, dc)

                # 5.a) Fórmula (usa tu mecanismo de fórmulas base si lo tienes;

                #      si no, conservamos el fallback de copiar la celda plantilla)

                try:

                    col_letter = get_column_letter(dc)

                    # Si tienes tus helpers de fórmulas (p. ej. _generate_formula / form_map), úsalo aquí:

                    f = _generate_formula(form_map, col_letter, dr)

                except Exception:

                    f = None

                if f:

                    dst_cell.value = f  # Excel recalcula al abrir

                elif template_row:

                    # (raro que falte)  copia de base si existiera algo

                    src_tpl = ws_dst.cell(row=template_row, column=dc)

                    dst_cell.value = src_tpl.value  # type: ignore[assignment]

                # else: no hay fórmula (mantener None)

                # 5.b) Estilo: copiar SIEMPRE desde la plantilla si está disponible;

                # en su defecto, copia estilo de la fila plantilla de la propia Base.

                if ws_tpl is not None:

                    src_tpl = ws_tpl.cell(row=tpl_seed_row, column=dc)

                    copy_cell_style(src_tpl, dst_cell)

                elif template_row:

                    src_tpl = ws_dst.cell(row=template_row, column=dc)

                    copy_cell_style(src_tpl, dst_cell)

                

        blocks.append({"oi": name, "rows": rows, "dst_first_row": dst_top, "dst_last_row": dst_top + rows - 1})

        total_rows += rows

        first_free += rows  # avanza puntero

        wb_src.close()

        if progress_cb:

            progress_cb({"stage": "processing", "message": f"Pegado {name}", "rows": rows})

    # --> Aplicar formatos condicionales de la PLANTILLA sobre TODA el área pegada (A:CQ)

    if ws_tpl is not None and total_rows > 0 and blocks:

        min_col = column_index_from_string("A")

        max_col = column_index_from_string("CQ")

        start = blocks[0]["dst_first_row"]

        end   = blocks[-1]["dst_last_row"]

        # borrar CF previas en esa banda y aplicar desde plantilla

        _clear_cf_in_band(ws_dst, start, end, min_col, max_col)

        _apply_cf_from_template_band(ws_tpl, ws_dst, start, end, min_col, max_col, seed_row=tpl_seed_row)

    # higiene: limpiar secretos

    passwords.default = None

    passwords.per_file.clear()

    wb_dst.save(out)

    wb_dst.close()

    try:

        tpl_wb.close()

    except Exception:

        pass

    out_bytes = out.getvalue()

    if drawing_bundle:

        out_bytes = _inject_sheet_drawing(out_bytes, drawing_bundle)

    out.close()

    result = {

        "rows_copied": total_rows,

        "blocks": blocks,

        "first_write_row": blocks[0]["dst_first_row"] if blocks else first_free,

        "last_write_row": blocks[-1]["dst_last_row"] if blocks else first_free - 1,

    }

    return out_bytes, result
