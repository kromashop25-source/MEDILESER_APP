"""""
Servicio de E/S (Entrada/Salida de Excel con dos backends (motorores/capas de procesamiento)):
- OpenPyXL: Para libros de trabajo (workbooks) sin contraseña(.xlxs/xlsm) cuando no se establece una contraseña
de apertura.
- Exxcel COM (pywin32): Para libros de trabajo protegidos con contraseña, conservando el formato,
las macro y las contraseñas.

Seguridad:
- Las contraseñas se reciben como SecreStr (Pydantic) en la capa de APi y se pasan aqui como 
un string plano (str).
- Nunca registres (log) las contraseñas.Elimina las referencias lo antes posible después de su uso.
"""

from __future__ import annotations
from typing import List, Optional, Dict, Any, Literal
from pathlib import Path
import os
import gc

from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook
from openpyxl.utils.cell import coordinate_to_tuple
import pandas as pd
import tempfile
import os
import re

WIN32_AVAILABLE = False
try:
    import win32com.client as win32 # type: ignore
    WIN32_AVAILABLE = True
except Exception:
    WIN32_AVAILABLE = False

class ExcelError(RuntimeError):
    """Error general de Excel."""

def close_workbook_safe(wb: Optional[Workbook]) -> None:
    """
    Cierra un Workbook asegurando el cierre del ZIP auxiliar cuando keep_vba=True.
    Tolera referencias None.
    """
    if wb is None:
        return
    vba_archive = getattr(wb, "vba_archive", None)
    try:
        wb.close()
    except Exception:
        pass
    if vba_archive is not None:
        try:
            vba_archive.close()
        except Exception:
            pass
        finally:
            try:
                wb.vba_archive = None  # type: ignore[attr-defined]
            except Exception:
                pass

def _ensure_path(path_str: str) -> Path:
    """Asegura que el path es un objeto Path."""
    p = Path(path_str)
    if not p.exists():
        raise FileNotFoundError(f"Archivo no encontrado: {p}")
    if not p.is_file():
        raise ExcelError(f"No es un archivo válido: {p}")
    return p

# --- Validación interna de referencia A1 (una sola celda) ---
_CELL_RE = re.compile(r"^[A-Z]{1,3}[1-9][0-9]{0,5}$")
def _a1_to_row_col(cell_ref: str) -> tuple[int, int]:
    """
    Convierte 'B3' -> (3,2). Rechaza rangos tipo 'A1:B2' o referencias inválidas
    """
    if not isinstance(cell_ref, str) or not _CELL_RE.fullmatch(cell_ref.strip().upper()):
        raise ExcelError(f"Celda inválida: {cell_ref}. Esperando formato A1 (ej. 'B3')")
    # coordinate_to_tuple levanta ValueError si es inválido
    row, col = coordinate_to_tuple(cell_ref.strip().upper())
    return row, col

# ------------------------------
# OpenPyXL backend (sin contrseña)
# ------------------------------
def inspect_openpyxl(path: str) -> Dict[str, Any]:
    p = _ensure_path(path)
    wb: Optional[Workbook] = None
    ws = None
    try:
        wb = load_workbook(filename=str(p), data_only=True, keep_vba=True)
        sheets = wb.sheetnames
        active_sheet_name = wb.active.title if wb.active else sheets[0]
        meta = {"sheets": sheets, "active": active_sheet_name}
        sample = []
        ws = wb[active_sheet_name]
        for r in range(1, min(ws.max_row, 5) + 1):
            sample.append(ws.cell(row=r, column=1).value)
        return {"engine": "openpyxl", "path": str(p), "meta": meta, "sample_colA_top5": sample}
    except Exception as ex:
        raise ExcelError(f"Error al inspeccionar  OpenPyXL: {type(ex).__name__}: {ex}") from ex
    finally:
        ws = None
        close_workbook_safe(wb)
        wb = None
        gc.collect()

def update_openpyxl(path: str, edits: List[Dict[str, Any]]) -> Dict[str, Any]:
    p = _ensure_path(path)
    wb: Optional[Workbook] = None
    try:
        wb = load_workbook(filename=str(p), data_only=False, keep_vba=True)
        applied = []
        for e in edits:
            sheet = e["sheet"]
            cell = e["cell"]
            value = e["value"]
            if sheet not in wb.sheetnames:
                raise ExcelError(f"Hoja no ecnontrada: {sheet}")
            ws = wb[sheet]
            # Usar coordenadas seguras para evitar ambiguedad tuple/range
            row, col = _a1_to_row_col(cell)
            ws.cell(row=row, column=col).value = value
            applied.append({"sheet": sheet, "cell": cell, "value": value})
        wb.save(str(p))
        return {"engine": "openpyxl", "saved": True, "applied": applied}
    except Exception as ex:
        raise ExcelError(f"Error actualizando (openpyxl) {type(ex).__name__}: {ex}") from ex
    finally:
        close_workbook_safe(wb)
        wb = None
        gc.collect()

def _detect_fileformat_by_ext(path: Path) -> int:
    """
    Returns Excel FileFormat number for SaveAs.
    51 = xlOpenXMLWorkbook (.xlsx)
    52 = xlOpenXMLWorkbookMacroEnabled (.xlsm)
    """
    ext = path.suffix.lower()
    if ext == ".xlsm":
        return 52
    # default .xlsx
    return 51

def inspect_excelcom(path: str, password: str) -> Dict[str, Any]:
    if not WIN32_AVAILABLE:
        raise ExcelError("Excel COM (pywin32) no disponible. Instala Excel y pywin32.")
    p = _ensure_path(path)
    app = None
    wb = None
    try:
        app = win32.Dispatch("Excel.Application")
        app.Visible = False
        app.DisplayAlerts = False
        wb = app.Workbooks.Open(Filename=str(p), ReadOnly=True, Password=password)
        sheets = [sh.Name for sh in wb.Worksheets]
        active = wb.ActiveSheet.Name
        sample = []
        try:
            used = wb.ActiveSheet.Range("A1:A5").Value
            if used is None:
                sample = []
            else:
                vals = used if isinstance(used, (list, tuple) ) else ((used,))
                sample = [row[0] if isinstance(row, (list, tuple)) else row for row in vals]
        except Exception:
            sample = []
        return {"engine": "excelcom", "path": str(p), "meta": {"sheets": sheets, "active": active}, "sample_colA_top5": sample}
    except Exception as ex:
        raise ExcelError(f"Error inspeccionando Excel COM: {type(ex).__name__}: {ex}") from ex
    finally:
        try:
            if wb is not None:
                wb.Close(SaveChanges=False)
        finally:
            if app is not None:
                app.Quit()
        del wb
        del app
        del password
        gc.collect()

def update_excelcom(
    path: str,
    password: str,
    edits: List[Dict[str, Any]],
    save_mode: Literal["same_password", "no_password", "new_password"] = "same_password",
    new_password: Optional[str] = None,
) -> Dict[str, Any]:
    if not WIN32_AVAILABLE:
        raise ExcelError("Excel COM (pywin32) no disponible. Instala Excel y pywin32.")
    p = _ensure_path(path)
    app = None
    wb = None
    try:
        app = win32.Dispatch("Excel.Application")
        app.Visible = False
        app.DisplayAlerts = False
        wb = app.Workbooks.Open(Filename=str(p), ReadOnly=False, Password=password)

        applied = []
        for e in edits:
            sheet = e["sheet"]
            cell = e["cell"]
            value = e["value"]
            try:
                ws = wb.Worksheets(sheet)
            except Exception:
                raise ExcelError(f"Hoja no encontrada: {sheet}")
            ws.Range(cell).Value = value
            applied.append({"sheet": sheet, "cell": cell, "value": value})

        if save_mode == "same_password":
            wb.Save()  # keeps the same open password and formatting/macros
        else:
            ff = _detect_fileformat_by_ext(p)
            if save_mode == "no_password":
                wb.SaveAs(Filename=str(p), FileFormat=ff, Password="")
            elif save_mode == "new_password":
                if not new_password:
                    raise ExcelError("Debe proporcionar 'new_password' cuando save_mode='new_password'.")
                wb.SaveAs(Filename=str(p), FileFormat=ff, Password=new_password)
            else:
                raise ExcelError(f"save_mode no soportado: {save_mode}")

        return {"engine": "excelcom", "saved": True, "applied": applied, "save_mode": save_mode}
    except Exception as ex:
        raise ExcelError(f"Error actualizando (Excel COM): {type(ex).__name__}: {ex}") from ex
    finally:
        try:
            if wb is not None:
                wb.Close(SaveChanges=False)
        finally:
            if app is not None:
                app.Quit()
        # wipe secrets and COM objects
        del wb
        del app
        del password
        del new_password
        gc.collect()

# ------------------------------
# Public API used by the router
# ------------------------------

def inspect_excel(path: str, password: Optional[str]) -> Dict[str, Any]:
    """
    Decide backend:
    - If password is None or "", use openpyxl (fast, no Excel dependency).
    - If password provided, use Excel COM to open with password (preserve format).
    """
    if password:
        return inspect_excelcom(path, password)
    return inspect_openpyxl(path)


def update_excel(
    path: str,
    edits: List[Dict[str, Any]],
    password: Optional[str],
    save_mode: Literal["same_password", "no_password", "new_password"] = "same_password",
    new_password: Optional[str] = None,
) -> Dict[str, Any]:
    """
    Apply edits and save. Password rules:
    - If no password was required to open: uses openpyxl (cannot add encryption).
      -> save_mode ignored (always saved sin contraseña).
    - If password required: uses Excel COM and save_mode/new_password are honored.
    """
    if password:
        return update_excelcom(path, password, edits, save_mode, new_password)
    # openpyxl path: cannot set a new open password
    if save_mode in ("new_password",):
        raise ExcelError("No se puede establecer nueva contraseña con openpyxl. Use Excel COM (proporcione password de apertura).")
    return update_openpyxl(path, edits)

def read_as_dataframe(
    path: str,
    sheet: Optional[str] = None, # si es None, usaremos la primera hoja (índice 0)
    header_row: int = 1,   
    password: Optional[str] = None,
) -> pd.DataFrame:
    """
    Lee una hoja de Excel a DataFrame.
    - Sin password: usa pandas.read_excel + openpyxl.
    - Con password: si COM disponible, abre y hace SaveAs temporal sin contraseña, lee con pandas, y borra el temporal.
    """
    p = _ensure_path(path)
    hdr = header_row - 1 if header_row > 0 else 0  # pandas header es 0-index
    sheet_name = sheet if sheet is not None else 0  # ✅ garantiza DataFrame
    if not password:
        return pd.read_excel(str(p), sheet_name=sheet_name, header=hdr, engine="openpyxl")

    if not WIN32_AVAILABLE:
        raise ExcelError("No se puede leer con contraseña: Excel COM no disponible en este entorno.")

    app = None
    wb = None
    tmp_path = None
    try:
        app = win32.Dispatch("Excel.Application")
        app.Visible = False
        app.DisplayAlerts = False
        wb = app.Workbooks.Open(Filename=str(p), ReadOnly=True, Password=password)

        # Guardar a un temporal sin password (xlsx)
        tmp_fd, tmp_name = tempfile.mkstemp(suffix=".xlsx")
        os.close(tmp_fd)
        tmp_path = Path(tmp_name)
        ff = 51  # .xlsx
        wb.SaveAs(Filename=str(tmp_path), FileFormat=ff, Password="")

        # Leer con pandas
        df = pd.read_excel(str(tmp_path), sheet_name=sheet_name, header=hdr, engine="openpyxl")
        return df
    except Exception as ex:
        raise ExcelError(f"Error leyendo DataFrame: {type(ex).__name__}: {ex}") from ex
    finally:
        try:
            if wb is not None:
                wb.Close(SaveChanges=False)
        finally:
            if app is not None:
                app.Quit()
        if tmp_path and tmp_path.exists():
            try:
                tmp_path.unlink()
            except Exception:
                pass
        # limpiar secretos
        del wb
        del app
        if password:
            del password
        gc.collect()


    


