# app/services/integrations/vima_to_lista.py
from __future__ import annotations

from dataclasses import dataclass
from typing import Optional, Tuple, List, cast, Pattern, Callable, Dict, Any
import re
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import column_index_from_string
from openpyxl.cell.cell import MergedCell, Cell as XLCell

from ..excel_styles import copy_style


@dataclass
class VimaToListaConfig:
    vima_sheet: Optional[str] = None
    lista_sheet: Optional[str] = None
    vima_start_row: int = 11
    vima_cols_range: Tuple[str, str] = ("B", "N")  # B..N en VIMA
    lista_start_col: str = "A"                     # A..M en LISTA
    lista_start_row: int = 11
    require_all_g_to_n: bool = True
    stop_blank_oi_streak: int = 50
    mode_replace: bool = True  # limpia valores en destino antes de pegar
    incremental: bool = False #NUEVO: agregar sol OIs posteriores al último en LISTA
    oi_pattern: str = r"^OI[-\u2010\u2011\u2012\u2013\u2014\u2212\uFE63\uFF0D](\d+)[-\u2010\u2011\u2012\u2013\u2014\u2212\uFE63\uFF0D](\d{4})$"
    strict_incremental : bool = False # si True, falla si el último OI en LISTA
    replicate_merges: bool = True  # NUEVO: permite desactivar la réplica de merges


def _row_is_valid(ws_vima: Worksheet, row: int, cfg: VimaToListaConfig) -> bool:
    """
    Regla: debe tener Nro OI en C; y G..N con datos (si require_all_g_to_n=True).
    """
    c = column_index_from_string("C")
    g = column_index_from_string("G")
    n = column_index_from_string("N")

    nro_oi = ws_vima.cell(row=row, column=c).value
    if nro_oi in (None, "", 0):
        return False

    def _value_with_merge(col: int):
        cell = ws_vima.cell(row=row, column=col)
        if isinstance(cell, MergedCell):
            for rng in ws_vima.merged_cells.ranges:
                if rng.min_row <= row <= rng.max_row and rng.min_col <= col <= rng.max_col:
                    top_left = ws_vima.cell(row=rng.min_row, column=rng.min_col)
                    return top_left.value
        return cell.value

    vals = [_value_with_merge(col) for col in range(g, n + 1)]
    if cfg.require_all_g_to_n:
        return all(v not in (None, "") for v in vals)
    return any(v not in (None, "") for v in vals)


def _clear_dest(ws_lista: Worksheet, cfg: VimaToListaConfig) -> None:
    """Limpia SOLO valores del bloque destino (no toca estilos/ocultos)."""
    dst_c0 = column_index_from_string(cfg.lista_start_col)
    src_c0 = column_index_from_string(cfg.vima_cols_range[0])
    src_c1 = column_index_from_string(cfg.vima_cols_range[1])
    width = (src_c1 - src_c0) + 1

    for r in range(cfg.lista_start_row, ws_lista.max_row + 1):
        for ci in range(dst_c0, dst_c0 + width):
            ws_lista.cell(row=r, column=ci).value = None

# Guiones bonitos y similares que Excel/Word suelen colar
_HYPHENS = "\u2010\u2011\u2012\u2013\u2014\u2212\uFE63\uFF0D"  
_ZERO_WIDTH = "\u200B\u200C\u200D\u2060"
_NBSP = "\u00A0"

def _normalize_oi_value(value: object) -> str:
    """
    - Convierte a str
    - recorta
    - reemplaza guiones unicode por '-'
    - elimina NBSP y zero-width
    """
    s = "" if value is None else str(value)
    # homogeniza guiones
    for h in _HYPHENS:
        s = s.replace(h, "-")
    # elimina espacios no separables y zero-width
    for ch in (_NBSP + _ZERO_WIDTH):
        s = s.replace(ch, "")
    return s.strip()

def _parse_oi(value: object, pat: Pattern[str]) -> Optional[Tuple[int, int]]:
    """
    Devuelve (year, number) si el valor cumple el patrÃ³n; si no, None.
    Ordenamos por (year, number) para avanzar cronolÃ³gicamente.
    """
    s = _normalize_oi_value(value)
    if not s:
        return None
    m = pat.match(s)
    if not m:
        return None
    num = int(m.group(1))  # correlativo
    year = int(m.group(2))
    return (year, num)


def _last_oi_in_lista(ws_lista: Worksheet, cfg: VimaToListaConfig) -> Tuple[Optional[Tuple[int,int]], int]:
    """
    Devuelve el último OI (clave=(year, num)) encontrado en la COLUMNA B (cfg.lista_start_col)
    que cumpla el patrón cfg.oi_pattern, buscando desde abajo hacia arriba.
    Si no hay ninguno que cumpla, retorna (None, última_fila_con_algún_valor_en_B_o_la_última_fila_de_datos).
    """
    pat = re.compile(cfg.oi_pattern, re.IGNORECASE)
    b_col = column_index_from_string("B")  # normalmente B

    # 1) Buscar el último que MACHEE el patrón en B (sin validar más columnas)
    for r in range(ws_lista.max_row, cfg.lista_start_row - 1, -1):
        b_val = ws_lista.cell(row=r, column=b_col).value
        if b_val is None or str(b_val).strip() == "":
            continue
        key = _parse_oi(b_val, pat)   # devuelve (year, num) o None
        if key:
            return key, r

    # 2) Si no hay OI válido, al menos devuelvo la última fila con algo en B (o la fila previa al bloque)
    last_row = cfg.lista_start_row - 1
    for r in range(ws_lista.max_row, cfg.lista_start_row - 1, -1):
        if ws_lista.cell(row=r, column=b_col).value not in (None, ""):
            last_row = r
            break

    return None, last_row




def map_vima_to_lista(
    wb_vima: Workbook,
    wb_lista: Workbook,
    cfg: Optional[VimaToListaConfig] = None,
    progress_cb: Optional[Callable[[Dict[str, Any]], None]] = None,
) -> dict:
    """
    Copia valores (no fórmulas) de VIMA B..N -> LISTA A..M, clonando estilos y
    replicando combinaciones (merged cells). No altera columnas ocultas ni anchos.
    """
    cfg = cfg or VimaToListaConfig()

    ws_vima = cast(Worksheet, wb_vima[cfg.vima_sheet] if cfg.vima_sheet else wb_vima.active)
    ws_lista = cast(Worksheet, wb_lista[cfg.lista_sheet] if cfg.lista_sheet else wb_lista.active)

    # Si es incremental, NUNCA limpiamos destino
    if cfg.mode_replace and not cfg.incremental:
        _clear_dest(ws_lista, cfg)

    src_c0 = column_index_from_string(cfg.vima_cols_range[0])
    src_c1 = column_index_from_string(cfg.vima_cols_range[1])
    dst_c0 = column_index_from_string(cfg.lista_start_col)

    rows_copied = 0
    rows_skipped = 0
    blank_oi_streak = 0
    total_rows = max(1, ws_vima.max_row - cfg.vima_start_row + 1)
    processed_rows = 0
    last_row_reported = cfg.vima_start_row

    def emit_progress(current_row: int, stage: str) -> None:
        nonlocal last_row_reported
        if not progress_cb:
            return
        last_row_reported = current_row
        percent = max(0.0, min(100.0, (processed_rows / total_rows) * 100.0))
        progress_cb({
            "stage": stage,
            "percent": round(percent, 2),
            "rows_processed": processed_rows,
            "rows_total": total_rows,
            "rows_copied": rows_copied,
            "rows_skipped": rows_skipped,
            "current_row": current_row,
        })

    # === utilidades locales (sin tipos exóticos para Pylance) ===
    def find_merge(ws: Worksheet, row: int, col: int):
        for rng in ws.merged_cells.ranges:
            if rng.min_row <= row <= rng.max_row and rng.min_col <= col <= rng.max_col:
                return rng
        return None

    def unmerge_overlaps(ws: Worksheet, r0: int, c0: int, r1: int, c1: int) -> None:
        to_unmerge: List[object] = []
        for rng in ws.merged_cells.ranges:
            if not (rng.max_row < r0 or rng.min_row > r1 or rng.max_col < c0 or rng.min_col > c1):
                to_unmerge.append(rng)
        for rng in to_unmerge:
            ws.unmerge_cells(str(rng))

    # Configuración incremental: obtener último OI en LISTA y confirmar fila base de escritura
    pat = re.compile(cfg.oi_pattern, re.IGNORECASE)
    last_key: Optional[Tuple[int, int]] = None
    dst_base_row = cfg.lista_start_row
    if cfg.incremental:
        last_key, last_row_idx = _last_oi_in_lista(ws_lista, cfg)
        if cfg.strict_incremental and last_key is None:
            # No podemos confiar en el punto de partida, abortamos
            raise ValueError(
                "Incremental estricto: el último valor en LISTA no coincide con el patrón de OI "
                f"({cfg.oi_pattern}). Corrige el dato o desactiva 'strict_incremental'."
            )
        dst_base_row = max(last_row_idx + 1, cfg.lista_start_row)

    emit_progress(cfg.vima_start_row, "init")

    # === bucle principal ===
    for offset, r in enumerate(range(cfg.vima_start_row, ws_vima.max_row + 1), start=1):
        processed_rows = offset

        if not _row_is_valid(ws_vima, r, cfg):
            rows_skipped += 1
            emit_progress(r, "skipped")
            if ws_vima.cell(row=r, column=column_index_from_string("C")).value in (None, ""):
                blank_oi_streak += 1
                if blank_oi_streak >= cfg.stop_blank_oi_streak:
                    emit_progress(r, "stopped_blank")
                    break
            continue

        blank_oi_streak = 0
        # Si incremental: saltar las OIs <= último en LISTA
        if cfg.incremental:
            oi_val = ws_vima.cell(row=r, column=column_index_from_string("C")).value
            oi_key = _parse_oi(oi_val, pat)
            # si el OI no cumple el patrón, por seguridad lo saltamos en incremental
            if not oi_key or (last_key is not None and oi_key <= last_key):
                rows_skipped += 1
                emit_progress(r, "skipped_incremental")
                continue
        # fila destino (incremental: continúa debajo del último; completo: desde lista_start_row)
        dst_r = (dst_base_row if cfg.incremental else cfg.lista_start_row) + rows_copied

        # C..N -> B..M (desplazamiento -1)
        for ci in range(src_c0, src_c1 + 1):
            src_cell = ws_vima.cell(row=r, column=ci)
            dst_c = dst_c0 + (ci - src_c0)
            dst_cell = ws_lista.cell(row=dst_r, column=dst_c)

            mrng = find_merge(ws_vima, r, ci) if cfg.replicate_merges else None

            # Si la celda de origen NO es top-left de un merge, openpyxl la expone como MergedCell
            if cfg.replicate_merges and isinstance(src_cell, MergedCell):
                continue

            if cfg.replicate_merges and mrng:
                # Top-left del merge en origen
                rowspan = mrng.max_row - mrng.min_row + 1
                colspan = mrng.max_col - mrng.min_col + 1

                # Limpia solapes en destino
                unmerge_overlaps(ws_lista, dst_r, dst_c, dst_r + rowspan - 1, dst_c + colspan - 1)

                # Copia estilos y limpia valores ANTES de combinar en destino
                for rr in range(dst_r, dst_r + rowspan):
                    for cc in range(dst_c, dst_c + colspan):
                        dst_merge_cell = cast(XLCell, ws_lista.cell(row=rr, column=cc))
                        dst_merge_cell.value = None
                        src_tpl = cast(XLCell, src_cell)
                        copy_style(src_tpl, dst_merge_cell, copy_number_format=True)

                # Replica merge en destino
                ws_lista.merge_cells(
                    start_row=dst_r, start_column=dst_c,
                    end_row=dst_r + rowspan - 1, end_column=dst_c + colspan - 1
                )

                # Escribe SIEMPRE en la top-left del rango destino
                top_left = cast(XLCell, ws_lista.cell(row=dst_r, column=dst_c))
                src_tpl = cast(XLCell, src_cell)
                top_left.value = src_tpl.value
            else:
                # Celda normal: si el destino cae en un merge anterior, descombina
                dmrng = find_merge(ws_lista, dst_r, dst_c) if cfg.replicate_merges else None
                if dmrng:
                    ws_lista.unmerge_cells(str(dmrng))
                    dst_cell = ws_lista.cell(row=dst_r, column=dst_c)

                dst_cell = cast(XLCell, dst_cell)
                src_tpl = cast(XLCell, src_cell)
                dst_cell.value = src_tpl.value
                copy_style(src_tpl, dst_cell, copy_number_format=True)

        rows_copied += 1
        emit_progress(r, "copied")

    if progress_cb:
        progress_cb({
            "stage": "complete",
            "percent": 100.0,
            "rows_processed": processed_rows,
            "rows_total": total_rows,
            "rows_copied": rows_copied,
            "rows_skipped": rows_skipped,
            "current_row": last_row_reported,
        })

    return {
        "rows_copied": rows_copied,
        "rows_skipped": rows_skipped,
        "start_row_vima": cfg.vima_start_row,
        "start_row_lista": cfg.lista_start_row,
    }

