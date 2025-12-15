from __future__ import annotations
import asyncio
import json
from fastapi import APIRouter, HTTPException, UploadFile, File, Form, Depends
from pydantic import BaseModel, Field
from typing import Optional, cast, Dict, Any, Tuple
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
from openpyxl.worksheet.worksheet import Worksheet
from app.oi_tools.services.excel_password import (
    load_workbook_maybe_encrypted,
    load_workbook_fast_for_scan,
    release_workbook_stream,
)
from time import perf_counter
import re
from app.oi_tools.services.excel_io import close_workbook_safe
from app.oi_tools.services.progress_manager import progress_manager, SENTINEL
from app.oi_tools.services.integrations.vima_to_lista import (
    VimaToListaConfig,
    map_vima_to_lista,
    _last_oi_in_lista,
    _parse_oi,
)
from fastapi.responses import StreamingResponse
from io import BytesIO
from zipfile import BadZipFile
from openpyxl.utils.exceptions import InvalidFileException
# Compat: usar un nombre interno para evitar conflictos de tipo con Pylance
try:
    from msoffcrypto.exceptions import InvalidKeyError as _InvalidKeyError # type: ignore
except Exception:  # ModuleNotFoundError u otros
    class _InvalidKeyError(Exception):
        """Dummy cuando msoffcrypto no está disponible."""
        pass
from app.api.auth import get_current_user_session

router = APIRouter(prefix="/integrations", tags=["integrations"], dependencies=[Depends(get_current_user_session)])

@router.get("/vima-to-lista/progress/{operation_id}")
async def vima_progress_stream(operation_id: str):
    channel, history = progress_manager.subscribe(operation_id)
    loop = asyncio.get_running_loop()

    async def event_stream():
        try:
            for event in history:
                yield progress_manager.encode_event(event)
            while True:
                item = await loop.run_in_executor(None, channel.queue.get)
                if item is SENTINEL:
                    break
                yield progress_manager.encode_event(item)
        finally:
            progress_manager.unsubscribe(operation_id)

    return StreamingResponse(event_stream(), media_type="application/x-ndjson")


class VimaToListaPayload(BaseModel):
    vima_path: str = Field(..., description="Ruta al VIMA (.xlsm)")
    lista_path: str = Field(..., description="Ruta a LISTA de OIs (.xlsx)")
    output_path: Optional[str] = Field(None, description="Si None, sobrescribe lista_path")
    vima_password: Optional[str] = Field(None, description="Clave del VIMA (no se persiste)")
    vima_sheet: Optional[str] = None
    lista_sheet: Optional[str] = None
    vima_start_row: int = 11
    lista_start_row: int = 11
    require_all_g_to_n: bool = True
    mode_replace: bool = True
    incremental: bool = False
    oi_pattern: Optional[str] = None # si quieres usar uno distinto, ej. "OI-(\d{4})-(\d+)"
    strict_incremental: bool = False
    replicate_merges: bool = True

@router.post("/vima-to-lista")
def vima_to_lista(payload: VimaToListaPayload):
    wb_vima = None
    wb_lista = None
    try:
        wb_vima = load_workbook_maybe_encrypted(
            payload.vima_path,
            password=payload.vima_password,
            keep_vba=True,
            data_only=True,
        )
        # borramos la clave del modelo para evitar filtraciones posteriores
        payload.vima_password = None

        wb_lista = load_workbook(payload.lista_path, data_only=True, keep_vba=False)
        cfg = VimaToListaConfig(
            vima_sheet=payload.vima_sheet,
            lista_sheet=payload.lista_sheet,
            vima_start_row=payload.vima_start_row,
            lista_start_row=payload.lista_start_row,
            require_all_g_to_n=payload.require_all_g_to_n,
            mode_replace=payload.mode_replace,
            incremental=payload.incremental,
            oi_pattern=payload.oi_pattern or VimaToListaConfig().oi_pattern,
            strict_incremental=payload.strict_incremental,
            replicate_merges=payload.replicate_merges,
        )
        try:
            res = map_vima_to_lista(wb_vima, wb_lista, cfg)
        except ValueError as e:
            # p.ej., incremental estricto: último OI en LISTA no mactchea el patrón
            raise HTTPException(status_code=400, detail=str(e))
        save_path = payload.output_path or payload.lista_path
        wb_lista.save(save_path)
        return {"ok": True, "saved_to": save_path, **res}
    except _InvalidKeyError:
        raise HTTPException(status_code=403, detail="Contraseña incorrecta.", headers={"X-Code": "WRONG_PASSWORD"})
    except (BadZipFile, InvalidFileException):
        # Probable archivo cifrado sin contraseña
        raise HTTPException(status_code=401, detail="El archivo VIMA está protegido. Ingresa la contraseña.", headers={"X-Code":"PASSWORD_REQUIRED"})
    except FileNotFoundError as e:
        raise HTTPException(status_code=404, detail=f"Archivo no encontrado: {e}")
    except HTTPException:
        # Deja pasar los HTTP 4xx/5xx explí­citos que generamos adrede
        raise
    except Exception as e:
        # Cualquier otro error inesperado => 500
        raise HTTPException(status_code=500, detail=f"Error al integrar VIMA → LISTA: {type(e).__name__}")
    finally:
        close_workbook_safe(wb_vima); release_workbook_stream(wb_vima)
        close_workbook_safe(wb_lista)

@router.post("/vima-to-lista/dry-run")
def vima_to_lista_dry_run(payload: VimaToListaPayload):
    wb_vima = None
    wb_lista = None
    try:
        t0 = perf_counter()
        wb_vima = load_workbook_fast_for_scan(payload.vima_path, password=payload.vima_password)
        t1 = perf_counter()
        payload.vima_password = None
        wb_lista = load_workbook(payload.lista_path, data_only=True, keep_vba=False)
        t2 = perf_counter()

        cfg = VimaToListaConfig(
            vima_sheet=payload.vima_sheet,
            lista_sheet=payload.lista_sheet,
            vima_start_row=payload.vima_start_row,
            lista_start_row=payload.lista_start_row,
            require_all_g_to_n=payload.require_all_g_to_n,
            mode_replace=False,
            incremental=payload.incremental,
            oi_pattern=payload.oi_pattern or VimaToListaConfig().oi_pattern,
            strict_incremental=payload.strict_incremental,
            replicate_merges=payload.replicate_merges,
        )

        ws_v = cast(Worksheet, wb_vima[cfg.vima_sheet] if cfg.vima_sheet else wb_vima.active)
        ws_l = cast(Worksheet, wb_lista[cfg.lista_sheet] if cfg.lista_sheet else wb_lista.active)

        pat = re.compile(cfg.oi_pattern, re.IGNORECASE)
        last_key, last_row_idx = (None, cfg.lista_start_row - 1)
        if cfg.incremental:
            last_key, last_row_idx = _last_oi_in_lista(ws_l, cfg)
            if cfg.strict_incremental and last_key is None:
                raise HTTPException(status_code=400, detail="Incremental estricto: último OI no válido en LISTA.")

        row_iter = ws_v.iter_rows(
            min_row=cfg.vima_start_row,
            min_col=3,
            max_col=14,
            values_only=True,
        )
        first_index = None
        first_row = None
        for idx, row in enumerate(row_iter, start=cfg.vima_start_row):
            oi = row[0]
            if oi in (None, "", 0):
                continue
            vals = row[4:12]
            valid = (
                all(v not in (None, "") for v in vals)
                if cfg.require_all_g_to_n
                else any(v not in (None, "") for v in vals)
            )
            if not valid:
                continue
            if cfg.incremental:
                key = _parse_oi(oi, pat)
                if not key or (last_key is not None and key <= last_key):
                    continue
            first_index = idx
            first_row = row
            break

        if first_row is None:
            dst_base_row = (
                max(last_row_idx + 1, cfg.lista_start_row)
                if cfg.incremental
                else cfg.lista_start_row
            )
            last_in_lista = None
            if last_row_idx >= cfg.lista_start_row:
                last_in_lista = ws_l.cell(
                    row=last_row_idx,
                    column=column_index_from_string(cfg.lista_start_col)
                ).value
            summary = {
                "ok": True,
                "would_copy": 0,
                "start_write_row": dst_base_row,
                "last_oi_in_lista": last_in_lista,
                "first_oi_to_copy": None,
                "last_oi_to_copy": None,
                "replicate_merges": cfg.replicate_merges,
            }
            t3 = perf_counter()
            print(f"[DRYRUN(JSON)] abrir_vima={t1-t0:.3f}s abrir_lista={t2-t1:.3f}s total={t3-t0:.3f}s (short)")
            # Garantizamos a Pylance que ya no son None (retornamos arriba si lo fueran)
            return summary
        
        assert first_index is not None
        assert first_row is not None

        would_copy = 0
        first_oi = None
        last_oi = None

        def process_row(r: int, row_values: Tuple[Any, ...]) -> None:
            nonlocal would_copy, first_oi, last_oi
            oi_val = row_values[0]
            if oi_val in (None, "", 0):
                return
            vals_local = row_values[4:12]
            valid_local = (
                all(v not in (None, "") for v in vals_local)
                if cfg.require_all_g_to_n
                else any(v not in (None, "") for v in vals_local)
            )
            if not valid_local:
                return
            if cfg.incremental:
                key_local = _parse_oi(oi_val, pat)
                if not key_local or (last_key is not None and key_local <= last_key):
                    return
                # Si llegó aquí, la fila es válida (en incremental o no incremental)
                would_copy += 1
                if first_oi is None:
                    first_oi = oi_val
                last_oi = oi_val

        process_row(first_index, first_row)
        for r, row_values in enumerate(row_iter, start=first_index + 1):
            process_row(r, row_values)

        dst_base_row = (
            max(last_row_idx + 1, cfg.lista_start_row)
            if cfg.incremental
            else cfg.lista_start_row
        )
        last_in_lista = None
        if last_row_idx >= cfg.lista_start_row:
            last_in_lista = ws_l.cell(
                row=last_row_idx,
                column=column_index_from_string(cfg.lista_start_col)
            ).value

        return {
            "ok": True,
            "would_copy": would_copy,
            "start_write_row": dst_base_row,
            "last_oi_in_lista": last_in_lista,
            "first_oi_to_copy": first_oi,
            "last_oi_to_copy": last_oi,
            "replicate_merges": cfg.replicate_merges,
        }
    except _InvalidKeyError:
        raise HTTPException(status_code=403, detail="Contraseña incorrecta.", headers={"X-Code": "WRONG_PASSWORD"})
    except (BadZipFile, InvalidFileException):
        raise HTTPException(status_code=401, detail="El archivo VIMA está protegido. Ingresa la contraseña.", headers={"X-Code":"PASSWORD_REQUIRED"})
    except FileNotFoundError as e:
        raise HTTPException(status_code=404, detail=f"Archivo no encontrado: {e}")
    finally:
        close_workbook_safe(wb_vima); release_workbook_stream(wb_vima)
        close_workbook_safe(wb_lista)


@router.post("/vima-to-lista/dry-run-upload")
def vima_to_lista_dry_run_upload(
    vima_file: UploadFile = File(...),
    lista_file: UploadFile = File(...),
    vima_password: Optional[str] = Form(None),
    vima_sheet: Optional[str] = Form(None),
    lista_sheet: Optional[str] = Form(None),
    vima_start_row: int = Form(11),
    lista_start_row: int = Form(11),
    require_all_g_to_n: str = Form("true"),
    incremental: str = Form("true"),
    strict_incremental: str = Form("true"),
    oi_pattern: Optional[str] = Form(None),
    replicate_merges: str = Form("true"),
    operation_id: Optional[str] = Form(None),
):
    from openpyxl import load_workbook
    from zipfile import BadZipFile
    from openpyxl.utils.exceptions import InvalidFileException
    import os, tempfile

    def to_bool(s: str) -> bool:
        return str(s).strip().lower() in ("1", "true", "on", "yes", "y")
    require_all_g_to_n_b = to_bool(require_all_g_to_n)
    incremental_b = to_bool(incremental)
    strict_incremental_b = to_bool(strict_incremental)
    replicate_merges_b = to_bool(replicate_merges)

    if operation_id:
        progress_manager.emit(operation_id, {
            "type": "status",
            "stage": "received",
            "message": "Preparando archivos",
            "progress": 0,
        })

    if oi_pattern:
        oi_pattern = oi_pattern.replace("\\\\", "\\")

    vima_tmp = None
    lista_tmp = None
    wb_vima = None
    wb_lista = None
    try:
        vima_name = (vima_file.filename or "").lower()
        lista_name = (lista_file.filename or "").lower()
        if vima_name.endswith(".xls") or lista_name.endswith(".xls"):
            raise HTTPException(status_code=400, detail="Formato .xls no soportado. Convierte a .xlsx/.xlsm.")

        vima_file.file.seek(0)
        lista_file.file.seek(0)
        vima_bytes = vima_file.file.read()
        lista_bytes = lista_file.file.read()

        vfd, vima_tmp = tempfile.mkstemp(suffix=os.path.splitext(vima_name)[1] or ".xlsm")
        os.close(vfd)
        with open(vima_tmp, "wb") as f:
            f.write(vima_bytes)
        lfd, lista_tmp = tempfile.mkstemp(suffix=os.path.splitext(lista_name)[1] or ".xlsx")
        os.close(lfd)
        with open(lista_tmp, "wb") as f:
            f.write(lista_bytes)

        if operation_id:
            progress_manager.emit(operation_id, {
                "type": "status",
                "stage": "loading",
                "message": "Abriendo libros",
            })

        t0 = perf_counter()
        try:
            wb_vima = load_workbook_fast_for_scan(vima_tmp, password=vima_password)
            t1 = perf_counter()
        except _InvalidKeyError:
            raise HTTPException(status_code=403, detail="Contraseña incorrecta.", headers={"X-Code": "WRONG_PASSWORD"})
        except (BadZipFile, InvalidFileException):
            if not vima_password:
                raise HTTPException(status_code=401, detail="El archivo VIMA está protegido. Ingresa la contraseña.", headers={"X-Code":"PASSWORD_REQUIRED"})
            raise HTTPException(status_code=400, detail="El archivo VIMA no es un Excel válido.")

        try:
            wb_lista = load_workbook(lista_tmp, keep_vba=False, data_only=True)
            t2 = perf_counter()
        except (BadZipFile, InvalidFileException):
            raise HTTPException(status_code=400, detail="El archivo LISTA no es un Excel válido.")

        cfg = VimaToListaConfig(
            vima_sheet=vima_sheet,
            lista_sheet=lista_sheet,
            vima_start_row=int(vima_start_row),
            lista_start_row=int(lista_start_row),
            require_all_g_to_n=require_all_g_to_n_b,
            mode_replace=False,
            incremental=incremental_b,
            oi_pattern=oi_pattern or VimaToListaConfig().oi_pattern,
            strict_incremental=strict_incremental_b,
            replicate_merges=replicate_merges_b,
        )

        ws_v = cast(Worksheet, wb_vima[cfg.vima_sheet] if cfg.vima_sheet else wb_vima.active)
        ws_l = cast(Worksheet, wb_lista[cfg.lista_sheet] if cfg.lista_sheet else wb_lista.active)

        pat = re.compile(cfg.oi_pattern, re.IGNORECASE)
        last_key, last_row_idx = (None, cfg.lista_start_row - 1)
        if cfg.incremental:
            last_key, last_row_idx = _last_oi_in_lista(ws_l, cfg)
            if cfg.strict_incremental and last_key is None:
                raise HTTPException(status_code=400, detail="Incremental estricto: último OI no válido en LISTA.")

        def emit_analysis_progress(current_row: int) -> None:
            if not operation_id:
                return
            progress_manager.emit(operation_id, {
                "type": "progress",
                "stage": "analysis",
                "current_row": current_row,
            })

        row_iter = ws_v.iter_rows(
            min_row=cfg.vima_start_row,
            min_col=3,
            max_col=14,
            values_only=True,
        )
        first_index = None
        first_row = None
        for idx, row in enumerate(row_iter, start=cfg.vima_start_row):
            oi = row[0]
            if oi in (None, "", 0):
                emit_analysis_progress(idx)
                continue
            vals = row[4:12]
            valid = (
                all(v not in (None, "") for v in vals)
                if cfg.require_all_g_to_n
                else any(v not in (None, "") for v in vals)
            )
            if not valid:
                emit_analysis_progress(idx)
                continue
            if cfg.incremental:
                key = _parse_oi(oi, pat)
                if not key or (last_key is not None and key <= last_key):
                    emit_analysis_progress(idx)
                    continue
            first_index = idx
            first_row = row
            break

        if first_row is None:
            dst_base_row = max(last_row_idx + 1, cfg.lista_start_row) if cfg.incremental else cfg.lista_start_row
            last_in_lista = None
            if last_row_idx >= cfg.lista_start_row:
                last_in_lista = ws_l.cell(
                    row=last_row_idx,
                    column=column_index_from_string(cfg.lista_start_col)
                ).value
            summary = {
                "ok": True,
                "would_copy": 0,
                "start_write_row": dst_base_row,
                "last_oi_in_lista": last_in_lista,
                "first_oi_to_copy": None,
                "last_oi_to_copy": None,
                "replicate_merges": cfg.replicate_merges,
            }
            t3 = perf_counter()
            print(f"[DRYRUN(UPLOAD)] abrir_vima={t1-t0:.3f}s abrir_lista={t2-t1:.3f}s escanear={(t3-t2):.3f}s total={t3-t0:.3f}s (short)")
            if operation_id:
                progress_manager.emit(operation_id, {
                    "type": "complete",
                    "message": "Dry-run sin filas nuevas",
                    "percent": 100.0,
                    "result": summary,
                })
            return summary
        # Garantizar non-None para Pylance (ya retornamos arriba si no hubo primera fila)
        assert first_index is not None
        assert first_row is not None

        would_copy = 0
        first_oi = None
        last_oi = None

        def process_row(r: int, row_values: Tuple[Any, ...]) -> None:
            nonlocal would_copy, first_oi, last_oi
            oi_val = row_values[0]
            if oi_val in (None, "", 0):
                emit_analysis_progress(r)
                return
            vals_local = row_values[4:12]
            valid_local = (
                all(v not in (None, "") for v in vals_local)
                if cfg.require_all_g_to_n
                else any(v not in (None, "") for v in vals_local)
            )
            if not valid_local:
                emit_analysis_progress(r)
                return
            if cfg.incremental:
                key_local = _parse_oi(oi_val, pat)
                if not key_local or (last_key is not None and key_local <= last_key):
                    emit_analysis_progress(r)
                    return
                # Fila válida (pasa filtro general y, si aplica, filtro incremental)
                would_copy += 1
                if first_oi is None:
                    first_oi = oi_val
                last_oi = oi_val
                emit_analysis_progress(r)

        process_row(first_index, first_row)
        for r, row_values in enumerate(row_iter, start=first_index + 1):
            process_row(r, row_values)

        dst_base_row = max(last_row_idx + 1, cfg.lista_start_row) if cfg.incremental else cfg.lista_start_row
        last_in_lista = None
        if last_row_idx >= cfg.lista_start_row:
            last_in_lista = ws_l.cell(
                row=last_row_idx,
                column=column_index_from_string(cfg.lista_start_col)
            ).value

        summary = {
            "ok": True,
            "would_copy": would_copy,
            "start_write_row": dst_base_row,
            "last_oi_in_lista": last_in_lista,
            "first_oi_to_copy": first_oi,
            "last_oi_to_copy": last_oi,
            "replicate_merges": cfg.replicate_merges,
        }
        if operation_id:
            progress_manager.emit(operation_id, {
                "type": "complete",
                "message": "Dry-run completado",
                "percent": 100.0,
                "result": summary,
            })
        return summary
    except HTTPException as exc:
        if operation_id:
            event: Dict[str, Any] = {
                "type": "error",
                "status": exc.status_code,
                "detail": exc.detail,
            }
            code = (exc.headers or {}).get("X-Code") if hasattr(exc, "headers") else None
            if code:
                event["code"] = code
            progress_manager.emit(operation_id, event)
        raise
    except Exception as exc:
        if operation_id:
            progress_manager.emit(operation_id, {
                "type": "error",
                "status": 500,
                "detail": str(exc),
            })
        raise
    finally:
        try:
            if wb_vima: wb_vima.close(); release_workbook_stream(wb_vima)
        except: pass
        try:
            if wb_lista: wb_lista.close()
        except: pass
        if vima_tmp and os.path.exists(vima_tmp):
            try: os.unlink(vima_tmp)
            except: pass
        if lista_tmp and os.path.exists(lista_tmp):
            try: os.unlink(lista_tmp)
            except: pass

        if operation_id:
            progress_manager.finish(operation_id)


@router.post("/vima-to-lista/upload")
def vima_to_lista_upload(
    vima_file: UploadFile = File(...),
    lista_file: UploadFile = File(...),
    vima_password: Optional[str] = Form(None),
    vima_sheet: Optional[str] = Form(None),
    lista_sheet: Optional[str] = Form(None),
    vima_start_row: int = Form(11),
    lista_start_row: int = Form(11),
    require_all_g_to_n: str = Form("true"),
    incremental: str = Form("true"),
    strict_incremental: str = Form("true"),
    oi_pattern: Optional[str] = Form(None),
    replicate_merges: str = Form("true"),
    operation_id: Optional[str] = Form(None),
):
    from openpyxl import load_workbook
    from zipfile import BadZipFile
    from openpyxl.utils.exceptions import InvalidFileException
    import os, tempfile

    def to_bool(s: str) -> bool:
        return str(s).strip().lower() in ("1", "true", "on", "yes", "y")
    require_all_g_to_n_b = to_bool(require_all_g_to_n)
    incremental_b = to_bool(incremental)
    strict_incremental_b = to_bool(strict_incremental)
    replicate_merges_b = to_bool(replicate_merges)

    if operation_id:
        progress_manager.emit(operation_id, {
            "type": "status",
            "stage": "received",
            "message": "Preparando archivos",
            "progress": 0,
        })

    if oi_pattern:
        oi_pattern = oi_pattern.replace("\\\\", "\\")

    vima_tmp = None
    lista_tmp = None
    wb_vima = None
    wb_lista = None
    out_stream = BytesIO()
    try:
        vima_name = (vima_file.filename or "").lower()
        lista_name = (lista_file.filename or "").lower()
        if vima_name.endswith(".xls") or lista_name.endswith(".xls"):
            raise HTTPException(status_code=400, detail="Formato .xls no soportado. Convierte a .xlsx/.xlsm.")

        vima_file.file.seek(0)
        lista_file.file.seek(0)
        vima_bytes = vima_file.file.read()
        lista_bytes = lista_file.file.read()

        vfd, vima_tmp = tempfile.mkstemp(suffix=os.path.splitext(vima_name)[1] or ".xlsm")
        os.close(vfd)
        with open(vima_tmp, "wb") as f:
            f.write(vima_bytes)
        lfd, lista_tmp = tempfile.mkstemp(suffix=os.path.splitext(lista_name)[1] or ".xlsx")
        os.close(lfd)
        with open(lista_tmp, "wb") as f:
            f.write(lista_bytes)

        if operation_id:
            progress_manager.emit(operation_id, {
                "type": "status",
                "stage": "loading",
                "message": "Abriendo libros",
            })

        cfg = VimaToListaConfig(
            vima_sheet=vima_sheet,
            lista_sheet=lista_sheet,
            vima_start_row=int(vima_start_row),
            lista_start_row=int(lista_start_row),
            require_all_g_to_n=require_all_g_to_n_b,
            mode_replace=False,
            incremental=incremental_b,
            oi_pattern=oi_pattern or VimaToListaConfig().oi_pattern,
            strict_incremental=strict_incremental_b,
            replicate_merges=replicate_merges_b,
        )
        pat = re.compile(cfg.oi_pattern, re.IGNORECASE)

        try:
            wb_lista = load_workbook(lista_tmp, keep_vba=False, data_only=True)
        except (BadZipFile, InvalidFileException):
            raise HTTPException(status_code=400, detail="El archivo LISTA no es un Excel válido.")

        ws_l = cast(Worksheet, wb_lista[cfg.lista_sheet] if cfg.lista_sheet else wb_lista.active)
        last_key, last_row_idx = (None, cfg.lista_start_row - 1)
        if cfg.incremental:
            last_key, last_row_idx = _last_oi_in_lista(ws_l, cfg)
            if cfg.strict_incremental and last_key is None:
                raise HTTPException(status_code=400, detail="Incremental estricto: último OI no válido en LISTA.")

        wb_vima_fast = None
        has_valid: Optional[bool] = None
        try:
            wb_vima_fast = load_workbook_fast_for_scan(vima_tmp, password=vima_password)
            ws_fast = cast(Worksheet, wb_vima_fast[cfg.vima_sheet] if cfg.vima_sheet else wb_vima_fast.active)
            for row in ws_fast.iter_rows(min_row=cfg.vima_start_row, min_col=3, max_col=14, values_only=True):
                oi = row[0]
                if oi in (None, "", 0):
                    continue
                vals = row[4:12]
                ok = (
                    all(v not in (None, "") for v in vals)
                    if cfg.require_all_g_to_n
                    else any(v not in (None, "") for v in vals)
                )
                if not ok:
                    continue
                if cfg.incremental:
                    key = _parse_oi(oi, pat)
                    if not key or (last_key is not None and key <= last_key):
                        continue
                has_valid = True
                break
            if has_valid is None:
                has_valid = False
        except _InvalidKeyError:
            raise HTTPException(status_code=403, detail="Contraseña incorrecta.", headers={"X-Code": "WRONG_PASSWORD"})
        except (BadZipFile, InvalidFileException):
            if not vima_password:
                raise HTTPException(status_code=401, detail="El archivo VIMA está protegido. Ingresa la contraseña.", headers={"X-Code":"PASSWORD_REQUIRED"})
            has_valid = True  # seguimos con la apertura completa
        finally:
            if wb_vima_fast is not None:
                try:
                    wb_vima_fast.close()
                finally:
                    release_workbook_stream(wb_vima_fast)

        if has_valid is False:
            if operation_id:
                progress_manager.emit(operation_id, {
                    "type": "complete",
                    "message": "Sin filas nuevas para copiar",
                    "percent": 100.0,
                    "download_ready": True,
                    "result": {"rows_copied": 0, "rows_skipped": 0},
                })
            out_stream = BytesIO(lista_bytes)
            out_stream.seek(0)
            headers = {"Content-Disposition": 'attachment; filename="LISTA_SALIDA.xlsx"'}
            wb_lista.close()
            wb_lista = None
            return StreamingResponse(
                out_stream,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers=headers
            )

        try:
            if vima_password:
                wb_vima = load_workbook_maybe_encrypted(
                    vima_tmp, password=vima_password, keep_vba=True, data_only=True
                )
            else:
                wb_vima = load_workbook(vima_tmp, keep_vba=True, data_only=True)
        except _InvalidKeyError:
            raise HTTPException(status_code=403, detail="Contraseña incorrecta.", headers={"X-Code": "WRONG_PASSWORD"})
        except (BadZipFile, InvalidFileException):
            if not vima_password:
                raise HTTPException(status_code=401, detail="El archivo VIMA está protegido. Ingresa la contraseña.", headers={"X-Code":"PASSWORD_REQUIRED"})
            raise HTTPException(status_code=400, detail="El archivo VIMA no es un Excel válido.")

        def forward_progress(payload: Dict[str, Any]) -> None:
            if not operation_id:
                return
            progress_manager.emit(operation_id, {
                "type": "progress",
                **payload,
            })

        if operation_id:
            progress_manager.emit(operation_id, {
                "type": "status",
                "stage": "processing",
                "message": "Procesando filas",
                "progress": 0,
            })

        res = map_vima_to_lista(wb_vima, wb_lista, cfg, progress_cb=forward_progress)

        if operation_id:
            progress_manager.emit(operation_id, {
                "type": "status",
                "stage": "saving",
                "message": "Generando archivo de salida",
            })

        wb_lista.save(out_stream)
        wb_lista.close()
        wb_lista = None
        out_stream.seek(0)
        headers = {"Content-Disposition": 'attachment; filename="LISTA_SALIDA.xlsx"'}

        if operation_id:
            progress_manager.emit(operation_id, {
                "type": "complete",
                "message": "Integración completada",
                "percent": 100.0,
                "download_ready": True,
                "result": res,
            })

        return StreamingResponse(
            out_stream,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers=headers
        )
    except HTTPException:
        raise
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Error al integrar (upload): {type(e).__name__}: {e}"
        )
    finally:
        try:
            if wb_vima: wb_vima.close(); release_workbook_stream(wb_vima)
        except: pass
        try:
            if wb_lista: wb_lista.close()
        except: pass
        if vima_tmp and os.path.exists(vima_tmp):
            try: os.unlink(vima_tmp)
            except: pass
        if lista_tmp and os.path.exists(lista_tmp):
            try: os.unlink(lista_tmp)
            except: pass

        if operation_id:
            progress_manager.finish(operation_id)
