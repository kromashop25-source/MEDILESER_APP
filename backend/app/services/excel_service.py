from io import BytesIO
from pathlib import Path
from typing import Iterable, Tuple, Optional, cast
from datetime import datetime, timezone
from zoneinfo import ZoneInfo
from zipfile import ZipFile, ZIP_DEFLATED
from xml.etree import ElementTree as ET

from openpyxl import load_workbook, Workbook
from openpyxl.formula.translate import Translator
from openpyxl.utils.cell import column_index_from_string
from openpyxl.utils.cell import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Border, Side
from openpyxl.workbook.protection import WorkbookProtection
from openpyxl.utils.protection import hash_password
from openpyxl.cell.cell import MergedCell
from openpyxl.styles.protection import Protection as CellProtection

from ..models import OI, Bancada
from ..core.settings import get_settings
from .rules_service import pma_to_pressure, find_exact_in_range, normalize_for_excel_list


HEADER_ROW = 8
DATA_START_ROW = 9
SHEET_NAME = "ERROR FINAL"
Q3_RANGE = "AZ2:BC2"       # lista para E4
ALCANCE_RANGE = "AZ1:BE1"  # lista para O4
# CORRECCIÓN: Copiamos fórmulas desde la Col Q (tiempo) hasta el final
FORMULA_START_COL = "Q"
FORMULA_END_COL = "BL"
WORKBOOK_XML_PATH = "xl/workbook.xml"
EXCEL_MAIN_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
MANUAL_LI_LF_COLS = {
    column_index_from_string("Y"),   # Q2 L.I.
    column_index_from_string("Z"),   # Q2 L.F.
    column_index_from_string("AK"),  # Q1 L.I.
    column_index_from_string("AL"),  # Q1 L.F.
}

def _inject_reservation_notice(xlsx_bytes: bytes, hashed_password: str, reserved_by: str | None = None) -> bytes:
    """
    Añade o actualiza la etiqueta <fileSharing> en workbook.xml para forzar
    el mensaje de "solo lectura" cuando no se conoce la contraseña.
    """
    src = BytesIO(xlsx_bytes)
    dst = BytesIO()
    with ZipFile(src, "r") as zin, ZipFile(dst, "w", ZIP_DEFLATED) as zout:
        for info in zin.infolist():
            data = zin.read(info.filename)
            if info.filename == WORKBOOK_XML_PATH:
                data = _update_workbook_xml(data, hashed_password, reserved_by)
            zout.writestr(info, data)
    dst.seek(0)
    return dst.read()

def _update_workbook_xml(xml_bytes: bytes, hashed_password: str, reserved_by: str | None) -> bytes:
    ET.register_namespace("", EXCEL_MAIN_NS)
    root = ET.fromstring(xml_bytes)
    tag = f"{{{EXCEL_MAIN_NS}}}fileSharing"
    node = root.find(tag)
    if node is None:
        node = ET.Element(tag)
        file_version = root.find(f"{{{EXCEL_MAIN_NS}}}fileVersion")
        insert_idx = 1 if file_version is not None else 0
        root.insert(insert_idx, node)
    node.set("reservationPassword", hashed_password)
    if reserved_by:
        node.set("userName", reserved_by)
    return ET.tostring(root, encoding="utf-8", xml_declaration=True)

def _find_header_col(ws: Worksheet, header_name: str, header_row: int = HEADER_ROW) -> Optional[int]:
    """
    Busca la columna de una cabecera por texto exacto (case-insensitive) en la fila `header_row`.
    Retorna el índice de columna (1..N) o None si no existe.
    """
    target = header_name.strip().lower()
    for cell in ws[header_row]:
        text = "" if cell.value is None else str(cell.value).strip().lower()
        if text == target:
            # openpyxl ofrece cell.col_idx (int). SI no, convertimos desde .column (letra o int).
            col = getattr(cell, "col_idx", None)
            if isinstance(col, int):
                return col
            raw = getattr(cell, "column", None)
            if isinstance(raw, int):
                return raw
            if isinstance(raw, str):
                return column_index_from_string(raw)
    return None

def _ensure_workbook() -> Tuple[Workbook, Worksheet]:
    settings = get_settings()
    tpl = Path(settings.template_abs_path)
    if tpl.exists():
        # Mantener vínculos externos tal cual en la plantilla para evitar
        # los avisos de “reparaciones” al abrir en Excel.
        wb = load_workbook(tpl, data_only=False, keep_links=True)
        active = wb.active or (wb.worksheets[0] if wb.worksheets else None)
        if active is None:
            wb.create_sheet("Sheet1")
            active = wb.worksheets[0]
        ws = cast(Worksheet, active)
    else:
        # Fallback para no bloquear pruebas si la plantilla no está
        wb = Workbook()
        active = wb.active or (wb.worksheets[0] if wb.worksheets else None)
        if active is None:
            wb.create_sheet("Sheet1")
            active = wb.worksheets[0]
        ws = cast(Worksheet, active)
        ws["A8"] = "Item"
        ws["B8"] = "# Medidor"
        ws["C8"] = "Estado"
    return wb, ws

def _get_sheet(wb: Workbook, name: str) -> Worksheet:
    if name in wb.sheetnames:
        return wb[name]
    active = wb.active or (wb.worksheets[0] if wb.worksheets else None)
    if active is None:
        wb.create_sheet("Sheet1")
        active = wb.worksheets[0]
    return cast(Worksheet, active)

def _iter_range_values(ws: Worksheet, range_ref: str) -> list[str]:
    """Devuelve los valores del rango ya normalizados al formato de la lista (coma y 1 decimal)."""
    vals: list[str] = []
    for row in ws[range_ref]:
        for cell in row:
            # normalize_for_excel_list puede devolver None → forzamos str
            vals.append(normalize_for_excel_list(cell.value) or "")
    return vals

def _copy_row_styles(src_ws: Worksheet, src_row: int, dst_ws: Worksheet, dst_row: int, max_col: int) -> None:
    """Replica estilos de la fila `src_row` en `dst_row` (A..max_col).
    Evita celdas fusionadas y clona `protection` para no pasar un StyleProxy."""
    # altura de fila idéntica (si aplica)
    try:
        dst_ws.row_dimensions[dst_row].height = src_ws.row_dimensions[src_row].height
    except Exception:
        pass
    for c in range(1, max_col + 1):
        src = src_ws.cell(row=src_row, column=c)
        dst = dst_ws.cell(row=dst_row, column=c)
        # no asignar estilo a celdas fusionadas no-ancla
        if isinstance(dst, MergedCell):
            continue
        if src.has_style:
            try:
                dst.font = src.font.copy()
                dst.number_format = src.number_format
                dst.alignment = src.alignment.copy()
                dst.fill = src.fill.copy()
                dst.border = src.border.copy()
                # Protection: clonar; evitar StyleProxy
                prot = getattr(src, "protection", None)
                if prot is not None:
                    try:
                        dst.protection = prot.copy()
                    except Exception:
                        dst.protection = CellProtection(locked=prot.locked, hidden=prot.hidden)
            except Exception:
                # si algo falla en una celda, continuar con la siguiente
                continue

def _apply_thick_bottom_border(ws: Worksheet, row: int, start_col_letter: str, end_col_letter: str) -> None:
    start_col = column_index_from_string(start_col_letter)
    end_col = column_index_from_string(end_col_letter)
    thick = Side(style="thick")
    for c in range(start_col, end_col + 1):
        cell = ws.cell(row=row, column=c)
        b = cell.border
        cell.border = Border(left=b.left, right=b.right, top=b.top, bottom=thick)

def _copy_formulas(ws: Worksheet, src_row: int, dst_row: int, start_col_letter: str, end_col_letter: str) -> None:
    start_col = column_index_from_string(start_col_letter)
    end_col = column_index_from_string(end_col_letter)
    for c in range(start_col, end_col + 1):
        # NO copiar nada en las columnas manuales de L.I/L.F de Q2/Q1
        if c in MANUAL_LI_LF_COLS:
            continue
        src_cell = ws.cell(row=src_row, column=c)
        dst_cell = ws.cell(row=dst_row, column=c)
        # Evitar escribir en celdas fusionadas (solo la ancla acepta value)
        if isinstance(dst_cell, MergedCell):
            continue
        if src_cell.data_type == "f" or (isinstance(src_cell.value, str) and str(src_cell.value).startswith("=")):
            formula = str(src_cell.value)
            dst_cell.value = Translator(formula, origin=src_cell.coordinate).translate_formula(dst_cell.coordinate)
        else:
            try:
                dst_cell.value = src_cell.value
            except AttributeError:
                # Si el destino está mergeado/protegido, omitir
                pass 

def _to_float_or_none(val) -> Optional[float]:
    """
    Intenta convertir un valor genérico a float.
    Acepta strings con coma o punto decimal.
    Devuelve None si no es convertible (None, "", texto no numérico).
    """
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    if isinstance(val, str):
        s = val.strip()
        if not s:
            return None
        # Intento directo
        for candidate in (s, s.replace(",", ".")):
            try:
                return float(candidate)
            except (TypeError, ValueError):
                continue
    return None
    
def _coerce_estado(val) -> Optional[int]:
    try:
        n = int(val)
    except (TypeError, ValueError):
        return None
    return max(0, min(5, n))

def generate_excel(
    oi: OI,
    bancadas: Iterable[Bancada],
    password: str | None = None,
    work_dt: Optional[datetime] = None,
) -> Tuple[bytes, str]:
    wb, _ws_active = _ensure_workbook()
    ws = _get_sheet(wb, SHEET_NAME)  # usar siempre "ERROR FINAL"

    # Celdas fijas de cabecera (selección exacta desde listas)
    q3_candidates = _iter_range_values(ws, Q3_RANGE)
    alcance_candidates = _iter_range_values(ws, ALCANCE_RANGE)
    
    # normalize_for_excel_list puede devolver None → forzamos str con ""
    q3_value = find_exact_in_range(q3_candidates, normalize_for_excel_list(oi.q3) or "")
    alcance_value = find_exact_in_range(alcance_candidates, normalize_for_excel_list(oi.alcance) or "")
    
    if q3_value is None:
        raise ValueError("Q3 no coincide con la lista de la plantilla")
    if alcance_value is None:
        raise ValueError("Alcance no coincide con la lista de la plantilla")
        
    ws["E4"] = q3_value
    ws["O4"] = alcance_value

    # Asegurar cabecera "Estado" (fila 8)
    estado_col = _find_header_col(ws, "Estado", header_row=HEADER_ROW)
    if estado_col is None:
        last_col = ws.max_column + 1
        ws.cell(row=HEADER_ROW, column=last_col, value="Estado")
        estado_col = last_col

    # Otras columnas conocidas (opcionales)
    medidor_col = (
        _find_header_col(ws, "# Medidor", header_row=HEADER_ROW)
        or _find_header_col(ws, "# Medidor", header_row=6)
    )
    if medidor_col is None:
        medidor_col = column_index_from_string("G")

    # Escribir filas desde la 9
    rows = list(bancadas)
    # Ordenar por item si existe
    rows.sort(key=lambda b: (b.item or 0))
    
    # Datos globales para columnas B, C, D, E, H
    work_dt = work_dt or datetime.utcnow()
    if work_dt.tzinfo is None:
        dt_utc = work_dt.replace(tzinfo=timezone.utc)
    else:
        dt_utc = work_dt.astimezone(timezone.utc)
    dt_pe = dt_utc.astimezone(ZoneInfo("America/Lima"))
    today_date = dt_pe.date()
    presion_val = pma_to_pressure(oi.pma) if oi.pma else None

    current_row = DATA_START_ROW
    for i, b in enumerate(rows, start=1):
        # 1. Detectar fuente de filas: ¿Tiene data del Grid (rows_data) o es legacy?
        rows_source = getattr(b, "rows_data", []) or []
        nrows = len(rows_source) if rows_source else int(getattr(b, "rows", 15) or 15)

        # 2. Iterar fila por fila
        for k in range(nrows):
            r = current_row + k
            # Obtener payload de la fila k (si existe)
            row_payload = rows_source[k] if (rows_source and k < len(rows_source)) else {}

            # Col A: Item incremental
            item_value = current_row - DATA_START_ROW + 1 + k
            ws.cell(row=r, column=1, value=item_value)

            # Col B y C: Fechas
            ws.cell(row=r, column=2, value=today_date)
            ws.cell(row=r, column=3, value=today_date)

            # Col D y E: Banco y Técnico
            ws.cell(row=r, column=4, value=oi.banco_id)
            ws.cell(row=r, column=5, value=oi.tech_number)

            # Col G: Medidor (Prioridad: Fila > Bancada > Vacío)
            val_medidor = row_payload.get("medidor") or b.medidor
            if medidor_col:
                ws.cell(row=r, column=medidor_col, value=val_medidor or "")

            # --- LÓGICA DE REPLICACIÓN VERTICAL (FILA MAESTRA VS ESCLAVA) ---
            # Col H: Presión (Referencia Vertical)
            if presion_val is not None:
                if k == 0:
                    ws.cell(row=r, column=8, value=presion_val)
                else:
                    ws.cell(row=r, column=8, value=f"=H{r-1}")

            # Col I: Estado
            row_estado = _coerce_estado(row_payload.get("estado") if isinstance(row_payload, dict) else None)
            bancada_estado = _coerce_estado(getattr(b, "estado", None))
            if rows_source:
                ws.cell(row=r, column=estado_col, value=row_estado if row_estado is not None else (bancada_estado or 0))
            else:
                if k == 0:
                    ws.cell(row=r, column=estado_col, value=bancada_estado if bancada_estado is not None else 0)
                else:
                    ws.cell(row=r, column=estado_col, value=f"=I{r-1}")

            # --- Funciones internas auxiliares (en scope de r y k) ---
            def _block_has_data(block) -> bool:
                if not isinstance(block, dict):
                    return False
                for key in ("c1", "c2", "c3", "c4", "c5", "c6", "c7", "c7_seconds"):
                    val = block.get(key, None)
                    if val not in (None, "", 0):
                        return True
                return False

            def _clear_block(row_idx: int, start_col: int) -> None:
                # c1..c7
                for offset in range(7):
                    # Limpiar a mano para evitar arrastrar datos de otras bancadas (openpyxl ignora value=None)
                    ws.cell(row=row_idx, column=start_col + offset).value = None

            def _write_block(start_col, block):
                # Columns: c1(Temp), c2(P.In), c3(P.Out), c4(LI), c5(LF), c6(Vol), c7(Time TXT)
                shared_indices = {0, 1, 2, 5, 6} 
                
                # TAREA 5: Leer valores de tiempo
                time_text_val = block.get("c7", None)
                time_seconds_val = block.get("c7_seconds", None)

                for idx, key in enumerate(["c1", "c2", "c3", "c4", "c5", "c6", "c7"]):
                    val = block.get(key)
                    num_val = _to_float_or_none(val)
                    target_col = start_col + idx
                    col_letter = get_column_letter(target_col)

                    # Lecturas manuales: solo escribir si vienen informadas
                    if idx in (3, 4) and (val is None or val == ""):
                        cell = ws.cell(row=r, column=target_col)
                        cell.value = None  # borra fórmula/valor
                        try:
                            cell.data_type = "n"
                        except Exception:
                            pass
                        continue

                    # Q1 L.I. (AK) independiente
                    if start_col == 34 and idx == 3:
                        if val is None:
                            ws.cell(row=r, column=target_col, value=None)
                        else:
                            ws.cell(
                                row=r,
                                column=target_col,
                                value=num_val if num_val is not None else val,
                            )
                    # Para las columnas compartidas, replicar verticalmente
                    elif k > 0 and idx in shared_indices:
                        ws.cell(row=r, column=target_col, value=f"={col_letter}{r-1}")
                    # Columna de tiempo (texto)
                    elif idx == 6:
                        ws.cell(row=r, column=target_col, value=time_text_val or "")
                    else:
                        # Fila base o campos individuales
                        if num_val is not None:
                            ws.cell(row=r, column=target_col, value=num_val)
                        elif val is not None:
                            ws.cell(row=r, column=target_col, value=val)

            # --- Fin funciones internas ---

            # Copiar primero las fórmulas
            _copy_formulas(ws, DATA_START_ROW, r, FORMULA_START_COL, FORMULA_END_COL)

            # Bloques de datos
            q3_block = row_payload.get("q3")
            q2_block = row_payload.get("q2")
            q1_block = row_payload.get("q1")

            # Q3
            if _block_has_data(q3_block):
                _write_block(10, q3_block)  # J=10
            else:
                _clear_block(r, 10)

            # Q2
            if _block_has_data(q2_block):
                _write_block(22, q2_block)  # V=22
            else:
                _clear_block(r, 22)

            # Q1
            if _block_has_data(q1_block):
                _write_block(34, q1_block)  # AH=34
            else:
                _clear_block(r, 34)

            # Replicar estilo
            _copy_row_styles(ws, DATA_START_ROW, ws, r, column_index_from_string(FORMULA_END_COL))
        
        # Actualizar puntero global de filas
        current_row += nrows

        # Borde inferior grueso
        _apply_thick_bottom_border(ws, current_row - 1, "A", FORMULA_END_COL)

    # --- Desbloquear celdas editables ---
    unlocked_protection = CellProtection(locked=False)

    # 1. Cabeceras Editables
    ws["E4"].protection = unlocked_protection
    ws["O4"].protection = unlocked_protection

    # 2. Filas de Datos
    editable_col_ranges = [1, 7] + list(range(9, 17)) + list(range(22, 29)) + list(range(34, 41))

    if current_row > DATA_START_ROW:
        for r_idx in range(DATA_START_ROW, current_row):
            for c_idx in editable_col_ranges:
                try:
                    ws.cell(row=r_idx, column=c_idx).protection = unlocked_protection
                except AttributeError:
                    pass

    # ----- Protección interna de libro y hoja (celdas bloqueadas) -----
    settings = get_settings()
    internal_pwd = getattr(settings, "cells_protection_password", None)

    # Proteger estructura de libro con contraseña interna
    wb.security = WorkbookProtection(lockStructure=True, lockRevision=True)
    if internal_pwd:
        wb.security.set_workbook_password(internal_pwd)
        wb.security.set_revisions_password(internal_pwd)

    # Proteger hoja "ERROR FINAL" con misma contraseña interna
    ws.protection.enable()
    if internal_pwd:
        ws.protection.set_password(internal_pwd)

    # ----- Contraseña ingresada por el usuario → solo lectura recomendada -----
    hashed_reservation: str | None = None
    if password:
        # Esta contraseña SOLO se usa para el mensaje de "libro reservado / abrir como solo lectura".
        hashed_reservation = hash_password(password)

    # Guardar en memoria
    buf = BytesIO()
    wb.save(buf)
    workbook_bytes = buf.getvalue()

    # Inyectar fileSharing con la contraseña del usuario (lectura recomendada)
    if hashed_reservation:
        reserved_by = f"Banco{oi.banco_id:02d}" if oi.banco_id is not None else None
        workbook_bytes = _inject_reservation_notice(workbook_bytes, hashed_reservation, reserved_by)

    filename = f"{oi.code}.xlsx"
    return workbook_bytes, filename
