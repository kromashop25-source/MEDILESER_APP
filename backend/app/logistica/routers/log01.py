from __future__ import annotations

import asyncio
import logging
import os
import queue
import re
import shutil
import tempfile
import threading
import time
import unicodedata
import uuid
from copy import copy
from dataclasses import dataclass
from datetime import date, datetime
from io import BytesIO
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple, cast
from typing import TypedDict

from fastapi import APIRouter, Depends, File, Form, HTTPException, UploadFile
from fastapi.responses import FileResponse, JSONResponse, Response, StreamingResponse
from openpyxl import load_workbook
from openpyxl.utils.datetime import WINDOWS_EPOCH, from_excel
from openpyxl.cell.cell import Cell
from openpyxl.styles import Alignment, Font
from openpyxl.worksheet.worksheet import Worksheet

from app.api.auth import get_current_user_session
from app.core.settings import get_settings
from app.oi_tools.services.progress_manager import progress_manager, _SENTINEL as SENTINEL # mismo sentinel
from app.oi_tools.services.cancel_manager import cancel_manager, CancelToken

router = APIRouter(
    prefix="/logistica/log01",
    tags=["logistica/log01"],
    dependencies=[Depends(get_current_user_session)],
)

logger = logging.getLogger(__name__)
_LOG01_HELLO_PAD = " " * 2048

class Log01Cancelled(Exception):
    pass


@dataclass
class Log01InputFile:
    name: str
    data: Optional[bytes] = None
    path: Optional[str] = None


@dataclass
class Log01Job:
    operation_id: str
    created_at: float
    status: str
    work_dir: str
    output_name: Optional[str] = None
    result_path: Optional[str] = None
    error: Optional[str] = None


LOG01_JOBS: Dict[str, Log01Job] = {}
LOG01_JOBS_LOCK = threading.Lock()
LOG01_TTL_SECONDS = 60 * 30

# ----------------------------
# Utilitarios
# ----------------------------

_OI_RE = re.compile(r"OI-(\d{4})-(\d{4})", re.IGNORECASE)

def _parse_oi_number_from_filename(filename: str) -> int:
    m = _OI_RE.search(filename or "")
    if not m:
        raise ValueError(
            f"INVALID_OI_FILENAME · Nombre inválido: no se encontró patrón OI-####-YYYY en '{filename}'"
        )
    return int(m.group(1))
    

def _norm_str(v: Any) -> str:
    return str(v).strip() if v is not None else ""

def _norm_header(v: Any) -> str:
    s = _norm_str(v).lower()
    if not s:
        return ""
    s = unicodedata.normalize("NFKD", s)
    s = "".join(ch for ch in s if not unicodedata.combining(ch))
    s = re.sub(r"[^a-z0-9]+", " ", s).strip()
    return s

def _classify_file_error(e: Exception) -> str:
    """Clasifica errores por archivo para auditoría/soporte (S1-T11)."""
    msg = str(e or "").lower()
    if "no se encontró ña cabecera 'item'" in msg or "cabecera 'item'" in msg:
        return "NO_ITEM_HEADER"
    if "faltan cabeceras requeridas" in msg:
        return "MISSING_HEADERS"
    if "nombre inválido no se encontró patron oi" in msg or "patrón oi-" in msg:
        return "INVALID_OI_FILENAME"
    if "archivo vacío" in msg:
        return "EMPTY_FILE"
    return "FILE_INVALID"

def _parse_oi_tag_from_filename(filename: str) -> Optional[str]:
    m = _OI_RE.search(filename or "")
    if not m:
        return None
    return f"OI-{m.group(1)}-{m.group(2)}"

def _split_error_code_detail(raw: str) -> tuple[str | None, str | None]:
    if "·" not in raw:
        return None, raw
    code, detail = raw.split("·", 1)
    code = code.strip() or None
    detail = detail.strip() or None
    return code, detail

def _normalize_estado_literal(v: Any) -> Optional[str]:
    """
    Normaliza y valida el estado SOLO por texto literal:
    - "CONFORME"
    - "NO CONFORME"
    Sin reglas numéricas. Valores numéricos se rechazan.
    Tolera variaciones comunes: espacios dobles, guiones, puntuación, saltos de línea.
    """
    if v is None:
        return None
    # Rechazar números (0/1/2...) explícitamente: no hay reglas numéricas en LOG-01
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        return None
    
    s = _norm_str(v)
    if not s:
        return None
    
    # Quitar tíldes/diacríticos
    s = unicodedata.normalize("NFKD", s)
    s = "".join(ch for ch in s if not unicodedata.combining(ch))
    
    s = s.upper()
    s = s.replace("-", " ")
    # Mantener solo letras A-Z y espacios (eliminar puntuación y otros símbolos)
    s = re.sub(r"[^A-Z\s]+", " ", s)
    s = re.sub(r"\s+", " ", s).strip()

    if s == "CONFORME":
        return "CONFORME"
    if s == "NO CONFORME":
        return "NO CONFORME"    
    return None
        

def _natural_key(s: str):
    # Natural sort: divide digitos y texto
    return [int(t) if t.isdigit() else t.lower() for t in re.split(r'(\d+)', s)]

def _find_item_header_cell(ws: Worksheet, max_rows: int = 50, max_cols: int = 50) -> Optional[Tuple[int, int]]:
    """
    Busca la cabecera 'Item' (fila/col variable) en el rango superior del Excel.
    Retorna (row, col) de la celda donde se encontró.
    """
    r_max = min(max_rows, ws.max_row or max_rows)
    c_max = min(max_cols, ws.max_column or max_cols)
    for r in range(1, r_max + 1):
        for c in range(1, c_max + 1):
            if _norm_header(ws.cell(row=r, column=c).value) == "item":
                return (r, c)
    return None


@dataclass
class SerieInfo:
    oi_num: int
    estado: str # CONFORME / NO CONFORME
    values: Dict[str, Any]

_OUTPUT_KEYS = [
    "medidor",
    "q3",
    "error_q3",
    "q2",
    "error_q2",
    "q1",
    "error_q1",
    "estado_pe",
    "fecha",
    "certificado",
    "estado",
    "precinto",
    "banco_numero",
    "certificado_banco",
    "organismo",
]

_INPUT_HEADER_ALIASES = {
    "medidor": ["serie del medidor"],
    "q3": ["q3 litros hora"],
    "q2": ["q2 litros hora"],
    "q1": ["q1 litros hora"],
    "estado_pe": ["ensayo de presion estatica"],
    "fecha": ["fecha de ejecucion"],
    "certificado": ["numero de certificado"],
    "precinto": ["numero de serie del precinto de verificacion inicial"],
    "banco_numero": ["numero de banco de ensayo"],
    "certificado_banco": ["numero de certificado del banco de pruebas"],
    "organismo": ["organismo de inspeccion"],
}

def _emit(operation_id: Optional[str], ev: Dict[str, Any]) -> None:
    if not operation_id:
        logger.warning("LOG01 emit skipped: operation_id None")
        return
    progress_manager.emit(operation_id, ev)
    logger.debug(
        "LOG01 emit operation_id=%s ts=%s type=%s stage=%s",
        operation_id,
        time.time(),
        ev.get("type"),
        ev.get("stage"),
    )


def _copy_cell_style(src, dst) -> None:
    dst.font = copy(src.font)
    dst.fill = copy(src.fill)
    dst.border = copy(src.border)
    dst.alignment = copy(src.alignment)
    dst.number_format = src.number_format
    dst.protection = copy(src.protection)


def _apply_output_format(cell) -> None:
    cell.alignment = Alignment(horizontal="center", vertical="center")
    font = cell.font
    cell.font = Font(
        name="Arial",
        size=8,
        b=font.b,
        i=font.i,
        strike=font.strike,
        outline=font.outline,
        shadow=font.shadow,
        condense=font.condense,
        extend=font.extend,
        color=font.color,
        vertAlign=font.vertAlign,
        u=font.u,
        charset=font.charset,
        family=font.family,
        scheme=font.scheme,
    )


def _normalize_output_date(value: Any) -> Any:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        s = value.strip()
        if not s:
            return value
        try:
            return datetime.fromisoformat(s).date()
        except ValueError:
            pass
        for fmt in (
            "%d/%m/%Y",
            "%d-%m-%Y",
            "%d.%m.%Y",
            "%d/%m/%Y %H:%M:%S",
            "%d/%m/%Y %H:%M",
            "%Y-%m-%d",
            "%Y-%m-%d %H:%M:%S",
            "%Y-%m-%d %H:%M",
            "%Y/%m/%d",
            "%Y/%m/%d %H:%M:%S",
            "%Y/%m/%d %H:%M",
        ):
            try:
                return datetime.strptime(s, fmt).date()
            except ValueError:
                continue
    return value

def _normalize_input_date(value: Any, *, epoch: datetime = WINDOWS_EPOCH) -> Any:
    """
    Normaliza fechas de entrada para exportarlas como date (sin hora).
    - Soporta datetime/date
    - Soporta strings en formatos comunes.
    - Soporta serial Excel (int/float) usando epoch del workbook.
    """
    if value is None:
        return None

    # Evitar que bool (subclase de int) se trate como serial Excel
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        # 0 suele equivaler a vacío / no aplica
        if value == 0:
            return None
        try:
            dtv = from_excel(value, epoch=epoch)
            if isinstance(dtv, datetime):
                return dtv.date()
            if isinstance(dtv, date):
                return dtv
        except ValueError:
            pass

    return _normalize_output_date(value)


                

def _cleanup_log01_job_files(job: Log01Job) -> None:
    if job.work_dir and os.path.isdir(job.work_dir):
        shutil.rmtree(job.work_dir, ignore_errors=True)


def _cleanup_log01_jobs() -> None:
    now = time.time()
    stale: List[Log01Job] = []
    with LOG01_JOBS_LOCK:
        for operation_id, job in list(LOG01_JOBS.items()):
            if job.status == "running":
                continue
            if now - job.created_at < LOG01_TTL_SECONDS:
                continue
            stale.append(job)
            LOG01_JOBS.pop(operation_id, None)
    for job in stale:
        _cleanup_log01_job_files(job)


def _read_input_bytes(item: Log01InputFile) -> bytes:
    if item.data is not None:
        return item.data
    if item.path:
        with open(item.path, "rb") as fh:
            return fh.read()
    return b""


def _persist_upload_files(files: List[UploadFile], work_dir: str) -> List[Log01InputFile]:
    items: List[Log01InputFile] = []
    for idx, up in enumerate(files, start=1):
        name = up.filename or f"archivo_{idx}.xlsx"
        safe_name = Path(name).name
        dest_path = os.path.join(work_dir, f"{idx}_{safe_name}")
        with open(dest_path, "wb") as out_f:
            shutil.copyfileobj(up.file, out_f)
        up.file.close()
        items.append(Log01InputFile(name=name, path=dest_path))
    return items


def _process_log01_files(
    file_items: List[Log01InputFile],
    operation_id: Optional[str],
    output_filename: Optional[str],
    cancel_token: Optional["CancelToken"],
) -> Tuple[bytes, str, Dict[str, Any]]:
    cancel_emitted = False

    def _raise_cancelled() -> None:
        nonlocal cancel_emitted
        if cancel_token and cancel_token.is_cancelled():
            if not cancel_emitted:
                _emit(
                    operation_id,
                    {"type": "status", "stage": "cancelled", "message": "Cancelado por el usuario"},
                )
                cancel_emitted = True
            raise Log01Cancelled()

    _emit(operation_id, {"type": "status", "stage": "received", "message": "Archivos recibidos", "progress": 0})

    # 1) Consolidar serie -> (oi mayor, estado final)
    series: Dict[str, SerieInfo] = {}
    total_files = len(file_items)
    ok_files = 0
    bad_files = 0
    rows_total_read = 0

    # Auditoría "de origen" (por archivo/OI), NO depende del dedupe
    audit_by_oi: List[Dict[str, Any]] = []
    files_rejected: List[Dict[str, Any]] = []
    input_conformes_total = 0
    input_no_conformes_total = 0

    for idx, item in enumerate(file_items, start=1):
        _raise_cancelled()

        fname = item.name or f"archivo_{idx}.xlsx"
        _emit(
            operation_id,
            {
                "type": "status",
                "stage": "file_start",
                "message": f"Procesando: {fname}",
                "progress": int((idx - 1) * 100 / max(total_files, 1)),
            },
        )

        oi_num: Optional[int] = None
        try:
            oi_num = _parse_oi_number_from_filename(fname)
            data = _read_input_bytes(item)
            if not data:
                raise ValueError("EMPTY_FILE · Archivo vacío.")

            wb = load_workbook(BytesIO(data), data_only=True)
            ws: Worksheet = wb.worksheets[0]

            item_pos = _find_item_header_cell(ws)
            if item_pos is None:
                raise ValueError("NO_ITEM_HEADER · No se encontró la cabecera 'Item' en la primera hoja.")
            header_row, _item_col = item_pos


            input_header_map = {}
            for c in range(1, ws.max_column + 1):
                name = _norm_header(ws.cell(row=header_row, column=c).value)
                if name and name not in input_header_map:
                    input_header_map[name] = c

            def find_input_col(key: str) -> Optional[int]:
                for alias in _INPUT_HEADER_ALIASES.get(key, []):
                    col = input_header_map.get(_norm_header(alias))
                    if col:
                        return col
                return input_header_map.get(_norm_header(key))

            input_cols = {key: find_input_col(key) for key in _OUTPUT_KEYS}

            missing = [k for k in _OUTPUT_KEYS if not input_cols.get(k)]
            if missing:
                raise ValueError("MISSING_HEADERS · Faltan cabeceras requeridas: " + ", ".join(missing))

            # Forzar tipo (Pylance): desde aquí son int, no Optional[int]
            col_serie = cast(int, input_cols["medidor"])
            col_estado = cast(int, input_cols["estado"])
            


            data_start = header_row + 1  # si hay fila en blanco, se ignora porque serie estará vacía
            def _is_empty_cell_value(v: Any) -> bool:
                if v is None:
                    return True
                if isinstance(v, str) and not v.strip():
                    return True
                return False

            required_cols_for_empty_check: List[int] = [cast(int, input_cols[k]) for k in _OUTPUT_KEYS]
            extracted = 0
            file_conformes = 0
            file_no_conformes = 0
            for r in range(data_start, ws.max_row + 1):
                if r % 200 == 0:
                    _raise_cancelled()

                # Opción A: fin por fila completamente vacía (pero solo si ya leímos al menos 1 registro)
                row_is_empty = True
                for c in required_cols_for_empty_check:
                    if not _is_empty_cell_value(ws.cell(row=r, column=c).value):
                        row_is_empty = False
                        break

                if row_is_empty:
                    if extracted == 0:
                        continue  # tolera filas vacías iniciales
                    break         # fin real del bloque de datos

                serie = _norm_str(ws.cell(row=r, column=col_serie).value)
                if not serie:
                    continue

                estado = _normalize_estado_literal(ws.cell(row=r, column=col_estado).value)
                if not estado:
                    # si viene vacío/ruido/variación no válida, se ignora
                    continue

                if estado == "CONFORME":
                    file_conformes += 1
                else:
                    file_no_conformes += 1



                row_values: Dict[str, Any] = {"medidor": serie, "estado": estado}
                for key in _OUTPUT_KEYS:
                    if key in ("medidor", "estado"):
                        continue
                    col = input_cols.get(key)
                    row_values[key] = ws.cell(row=r, column=col).value if col else None

                # Normalizar fecha para exportar estrictamente dd/mm/yyyy (sin hora)
                row_values["fecha"] = _normalize_input_date(
                    row_values.get("fecha"),
                    epoch=wb.epoch
                )

                extracted += 1
                prev = series.get(serie)
                if prev is None or oi_num > prev.oi_num:
                    series[serie] = SerieInfo(oi_num=oi_num, estado=estado, values=row_values)

            rows_total_read += extracted
            ok_files += 1

            input_conformes_total += file_conformes
            input_no_conformes_total += file_no_conformes
            audit_by_oi.append(
                {
                    "filename": fname,
                    "oi_num": oi_num,
                    "status": "OK",
                    "rows_read": extracted,
                    "conformes": file_conformes,
                    "no_conformes": file_no_conformes,
                }
            )

            _emit(
                operation_id,
                {
                    "type": "status",
                    "stage": "file_ok",
                    "message": f"OK: {fname} (registros leídos: {extracted})",
                    "progress": int(idx * 100 / max(total_files, 1)),
                },
            )

        except Exception as e:
            bad_files += 1
            raw = str(e).strip()
            err_code, err_detail = _split_error_code_detail(raw)
            if not err_code:
                err_code = "FILE_INVALID"
            if not err_detail:
                err_detail = "Error no especificado"
            oi_tag = None
            if err_code != "INVALID_OI_FILENAME":
                oi_tag = _parse_oi_tag_from_filename(fname)
            audit_by_oi.append(
                {
                    "filename": fname,
                    "oi_num": oi_num,
                    "status": "ERROR",
                    "rows_read": 0,
                    "conformes": 0,
                    "no_conformes": 0,
                    "error": raw,
                    "error_code": err_code,
                    }
            )
            files_rejected.append(
                {
                    "filename": fname,
                    "oi": oi_tag,
                    "code": err_code,
                    "detail": err_detail,
                }
            )
            _emit(
                operation_id,
                {
                    "type": "error",
                    "stage": "file_error",
                    "message": f"Error en {fname}",
                    "detail": err_detail,
                    "code": err_code,
                },
            )
            # Importante: continuar con el lote
            continue

    # 2) Filtrar solo CONFORME
    conformes = [s for s, info in series.items() if info.estado == "CONFORME"]
    conformes.sort(key=_natural_key)

    series_total_dedup = len(series)
    series_conformes = len(conformes)
    series_no_conformes_final = series_total_dedup - series_conformes
    series_duplicates_eliminated = max(rows_total_read - series_total_dedup, 0)

    # 3) Render a plantilla LOG01 (fila 2+, item desde 1)
    st = get_settings()
    template_path = getattr(st, "log01_template_abs_path", None)
    if not template_path:
        # fallback (por si aún no se agregan property en settings)
        template_path = str((st.data_dir.parent / "app" / "data" / "templates" / "logistica" / "LOG01_PLANTILLA_SALIDA.xlsx").resolve())

    wb_out = load_workbook(template_path)
    ws_out = next(
        (ws for ws in wb_out.worksheets if ws.sheet_state == "visible"),
        None,
    )
    if ws_out is None:
        ws_out = wb_out.active
    ws_out = cast(Worksheet, ws_out)

    # Detect output columns from header row to stay aligned with template changes.
    header_map = {}
    for c in range(1, ws_out.max_column + 1):
        name = _norm_header(ws_out.cell(row=1, column=c).value)
        if name and name not in header_map:
            header_map[name] = c

    col_item = header_map.get(_norm_header("item"), 1)
    output_cols = {}
    for key in _OUTPUT_KEYS:
        col = header_map.get(_norm_header(key))
        if col is not None:
            output_cols[key] = col

    cols_to_clear = []
    for c in [col_item, *output_cols.values()]:
        if c is None or c in cols_to_clear:
            continue
        cols_to_clear.append(c)
    cols_to_fill = cols_to_clear.copy()

    # Limpieza simple: borrar desde fila 2 hacia abajo en columna A..L (mínimo)
    max_clear = max(ws_out.max_row, 2)
    for r in range(2, max_clear + 1):
        for c in cols_to_clear:
            ws_out.cell(row=r, column=c, value=None)

    for i, serie in enumerate(conformes, start=1):
        r = i + 1  # inicia en fila 2
        if r != 2:
            for c in cols_to_fill:
                src = ws_out.cell(row=2, column=c)
                dst = ws_out.cell(row=r, column=c)
                _copy_cell_style(src, dst)

        ws_out.row_dimensions[r].height = 15

        info = series[serie]
        item_cell = cast(Cell, ws_out.cell(row=r, column=col_item, value=i))  # item
        _apply_output_format(item_cell)
        for key, col in output_cols.items():
            value = info.values.get(key)
            if key == "medidor" and not value:
                value = serie
            if value is None:
                continue
            cell = cast(Cell, ws_out.cell(row=r, column=col))
            if key == "fecha":
                value = _normalize_output_date(value)
                cell.number_format = "dd/mm/yyyy"
            cell.value = value
            _apply_output_format(cell)

    bio = BytesIO()
    wb_out.save(bio)
    xlsx_bytes = bio.getvalue()

    # 4) Nombre de salida sugerido/confirmable
    if output_filename and output_filename.strip():
        out_name = output_filename.strip()
    else:
        if conformes:
            out_name = f"BD_{conformes[0]}_AL_{conformes[-1]}.xlsx"
        else:
            out_name = "BD_SIN_CONFORMES.xlsx"

    summary = {
        "files_total": total_files,
        "files_ok": ok_files,
        "files_error": bad_files,
        "series_total_dedup": series_total_dedup,
        "series_conformes": series_conformes,
        "series_no_conformes_final": series_no_conformes_final,
        # Auditoría "de origen"
        "audit_by_oi": audit_by_oi,
        "files_rejected": files_rejected,
        "totals_input": {
            "rows_read": rows_total_read,
            "conformes": input_conformes_total,
            "no_conformes": input_no_conformes_total,
        },
        # Detalle técnico
        "technical": {
            "rows_total_read": rows_total_read,
            "series_duplicates_eliminated": series_duplicates_eliminated,
        },
        # (Compatibilidad hacia atrás si ya lo consumes en UI actual)
        "series_duplicates_eliminated": series_duplicates_eliminated,
        "rows_total_read": rows_total_read,
    }

    _emit(
        operation_id,
        {"type": "complete", "message": "Consolidación completada", "percent": 100.0, "result": summary},
    )
    return xlsx_bytes, out_name, summary


def _run_log01_job(
    job: Log01Job,
    file_items: List[Log01InputFile],
    output_filename: Optional[str],
) -> None:
    operation_id = job.operation_id
    cancel_token = cancel_manager.get(operation_id)
    try:
        xlsx_bytes, out_name, _summary = _process_log01_files(
            file_items,
            operation_id,
            output_filename,
            cancel_token,
        )
        result_path = os.path.join(job.work_dir, "result.xlsx")
        with open(result_path, "wb") as out_f:
            out_f.write(xlsx_bytes)
        with LOG01_JOBS_LOCK:
            current = LOG01_JOBS.get(operation_id)
            if current:
                current.status = "complete"
                current.output_name = out_name
                current.result_path = result_path
    except Log01Cancelled:
        with LOG01_JOBS_LOCK:
            current = LOG01_JOBS.get(operation_id)
            if current:
                current.status = "cancelled"
    except Exception as exc:  # noqa: BLE001
        logger.exception("LOG01 job failed operation_id=%s", operation_id)
        _emit(
            operation_id,
            {"type": "error", "stage": "failed", "detail": str(exc), "code": "JOB_FAILED"},
        )
        with LOG01_JOBS_LOCK:
            current = LOG01_JOBS.get(operation_id)
            if current:
                current.status = "error"
                current.error = str(exc)
    finally:
        progress_manager.finish(operation_id)
        cancel_manager.remove(operation_id)

# ----------------------------
# Progreso (NDJSON stream)
# ----------------------------
@router.get("/progress/{operation_id}")
async def log01_progress_stream(operation_id: str):
    logger.info("LOG01 progress client connected operation_id=%s", operation_id)
    channel, history = progress_manager.subscribe(operation_id)

    async def event_stream():
        last_heartbeat = time.monotonic()
        try:
            # Primer evento JSON para handshake; evita carrera entre stream y start.
            hello_event = {
                "type": "hello",
                "ts": time.time(),
                "operation_id": operation_id,
                "pad": _LOG01_HELLO_PAD,
            }
            yield progress_manager.encode_event(
                {
                    **hello_event,
                }
            )
            logger.info("LOG01 progress hello sent operation_id=%s", operation_id)
            for event in history:
                logger.debug(
                    "LOG01 progress yield operation_id=%s type=%s stage=%s",
                    operation_id,
                    event.get("type"),
                    event.get("stage"),
                )
                yield progress_manager.encode_event(event)
            while True:
                try:
                    item = channel.queue.get_nowait()
                except queue.Empty:
                    now = time.monotonic()
                    if now - last_heartbeat >= 0.8:
                        yield b"\n"
                        last_heartbeat = now
                    await asyncio.sleep(0.1)
                    continue
                if item is SENTINEL:
                    break
                logger.debug(
                    "LOG01 progress yield operation_id=%s type=%s stage=%s",
                    operation_id,
                    item.get("type"),
                    item.get("stage"),
                )
                yield progress_manager.encode_event(item)
                last_heartbeat = time.monotonic()
        finally:
            logger.info("LOG01 progress client disconnected operation_id=%s", operation_id)
            progress_manager.unsubscribe(operation_id)
    headers = {
        # Streaming NDJSON: evitar buffering/transformaciones (sin gzip).
        "Cache-Control": "no-cache, no-transform",
        "Connection": "keep-alive",
        "X-Accel-Buffering": "no",
    }
    return StreamingResponse(
        event_stream(),
        media_type="application/x-ndjson; charset=utf-8",
        headers=headers,
    )


@router.get("/poll/{operation_id}")
def log01_poll(operation_id: str, cursor: int = -1):
    channel, events, cursor_next = progress_manager.get_events_since(operation_id, cursor)
    done = channel.closed
    summary = None
    if events:
        for ev in reversed(events):
            if ev.get("type") == "complete":
                summary = ev.get("result")
                break
    if summary is None and done and channel.history:
        for ev in reversed(channel.history):
            if ev.get("type") == "complete":
                summary = ev.get("result")
                break
    logger.info(
        "LOG01 poll operation_id=%s cursor=%s cursor_next=%s done=%s events=%s",
        operation_id,
        cursor,
        cursor_next,
        done,
        len(events),
    )
    return {
        "cursor_next": cursor_next,
        "events": events,
        "done": done,
        "summary": summary,
    }


@router.post("/cancel/{operation_id}")
def log01_cancel(operation_id: str):
    if not cancel_manager.cancel(operation_id):
        raise HTTPException(
            status_code=404,
            detail="Operacion no encontrada.",
            headers={"X-Code": "NOT_FOUND"},
        )
    _emit(operation_id, {"type": "status", "stage": "cancelled", "message": "Cancelado por el usuario"})
    progress_manager.finish(operation_id)
    return {"ok": True}


# ----------------------------
# Start + result (async)
# ----------------------------
@router.post("/start")
def log01_start(
    files: List[UploadFile] = File(...),
    operation_id: Optional[str] = Form(None),
    output_filename: Optional[str] = Form(None),
):
    _cleanup_log01_jobs()
    op_id = (operation_id or "").strip() or str(uuid.uuid4())
    logger.info("LOG01 start operation_id=%s", op_id)

    if not files:
        raise HTTPException(status_code=400, detail="Debes seleccionar al menos 1 Excel.")

    with LOG01_JOBS_LOCK:
        if op_id in LOG01_JOBS:
            raise HTTPException(status_code=409, detail="Operacion ya existe.")

    work_dir = tempfile.mkdtemp(prefix=f"log01_{op_id}_")
    try:
        file_items = _persist_upload_files(files, work_dir)
    except Exception:
        shutil.rmtree(work_dir, ignore_errors=True)
        raise

    cancel_manager.create(op_id)
    progress_manager.ensure(op_id)

    job = Log01Job(
        operation_id=op_id,
        created_at=time.time(),
        status="running",
        work_dir=work_dir,
    )
    with LOG01_JOBS_LOCK:
        LOG01_JOBS[op_id] = job

    thread = threading.Thread(
        target=_run_log01_job,
        args=(job, file_items, output_filename),
        daemon=True,
    )
    thread.start()
    return {"operation_id": op_id, "status": "started"}


@router.get("/result/{operation_id}")
def log01_result(operation_id: str):
    _cleanup_log01_jobs()
    with LOG01_JOBS_LOCK:
        job = LOG01_JOBS.get(operation_id)

    if job is None:
        raise HTTPException(status_code=404, detail="Resultado no encontrado.")

    if job.status == "running":
        return JSONResponse(status_code=202, content={"status": "processing"})

    if job.status == "cancelled":
        return JSONResponse(status_code=409, content={"detail": "Operacion cancelada."})

    if job.status == "error":
        return JSONResponse(status_code=409, content={"detail": job.error or "Error de procesamiento."})

    if not job.result_path or not os.path.exists(job.result_path):
        raise HTTPException(status_code=404, detail="Resultado no disponible.")

    filename = job.output_name or "BD_CONSOLIDADO.xlsx"
    headers = {
        "X-File-Name": filename,
    }
    return FileResponse(
        job.result_path,
        filename=filename,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )


# ----------------------------
# Upload + procesamiento + respuesta XLSX (sync)
# ----------------------------
@router.post("/upload")
def log01_upload(
    files: List[UploadFile] = File(...),
    operation_id: Optional[str] = Form(None),
    output_filename: Optional[str] = Form(None),
):
    logger.info("LOG01 upload operation_id=%s", operation_id)
    cancel_token = cancel_manager.create(operation_id) if operation_id else None
    if operation_id:
        progress_manager.ensure(operation_id)

    file_items: List[Log01InputFile] = []
    for idx, up in enumerate(files, start=1):
        name = up.filename or f"archivo_{idx}.xlsx"
        data = up.file.read()
        up.file.close()
        file_items.append(Log01InputFile(name=name, data=data))

    try:
        xlsx_bytes, out_name, _summary = _process_log01_files(
            file_items,
            operation_id,
            output_filename,
            cancel_token,
        )
    except Log01Cancelled:
        progress_manager.finish(operation_id)
        raise HTTPException(
            status_code=499,
            detail="Operacion cancelada por el usuario.",
            headers={"X-Code": "CANCELLED"},
        )
    finally:
        if operation_id:
            cancel_manager.remove(operation_id)

    progress_manager.finish(operation_id)

    headers = {
        "X-File-Name": out_name,
        "Content-Disposition": f'attachment; filename="{out_name}"',
    }
    return Response(
        content=xlsx_bytes,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )
