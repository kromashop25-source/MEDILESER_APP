from __future__ import annotations

from typing import Any, Dict, List, Optional, Set, Tuple
import os
import json
import csv
import io
import re
import unicodedata
import tempfile
import queue
import threading
import uuid
import time
import shutil
import errno
import random
from pathlib import Path
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from pypdf import PdfReader, PdfWriter

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from app.api.auth import get_current_user_session
from pydantic import BaseModel, Field
from app.core.settings import get_settings
from sqlmodel import Session, select
from app.core.db import engine
from app.models import Log01Artifact
from app.oi_tools.services.progress_manager import progress_manager, SENTINEL as PM_SENTINEL
from app.oi_tools.services.cancel_manager import cancel_manager, CancelToken



router = APIRouter(
    prefix="/logistica/log02",
    tags=["logistica/log02"],
    dependencies=[Depends(get_current_user_session)],

)

@router.get("/ping")
def log02_ping() -> Dict[str, Any]:
    return {"ok": True, "module": "LOG-02"}

class Log02RutaCheck(BaseModel):
    ruta: str
    existe: bool
    es_directorio: bool
    lectura: bool
    escritura: Optional[bool] = None
    detalle: Optional[str] = None

class Log02ValidarRutasUncRequest(BaseModel):
    rutas_origen: List[str] = Field(default_factory=list, description= "Lista de rutas origen (lectura)")
    ruta_destino: str = Field(..., description="Ruta destino (lesctura/escritura)")

class Log02ValidarRutasUncResponse(BaseModel):
    ok: bool
    origenes: List[Log02RutaCheck]
    destino: Log02RutaCheck

def _clean_path(value: str) -> str:
    return (value or "").strip()


def _check_read_dir(path_str: str) -> Log02RutaCheck:
    ruta = _clean_path(path_str)
    if not ruta:
        return Log02RutaCheck(
            ruta="",
            existe=False,
            es_directorio=False,
            lectura=False,
            detalle="Ruta vacia.",
        )
    try:
        p = Path(ruta)
        exists = p.exists()
        is_dir = p.is_dir() if exists else False
        if not exists:
            return Log02RutaCheck(ruta=ruta, existe=False, es_directorio=False, lectura=False, detalle="No existe.")
        if not is_dir:
            return Log02RutaCheck(ruta=ruta, existe=True, es_directorio=False, lectura=False, detalle="No es una carpeta.")
        
        # Lectura: intenta listar (puede fallar por permisos)
        try:
            os.listdir(ruta)
            return Log02RutaCheck(ruta=ruta, existe=True, es_directorio=True, lectura=True, detalle=None)
        except PermissionError:
            return Log02RutaCheck(
                ruta=ruta,
                existe=True,
                es_directorio=True,
                lectura=False,
                detalle="No tiene permisos de lectura.",
            )
        except Exception as e:
            return Log02RutaCheck(
                ruta=ruta,
                existe=True,
                es_directorio=True,
                lectura=False,
                detalle=f"No se puede leer la carpeta. {type(e).__name__}: {e}",
            )
    except Exception as e:
        return Log02RutaCheck(
            ruta=ruta,
            existe=False,
            es_directorio=False,
            lectura=False,
            detalle=f"Ruta inválida. {type(e).__name__}: {e}",
        )
    
def _check_dest_dir(path_str: str) -> Log02RutaCheck:
    base = _check_read_dir(path_str)
    # Si ya falló por no existir / no ser dir, no intentamos escribir.
    if not base.existe or not base.es_directorio:
        base.escritura = False
        return base
    
    # Escritura: crear archivos temporal y borrarl
    try:
        tmp_fd, tmp_path = tempfile.mkstemp(prefix="log02_check", suffix=".tmp", dir=base.ruta)
        try:
            with os.fdopen(tmp_fd, "wb") as f:
                f.write(b"ok")
        finally:
            try:
                os.unlink(tmp_path)
            except Exception:
                pass
        base.escritura = True
        return base
    except PermissionError:
        base.escritura = False
        base.detalle = "Sin permisos de escritura."
        return base
    except Exception as e:
        base.escritura = False
        base.detalle = f"No se pudo escribir en la carpeta. {type(e).__name__}: {e}"
        return base


@router.post("/validar-rutas-unc", response_model=Log02ValidarRutasUncResponse)
def log02_validar_rutas_unc(payload: Log02ValidarRutasUncRequest) -> Log02ValidarRutasUncResponse:
    roots_abs = _allowed_roots_abs()

    # Normalizar + limitar para evitar abuso accidental
    rutas_origen = [_clean_path(x) for x in (payload.rutas_origen or []) if _clean_path(x)]
    if len(rutas_origen) == 0:
        # Permitimos que UI valide y muestre mensaje claro desde backend también
        origenes = [Log02RutaCheck(ruta="", existe=False, es_directorio=False, lectura=False, detalle="Debe ingresar al menos una ruta de origen.")]
    else:
        if len(rutas_origen) > 20:
            rutas_origen = rutas_origen[:20]
        origenes = []
        for x in rutas_origen:
            detail = _check_allowed_detail(x, roots_abs)
            if detail:
                origenes.append(
                    Log02RutaCheck(
                        ruta=_clean_path(x),
                        existe=False,
                        es_directorio=False,
                        lectura=False,
                        detalle=detail,
                    )
                )
            else:
                origenes.append(_check_read_dir(x))
        
    dest_detail = _check_allowed_detail(payload.ruta_destino, roots_abs)
    if dest_detail:
        destino = Log02RutaCheck(
            ruta=_clean_path(payload.ruta_destino),
            existe=False,
            es_directorio=False,
            lectura=False,
            escritura=False,
            detalle=dest_detail,
        )
    else:
        destino = _check_dest_dir(payload.ruta_destino)

    ok_origen = all(o.existe and o.es_directorio and o.lectura for o in origenes) if rutas_origen else False
    ok_destino = bool(destino.existe and destino.es_directorio and destino.lectura and destino.escritura)
    ok = bool(ok_origen and ok_destino)

    return Log02ValidarRutasUncResponse(ok=ok, origenes=origenes, destino=destino)


# ====================================
# Explorador de carpetas (server-side)
# ====================================
class Log02ExplorerRootsResponse(BaseModel):
    roots: List[str]


class Log02ExplorerListItem(BaseModel):
    name: str
    path: str


class Log02ExplorerListResponse(BaseModel):
    path: str
    folders: List[Log02ExplorerListItem]


def _norm_abs(p: str) -> str:
    # Normalización simple para Windows/UNC.
    # No "resolve()" para evitar cambios extra; se apoya en normpath + abspath.
    s = (p or "").strip()
    if not s:
        return ""
    return os.path.normpath(os.path.abspath(s))


def _is_within_allowed(path_abs: str, roots_abs: List[str]) -> bool:
    if not path_abs:
        return False
    for r in roots_abs:
        try:
            common = os.path.commonpath([path_abs, r])
            if common == r:
                return True
        except Exception:
            continue
    return False


@router.get("/explorador/roots", response_model=Log02ExplorerRootsResponse)
def log02_explorer_roots() -> Log02ExplorerRootsResponse:
    settings = get_settings()
    roots = [(x or "").strip() for x in (settings.log02_unc_roots or []) if (x or "").strip()]
    return Log02ExplorerRootsResponse(roots=roots)


@router.get("/explorador/listar", response_model=Log02ExplorerListResponse)
def log02_explorer_listar(path: str = Query(..., description="Ruta absoluta dentro de raíces permitidas")) -> Log02ExplorerListResponse:
    settings = get_settings()
    roots = [(x or "").strip() for x in (settings.log02_unc_roots or []) if (x or "").strip()]
    if not roots:
        raise HTTPException(status_code=400, detail="No hay raíces configuradas para LOG-02. Configure VI_LOG02_UNC_ROOTS.")

    roots_abs = [_norm_abs(r) for r in roots]
    path_abs = _norm_abs(path)

    if not _is_within_allowed(path_abs, roots_abs):
        raise HTTPException(status_code=403, detail="Ruta fuera de las áreas permitidas (VI_LOG02_UNC_ROOTS).")
    
    p = Path(path_abs)
    if not p.exists():
        raise HTTPException(status_code=404, detail="La carpeta no existe.")
    if not p.is_dir():
        raise HTTPException(status_code=400, detail="La ruta no es una carpeta.")
    
    try:
        folders: List[Log02ExplorerListItem] = []
        with os.scandir(path_abs) as it:
            for entry in it:
                # solo en carpetas
                try:
                    if entry.is_dir(follow_symlinks=False):
                        folders.append(Log02ExplorerListItem(name=entry.name, path=os.path.join(path_abs, entry.name)))
                except Exception:
                    continue
        # orden alfabético
        folders.sort(key=lambda x: x.name.lower())
        return Log02ExplorerListResponse(path=path_abs, folders=folders)
    except PermissionError:
        raise HTTPException(status_code=403, detail="Sin permisos de lectura para listar en carpeta.")


# =============================================
# Copiado de PDFs conformes por OI (PB-LOG-015)
# =============================================

class Log02CopyConformesStartRequest(BaseModel):
    run_id: int = Field(..., description="ID de corrida de LOG-01 (historial) para usar sus artefactos (MANIFIESTO/NO_CONFORME_FINAL).")
    rutas_origen: List[str] = Field(default_factory=list, description="Rutas origen (lectura). Se buscarán carpetas OI-####-YYYY-LOTE-#### dentro de estas rutas.")
    ruta_destino: str = Field(..., description="Ruta destino (lectura/escritura). Se crearán carpetas por OI (mismo nombre exacto del lote).")
    output_mode: str = Field("keep_structure", description="Modo de salida: keep_structure (actual) o consolidate (carpeta consolidada).")
    group_size: int = Field(0, description="Tamaño N de grupo para subcarpetas en modo consolidate (0 => sin subcarpetas).")
    merge_group_size: int = Field(0, description="Tamaño N por PDF consolidado (0 => 1 PDF global con todos los certificados).")
    generate_merged_pdfs: bool = Field(False, description="En modo consolidate: genera PDFs consoldados (global si merge_group_size=0; por grupo si merge_group_size>0).")

class Log02CopyConformesStartResponse(BaseModel):
    operation_id: str

class Log02CopyConformesPollResponse(BaseModel):
    cursor_next: int = -1
    events: List[dict] = []
    done: bool = False
    summary: Optional[Any] = None

def _norm_str(v: Any) -> str:
    return ("" if v is None else str(v)).strip()

def _read_artifact_bytes(run_id: int, kind: str) -> bytes:
    """
    Lee en artefacto LOG-01 desde Log01Artifact (storage_rel_path relativo a settings.data_dir).

    """
    st = get_settings()
    with Session(engine) as session:
        art = session.exec(
            select(Log01Artifact).where(Log01Artifact.run_id == run_id, Log01Artifact.kind == kind)
        ).first()
        if not art:
            raise HTTPException(
                status_code=404,
                detail=f"No se encontró el artefacto requerido ({kind}) para la corrida {run_id}.",
            )
        abs_path = (st.data_dir / art.storage_rel_path).resolve()
        if not abs_path.exists():
            raise HTTPException(
                status_code=404,
                detail=f"El artefacto {kind} no existe en disco. Ruta: {abs_path}",
            )
        try:
            return abs_path.read_bytes()
        except Exception as e:
            raise HTTPException(
                status_code=500,
                detail=f"No se pudo leer el artefacto {kind}. {type(e).__name__}: {e}",
            )
        
def _allowed_roots_abs() -> List[str]:
    st = get_settings()
    roots = [(x or "").strip() for x in (st.log02_unc_roots or []) if (x or "").strip()]
    return [_norm_abs(r) for r in roots] if roots else []

def _check_allowed_detail(path_str: str, roots_abs: List[str]) -> Optional[str]:
    """
    Devuelve un mensaje de detalle si la ruta NO está permitida por VI_LOG02_UNC_ROOTS.
    - Si no hay roots configuradas, devolvemos detalle (para que el usuario lo corrija).
    - Si está dentro, devuelve None.
    """
    ruta = _clean_path(path_str)
    if not ruta:
        return None
    if not roots_abs:
        return "No hay raíces configuradas para LOG-02. Configure VI_LOG02_UNC_ROOTS."
    path_abs = _norm_abs(ruta)
    if not _is_within_allowed(path_abs, roots_abs):
        return "Ruta fuera de las áreas permitidas (VI_LOG02_UNC_ROOTS)."
    return None


def _ensure_within_allowed_or_400(path_abs: str, roots_abs: List[str], label: str) -> None:
    if not roots_abs:
        return # si no hay allowlist configurada, no bloqueamos (misma lógica que ya usas en roots/listar)
    p_abs = _norm_abs(path_abs)
    if not _is_within_allowed(p_abs, roots_abs):
        raise HTTPException(status_code=403, detail=f"{label}: ruta fuera de las áreas permitidas.")
    

def _emit(operation_id: str, ev: Dict[str, Any]) -> None:
    progress_manager.emit(operation_id, ev)

def _record_oi_error(
        audit: Dict[str, Any],
        operation_id: str,
        *,
        oi_tag: str,
        code: str,
        detail: str,
) -> None:
    audit["ois_error"].append({"oi": oi_tag, "code": code, "detail": detail})
    _emit(operation_id, {"type": "oi_error", "oi": oi_tag, "code": code, "message": detail})

def _record_oi_warn(
        operation_id: str,
        *,
        oi_tag: str,
        code: str,
        detail: str,
) -> None:
    _emit(operation_id, {"type": "oi_warn", "oi": oi_tag, "code": code, "message": detail})


def _cleanup_dest_folder(dest_folder: Path) -> None:
    try:
        if dest_folder.exists():
            shutil.rmtree(dest_folder)
    except Exception:
        pass

def _get_complete_audit(operation_id: str) -> Optional[Dict[str, Any]]:
    ch = progress_manager.get_channel(operation_id )
    if ch is None:
        return None
    for ev in reversed(ch.history):
        if ev.get("type") == "complete" and isinstance(ev.get("audit"), dict):
            return ev["audit"]
    return None

def _autosize_ws(ws) -> None:
    try:
        widths: Dict[int, int] = {}
        for row in ws.iter_rows(values_only=True):
            for i, v in enumerate(row, start=1):
                s = "" if v is None else str(v)
                widths[i] = min(70, max(widths.get(i, 0), len(s)))
        for i, w in widths.items():
            ws.column_dimensions[get_column_letter(i)].width = max(10, w + 2)
    except Exception:
        return
    
def _build_report_csv(audit: Dict[str, Any]) -> bytes:
    """
    CSV unificado con columna 'sección' para: resumen/por_oi/duplicados/faltantes/errores_oi
    """
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(["seccion", "oi", "serie", "campo", "valor"])

    # resumen
    w.writerow(["resumen", "", "", "run_id", audit.get("run_id", "")])
    w.writerow(["resumen", "", "", "total_ois", audit.get("total_ois", "")])
    w.writerow(["resumen", "", "", "ois_ok", audit.get("ois_ok", "")])
    arch_raw = audit.get("archivos")
    arch: Dict[str, Any] = arch_raw if isinstance(arch_raw, dict) else {}
    for k in [
        "pdf_detectados",
        "pdf_copiados",
        "pdf_omitidos_no_conforme",
        "pdf_omitidos_duplicados",
        "pdf_omitidos_no_encontrado",
        "archivos_no_pdf_omitidos",
    ]:
        w.writerow(["resumen", "", "", k, arch.get(k, "")])

    # io
    io_raw = audit.get("io")
    io_d: Dict[str, Any] = io_raw if isinstance(io_raw, dict) else {}
    for k in [
        "copy_ok",
        "copy_fail",
        "copy_attempts",
        "copy_retries",
        "copy_retryable_errors",
        "copy_locked",
        "copy_slow",
    ]:
        w.writerow(["io", "", "", k, io_d.get(k, "")])

    # por OI
    det = audit.get("detalle_por_oi")
    if isinstance(det, list):
        for r in det:
            if not isinstance(r, dict):
                continue
            oi = r.get("oi", "")
            for k, v in r.items():
                if k == "oi":
                    continue
                w.writerow(["por_oi", oi, "", k, v])

    # duplicados
    dups = audit.get("series_duplicadas")
    if isinstance(dups, list):
        for d in dups:
            if not isinstance(d, dict):
                continue
            w.writerow(["duplicados", d.get("oi", ""), d.get("serie", ""), "files", "|".join(d.get("files", []) if isinstance(d.get("files"), list) else [])])

    # faltantes (detalle completo)
    falt = audit.get("faltantes_detalle")
    if isinstance(falt, list):
        for f in falt:
            if not isinstance(f, dict):
                continue
            w.writerow(["faltantes", f.get("oi", ""), f.get("serie", ""), "missing", 1])

    # errores OI
    errs = audit.get("ois_error")
    if isinstance(errs, list):
        for e in errs:
            if not isinstance(e, dict):
                continue
            w.writerow(["errores_oi", e.get("oi",""), "", e.get("code",""), e.get("detail","")])
            
    return ("\ufeff" + buf.getvalue()).encode("utf-8")

def _build_report_xlsx(audit: Dict[str, Any]) -> bytes:
    wb = Workbook()
    ws0 = wb.active
    if ws0 is None:
        ws0 = wb.create_sheet()
    ws0.title = "Resumen"

    ws0.append(["campo", "valor"])
    ws0.append(["run_id", audit.get("run_id", "")])
    ws0.append(["total_ois", audit.get("total_ois", "")])
    ws0.append(["ois_ok", audit.get("ois_ok", "")])
    arch_raw = audit.get("archivos")
    arch: Dict[str, Any] = arch_raw if isinstance(arch_raw, dict) else {}
    for k in [
        "pdf_detectados",
        "pdf_copiados",
        "pdf_omitidos_no_conforme",
        "pdf_omitidos_duplicados",
        "pdf_omitidos_no_encontrado",
        "archivos_no_pdf_omitidos",
    ]:
        ws0.append([k, arch.get(k, "")])

    io_raw = audit.get("io")
    io_d: Dict[str, Any] = io_raw if isinstance(io_raw, dict) else {}
    ws0.append(["---", "---"])
    ws0.append(["IO", ""])
    for k in [
        "copy_ok",
        "copy_fail",
        "copy_attempts",
        "copy_retries",
        "copy_retryable_errors",
        "copy_locked",
        "copy_slow",
    ]:
        ws0.append([k, io_d.get(k, "")])
    _autosize_ws(ws0)

    ws1 = wb.create_sheet("Por OI")
    cols_oi = [
        "oi",
        "origen_folder",
        "dest_folder",
        "pdf_detectados",
        "pdf_copiados",
        "omitidos_no_conforme",
        "omitidos_duplicados",
        "no_pdf_omitidos",
        "faltantes_pdf",
        "file_errors",
        "status",
        "detail",
    ]
    ws1.append(cols_oi)
    det = audit.get("detalle_por_oi")
    if isinstance(det, list):
        for r in det:
            if not isinstance(r, dict):
                continue
            ws1.append([r.get(c, "") for c in cols_oi])
    _autosize_ws(ws1)

    ws2 = wb.create_sheet("Duplicados")
    ws2.append(["oi", "serie", "files"])
    dups = audit.get("series_duplicadas")
    if isinstance(dups, list):
        for d in dups:
            if not isinstance(d, dict):
                continue
            files = d.get("files", [])
            ws2.append([d.get("oi", ""), d.get("serie", ""), "|".join(files) if isinstance(files, list) else ""]) 
    _autosize_ws(ws2)

    ws3 = wb.create_sheet("Faltantes")
    ws3.append(["oi", "serie"])
    falt = audit.get("faltantes_detalle")
    if isinstance(falt, list):
        for f in falt:
            if not isinstance(f, dict):
                continue
            ws3.append([f.get("oi", ""), f.get("serie", "")])
    _autosize_ws(ws3)

    ws4 = wb.create_sheet("Errores OI")
    ws4.append(["oi", "code", "detail"])
    errs = audit.get("ois_error")
    if isinstance(errs, list):
        for e in errs:
            if not isinstance(e, dict):
                continue
            ws4.append([e.get("oi",""), e.get("code",""), e.get("detail","")])
    _autosize_ws(ws4)

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def _emit_progress(
        operation_id: str,
        *,
        i: int,
        total_ois: int,
        oi_tag: str,
        processed_in_oi: int | None = None,
        total_in_oi: int | None = None,
        message: str | None = None,
) -> None:
    """
    Emite progreso 0..100.
     - Si processed_in_oi/total_in_oi están presentes -> progreso fino dentro de la OI.
     - Si no -> progreso por OI (i/total_ois).
    """
    tot = max(int(total_ois), 1)
    base = (max(int(i), 1) -1) / tot # inicio de la OI actual (0..1)

    if processed_in_oi is not None and total_in_oi is not None and total_in_oi > 0:
        frac_oi = min(max(processed_in_oi / total_in_oi, 0.0), 1.0)
        pct = (base + (frac_oi / tot)) * 100.0
    else:
        pct = (max(int(i), 1) / tot) * 100.0

    msg = message
    if not msg:
        if processed_in_oi is not None and total_in_oi is not None and total_in_oi > 0:
            msg = f"{oi_tag}: {processed_in_oi}/{total_in_oi} PDFs •  {i}/{tot}"
        else:
            msg = f"Procesadas {i}/{tot} OIs"

    _emit(
        operation_id,
        {
            "type": "status",
            "stage": "progreso",
            "progress": float(round(pct, 2)),
            "percent": float(round(pct, 2)),
            "message": msg,
        },
    )

    


def _series_from_filename(name:str) -> str:
    # Serie = nombre del PDF sin extensión
    p = Path(name)
    return p.stem.strip()

_SERIE_SORT_RE = re.compile(r"^([A-Z]+)(\d+)$", re.IGNORECASE)

def _serie_sort_key(serie: str) -> Tuple[str, int, int, str]:
    """
    Orden determinista por "serie" (según nombre sin extensión).
    - Si coincide PREFIX+NUM: ordena por (PREFIX, NUM, width, raw)
    - Si no coincide: ordena por (RAW_UPPER, inf, 0, raw)
    """
    s = (serie or "").strip()
    if not s:
        return ("", 10**18, 0, "")
    m = _SERIE_SORT_RE.match(s.upper())
    if not m:
        up = s.upper()
        return (up, 10**18, 0, s)
    pref = (m.group(1) or "").upper()
    num_s = m.group(2) or "0"
    try:
        num = int(num_s)
    except Exception:
        num = 10**18
    return (pref, num, len(num_s), s)


_SERIE_RANGE_RE = re.compile(r"([A-Z]+)\s*(\d{2,})\s*(?:AL|-|A)\s*([A-Z]+)?\s*(\d{2,})", re.IGNORECASE)
_SERIE_RANGE_MAX = 200000

def _expand_series_from_text(value: str) -> List[str]:
    """
    Intenta extraer y expandir un rango de series dentro de un texto (p.ej. "PA0001 AL PA0100").
    Si no hay rango válido, devuelve [].
    """
    raw = _norm_str(value)
    if not raw:
        return []
    s = raw.replace("–", "-").replace("—", "-").upper()
    for m in _SERIE_RANGE_RE.finditer(s):
        prefix1 = (m.group(1) or "").upper()
        start_s = m.group(2) or ""
        prefix2 = (m.group(3) or prefix1).upper()
        end_s = m.group(4) or ""
        if not prefix1 or not start_s or not end_s:
            continue
        if prefix2 and prefix2 != prefix1:
            continue
        try:
            start_n = int(start_s)
            end_n = int(end_s)
        except ValueError:
            continue
        if end_n < start_n:
            start_n, end_n = end_n, start_n
        total = end_n - start_n + 1
        if total <= 0 or total > _SERIE_RANGE_MAX:
            continue
        width = max(len(start_s), len(end_s))
        return [f"{prefix1}{str(n).zfill(width)}" for n in range(start_n, end_n + 1)]
    return []

def _expand_conforme_set(conforme_set: Set[str]) -> Set[str]:
    """
    Expande entradas tipo rango (p.ej. "PA0001 AL PA0100") a series individuales.
    Si no se puede expandir, mantiene la serie original.
    """
    expanded: Set[str] = set()
    for serie in conforme_set:
        series_from_range = _expand_series_from_text(serie)
        if series_from_range:
            expanded.update(series_from_range)
        else:
            expanded.add(serie)
    return expanded

def _gaselag_key_from_name(name: str) -> str:
    """
    Normaliza nombre BD_/CD_ para match Gaselag:
    - quita prefijo BD_/CD_
    quita extensión
    elimina diacrítios
    elimina separadores (espacios, guines, puntos, etc.)
    """
    s = Path(name).stem
    s = s.strip()
    if not s:
        return ""
    s = re.sub(r"^(BD|CD)[-_\\s]+", "", s, flags=re.IGNORECASE)
    s = unicodedata.normalize("NFKD", s)
    s = "".join(ch for ch in s if not unicodedata.combining(ch))
    s = s.upper()
    s = re.sub(r"[^A-Z0-9]+", "", s)
    return s

def _gaselag_display_name(name: str) -> str:
    # Para mensajes legibles: quita BD_/CD_ y extensión, conserva separadores
    s = Path(name).stem
    s = re.sub(r"^(BD|CD)[-_\\s]+", "", s, flags=re.IGNORECASE)
    return s.strip()

def _build_gaselag_serie_map(manifest_payload: Dict[str, Any]) -> Dict[str, List[str]]:
    """
    Retorna: serie_key_normalizada -> [source_files BD_*]
    Basado en MANIFIESTO.by_oi_origen donde oi == GASELAG
    """
    out: Dict[str, List[str]] = {}
    items = manifest_payload.get("by_oi_origen")
    if not isinstance(items, list):
        return out
    for it in items:
        if not isinstance(it, dict):
            continue
        oi = _norm_str(it.get("oi")).upper()
        if oi != "GASELAG":
            continue
        files = it.get("source_files")
        if not isinstance(files, list):
            continue
        for fname in files:
            if not isinstance(fname, str):
                continue
            key = _gaselag_key_from_name(fname)
            if not key:
                continue
            out.setdefault(key, []).append(fname)
    return out

def _find_gaselag_folders_in_origins(series_keys: set[str], rutas_origen: List[str]) -> Dict[str, List[Path]]:
    """
    Busca carpetas Gaselag a 1 nivel dentro de cada origen,
    y las agrupa por clave normalizada.
    """
    found: Dict[str, List[Path]] = {k: [] for k in series_keys}
    if not series_keys:
        return found
    for root in rutas_origen:
        try:
            base = Path(root)
            if not base.exists() or not base.is_dir():
                continue
            for entry in base.iterdir():
                try:
                    if not entry.is_dir():
                        continue
                    key = _gaselag_key_from_name(entry.name)
                    if key in found:
                        found[key].append(entry)
                except Exception:
                    continue
        except Exception:
            continue
    for k in found:
        found[k].sort(key=lambda p: str(p).lower())
    return found


def _build_no_conforme_map(no_conforme_payload: Dict[str, Any]) -> Dict[str, set[str]]:
    """
    Retorna: oi_tag -> set(series_no_conforme)
    Basado en NO_CONFORME_FINAL.items: [{oi, serie, ...}]
    """
    out: Dict[str, set[str]] = {}
    items = no_conforme_payload.get("items")
    if not isinstance(items, list):
        return out
    for it in items:
        if not isinstance(it, dict):
            continue
        oi = _norm_str(it.get("oi"))
        serie = _norm_str(it.get("serie"))
        if not oi or not serie:
            continue
        out.setdefault(oi, set()).add(serie)
    return out


def _build_conforme_map(manifest_payload: Dict[str, Any]) -> Dict[str, set[str]]:
    """
    Retorna: oi_tag -> set(series_conforme)
    Basado en MANIFIESTO.by_oi: [{oi, series_conforme, ...}]
    """
    out: Dict[str, set[str]] = {}
    items = manifest_payload.get("by_oi")
    if not isinstance(items, list):
        return out
    for it in items:
        if not isinstance(it, dict):
            continue
        oi = _norm_str(it.get("oi"))
        if not oi:
            continue
        series_list = it.get("series_conforme")
        if not isinstance(series_list, list):
            continue
        series_set: set[str] = set()
        for s in series_list:
            serie = _norm_str(s)
            if serie:
                series_set.add(serie)
        out[oi] = series_set
    return out


def _is_log02_verbose() -> bool:
    v = os.getenv("VI_LOG02_VERBOSE", "").strip().lower()
    return v in ("1", "true", "yes", "on")

def _is_retryable_copy_error(e: BaseException) -> bool:
    """
    Determina si un error de I/O al copiar es razonablemente retryable:
    - Windows shares: WinError 32/33 (archivo en uso / lock)
    - PermissionError transitorio (antivirus/SMB)
    - errno EACCES/EBUSY/EPERM
    """
    if isinstance(e, PermissionError):
        return True
    if isinstance(e, OSError):
        winerr = getattr(e, "winerror", None)
        if winerr in (32, 33): # sharing violation / lock violation
            return True
        if e.errno in (errno.EACCES, errno.EBUSY, errno.EPERM):
            return True
    return False

def _sleep_ms_with_cancel(ms: int, cancel_token: CancelToken) -> None:
    """
    Sleep en chunks para respetar cancelación sin timeouts globales.
    """
    remaining = max(0, int(ms))
    step = 100 #ms
    while remaining > 0:
        if cancel_token.is_cancelled():
            return
        s = min(step, remaining) / 1000.0
        time.sleep(s)
        remaining -= step

def _copy2_atomic_with_retries(
        *,
        src_path: str | Path,
        dest_path: Path,
        cancel_token: CancelToken,
        operation_id: str,
        audit_io: Dict[str, Any],
        oi_label: str,
        serie: str,
        filename: str,
        max_attempts: int,
        base_ms: int,
        max_ms: int,
        slow_ms: int,
        verbose_events: bool,
) -> bool:
    """
    Copia con:
    - reintentos controlados en errores retryables
    - copia atómica: a .tmp + os.replace
    - medición simple de tiempo (slow copy)
    """
    dest_path.parent.mkdir(parents=True, exist_ok=True)
    tmp_name = f".tmp_{uuid.uuid4().hex}_{dest_path.name}"
    tmp_path = dest_path.parent / tmp_name

    attempt = 0
    while attempt < max_attempts:
        if cancel_token.is_cancelled():
            return False
        attempt += 1
        audit_io["copy_attempts"] += 1
        t0 = time.perf_counter()
        try:
            # Limpieza preventiva del tmp si quedó de un intento previo
            try:
                if tmp_path.exists():
                    tmp_path.unlink(missing_ok=True)
            except Exception:
                pass

            shutil.copy2(src_path, str(tmp_path))
            os.replace(str(tmp_path), str(dest_path))

            elapsed_ms = int((time.perf_counter() - t0) * 1000.0)
            audit_io["copy_ok"] += 1

            if elapsed_ms >= slow_ms:
                audit_io["copy_slow"] += 1
                if len(audit_io["slow_samples"]) < 10:
                    audit_io["slow_samples"].append(
                        {"oi": oi_label, "serie": serie, "file": filename, "ms": elapsed_ms}
                    )
                if verbose_events:
                    _emit(operation_id, {"type": "file_slow", "oi": oi_label, "serie": serie, "file": filename, "ms": elapsed_ms})

            return True
        except Exception as e:
            # Best-effort cleanup tmp
            try:
                if tmp_path.exists():
                    tmp_path.unlink(missing_ok=True)
            except Exception:
                pass
            
            retryable = _is_retryable_copy_error(e)
            if not retryable:
                audit_io["copy_fail"] += 1
                return False
            
            # Retryable
            audit_io["copy_retryable_errors"] += 1
            winerr = getattr(e, "winerror", None)
            if winerr in (32, 33): 
                audit_io["copy_locked"] += 1

            if attempt >= max_attempts:
                audit_io["copy_fail"] += 1
                return False

            # backoff con jitter, acotado
            backoff = min(max_ms, base_ms * (2 ** (attempt -1)))
            backoff = int(backoff + random.randint(0, 75))
            audit_io["copy_retries"] += 1

            if verbose_events:
                _emit(
                    operation_id,
                    {
                        "type": "file_retry",
                        "oi": oi_label,
                        "serie": serie,
                        "file": filename,
                        "attempt": attempt,
                        "wait_ms": backoff,
                        "message": f"Retry por I/O ({type(e).__name__}).",
                    },
                )
            _sleep_ms_with_cancel(backoff, cancel_token)
    audit_io["copy_fail"] += 1
    return False





def _find_oi_folders_in_origins(oi_tag: str, rutas_origen: List[str]) -> List[Path]:
    """
    Busca carpetas tipo: {oi_tag}-LOTE-#### (o variantes que empiecen con oi_tag + '-')
    solo a 1 nivel dentro de cada ruta origen.
    """
    found: List[Path] = []
    prefix = f"{oi_tag}-".lower()
    for root in rutas_origen:
        try:
            base = Path(root)
            if not base.exists() or not base.is_dir():
                continue
            for entry in base.iterdir():
                try:
                    if not entry.is_dir():
                        continue
                    name = entry.name.strip()
                    if name.lower().startswith(prefix):
                        # prioriza patrón con "LOTE"
                        found.append(entry)
                except Exception:
                    continue
        except Exception:
            continue
    # Orden determístico por ruta
    found.sort(key=lambda p: str(p).lower())
    return found


def _copy_conformes_worker(
        *,
        operation_id: str,
        cancel_token: CancelToken,
        run_id: int,
        rutas_origen: List[str],
        ruta_destino: str,
        output_mode: str = "keep_structure",
        group_size: int = 0,
        merge_group_size: int = 0,
        generate_merged_pdfs: bool = False,
) -> None:
    """
    Worker en hilo: copia PDFs conformes por OI, emitiendo progreso NDJSON.
    Reglas:
    - Match de carpeta por prefijo OI-####-YYYY (sin lote) sobre carpetas OI-####-YYYY-LOTE-####.
    - Si una OI tiene 0 carpetas => auditoría 'faltante'
    - Si una OI tiene >1 carpeta => auditoría 'duplicada' (no copiar)
    - Destino: crear carpeta con el mismo nombre exacto del lote.
      Si ya existe => auditoría 'destino duplicado' (no copiar)
    - Copia solo PDFs cuyo nombre (serie) NO esté en NO_CONFORME para esa OI.
      (Si hay otros archivos: se omiten.)
    """
    try:
        _emit(
            operation_id,
            {
                "type": "status",
                "stage": "inicio",
                "message": "Iniciando copiado de PDFs conformes...",
                "progress": 0,
                "percent": 0,
            },
        )

        # 1) Cargar artefactos (MANIFIESTO + NO_CONFORME_FINAL) desde LOG-01
        manifest_bytes = _read_artifact_bytes(run_id, "JSON_MANIFIESTO")
        no_conf_bytes = _read_artifact_bytes(run_id, "JSON_NO_CONFORME_FINAL")

        try:
            manifest = json.loads(manifest_bytes.decode("utf-8"))
        except Exception as e:
            raise RuntimeError(f"MANIFIESTO inválido. {type(e).__name__}: {e}")
        try:
            no_conf = json.loads(no_conf_bytes.decode("utf-8"))
        except Exception as e:
            raise RuntimeError(f"NO_CONFORME_FINAL inválido. {type(e).__name__}: {e}")

        by_oi = manifest.get("by_oi")
        if not isinstance(by_oi, list):
            raise RuntimeError("MANIFIESTO no contiene 'by_oi' válido.")
        
        no_conf_map = _build_no_conforme_map(no_conf)
        conformes_map = _build_conforme_map(manifest)
        use_conforme_allowlist = bool(conformes_map)
        verbose_events = _is_log02_verbose()
        
        # PB-LOG-021 knobs (Settings)
        st = get_settings()
        io_max_attempts = max(1, int(getattr(st, "log02_copy_max_attempts", 5)))
        io_base_ms = max(0, int(getattr(st, "log02_copy_retry_base_ms", 200)))
        io_max_ms = max(io_base_ms, int(getattr(st, "log02_copy_retry_max_ms", 2000)))
        io_slow_ms = max(0, int(getattr(st, "log02_copy_slow_ms", 3000)))

        gaselag_series_map = _build_gaselag_serie_map(manifest)
        gaselag_keys = sorted(gaselag_series_map.keys())

        # 2) Preparar lista de OIs BASES a procesar (excluye GASELAG)
        oi_tags: List[str] = []
        for b in by_oi:
            if not isinstance(b, dict):
                continue
            oi = _norm_str(b.get("oi"))
            if not oi:
                continue
            if oi.upper() == "GASELAG":
                continue
            oi_tags.append(oi)
        # único + orden estable
        seen: set[str] = set()
        oi_tags_u: List[str] = []
        for oi in oi_tags:
            if oi in seen:
                continue
            seen.add(oi)
            oi_tags_u.append(oi)
        oi_tags = oi_tags_u

        total_ois = len(oi_tags) + len(gaselag_keys)
        if total_ois == 0:
            raise RuntimeError("No hay OIs BASES ni GASELAG en el manifiesto (by_oi)")
        
        audit: Dict[str, Any] = {
            "run_id": run_id,
            "total_ois": total_ois,
            "ois_ok": 0,
            "detalle_por_oi": [],
            "faltantes_detalle": [],
            "series_duplicadas": [],
            "series_duplicadas_globales": [],
            "series_faltantes": [],
            "ois_faltantes": [],
            "ois_duplicadas": [],
            "destinos_duplicados": [],
            "ois_error": [],
            "archivos": {
                "pdf_detectados": 0,
                "pdf_copiados": 0,
                "pdf_omitidos_no_conforme": 0,
                "pdf_omitidos_duplicados": 0,
                "pdf_omitidos_no_encontrado": 0,  # series esperadas sin PDF
                "archivos_no_pdf_omitidos": 0,
            },
            "io": {
                "copy_ok": 0,
                "copy_fail": 0,
                "copy_attempts": 0,
                "copy_retries": 0,
                "copy_retryable_errors": 0,
                "copy_locked": 0,
                "copy_slow": 0,
                "slow_samples": [],
            },
        }
        audit_io = audit["io"]

        output_mode_norm = (output_mode or "keep_structure").strip().lower()
        if output_mode_norm not in ("keep_structure", "consolidate"):
            raise RuntimeError("output_mode inválido. Use keep_structure o consolidate.")
        if group_size is None:
            group_size = 0
        try:
            group_size_i = int(group_size)
        except Exception:
            group_size_i = 0
        if group_size_i < 0:
            group_size_i = 0
        group_size = group_size_i

        # ====================================
        # Modo consolidate: una sola carpeta
        # ====================================
        if output_mode_norm == "consolidate":
            tasks: List[Dict[str, Any]] = []
            seen_global: Dict[str, Dict[str, Any]] = {}
            dup_global: Dict[str, Set[str]] = {}

            def _mark_global_duplicate(serie: str, oi_label: str) -> None:
                prev = seen_global.get(serie)
                if prev:
                    dup_global.setdefault(serie, set()).add(str(prev.get("oi") or ""))
                dup_global.setdefault(serie, set()).add(oi_label)

            # ---- BASES ----
            for i, oi_tag in enumerate(oi_tags, start=1):
                if cancel_token.is_cancelled():
                    _emit(operation_id, {"type": "status", "stage": "cancelado", "message": "Cancelado por el usuario"})
                    return
                
                _emit(operation_id, {"type": "status", "stage": "oi", "oi": oi_tag, "message": f"Escaneando {oi_tag} ({i}/{total_ois})"})

                folders = _find_oi_folders_in_origins(oi_tag, rutas_origen)
                if len(folders) == 0:
                    audit["ois_faltantes"].append({"oi": oi_tag, "detalle": "No se encontró carpeta de lote en rutas origen."})
                    audit["detalle_por_oi"].append({
                        "oi": oi_tag,
                        "origen_folder": "",
                        "dest_folder": "",
                        "pdf_detectados": 0,
                        "pdf_copiados": 0,
                        "omitidos_no_conforme": 0,
                        "omitidos_duplicados": 0,
                        "no_pdf_omitidos": 0,
                        "faltantes_pdf": 0,
                        "file_errors": 0,
                        "status": "FALTANTE",
                        "detail": "No se encontró carpeta en orígenes",
                    })
                    _emit(operation_id, {"type": "oi_warn", "oi": oi_tag, "code": "OI_SIN_CARPETA", "message": "No se encontró carpeta para la OI en los orígenes."})
                    _emit_progress(operation_id, i=i, total_ois=total_ois, oi_tag=oi_tag)
                    continue
                if len(folders) > 1:
                    audit["ois_duplicadas"].append({"oi": oi_tag, "carpetas": [str(p) for p in folders]})
                    audit["detalle_por_oi"].append({
                        "oi": oi_tag,
                        "origen_folder": " | ".join(str(p) for p in folders),
                        "dest_folder": "",
                        "pdf_detectados": 0,
                        "pdf_copiados": 0,
                        "omitidos_no_conforme": 0,
                        "omitidos_duplicados": 0,
                        "no_pdf_omitidos": 0,
                        "faltantes_pdf": 0,
                        "file_errors": 0,
                        "status": "DUPLICADA",
                        "detail": "Múltiples carpetas en orígenes",
                    })
                    _emit(operation_id, {"type": "oi_warn", "oi": oi_tag, "code": "OI_CARPETA_DUPLICADA", "message": "Se encontraron múltiples carpetas para la misma OI. No se copiará hasta corregir."})
                    _emit_progress(operation_id, i=i, total_ois=total_ois, oi_tag=oi_tag)
                    continue

                src_folder = folders[0]
                no_conf_set = no_conf_map.get(oi_tag, set())
                raw_conforme_set = conformes_map.get(oi_tag) if use_conforme_allowlist else None
                conforme_set = _expand_conforme_set(raw_conforme_set) if raw_conforme_set is not None else None

                serie_files: Dict[str, List[str]] = {}
                series_present: Set[str] = set()
                dup_primary: Dict[str, str] = {}

                total_pdfs_in_oi = 0
                omitted_nonpdf = 0
                omitted_dup = 0
                faltantes_pdf = 0

                try:
                    with os.scandir(src_folder) as it:
                        for entry in it:
                            if cancel_token.is_cancelled():
                                _emit(operation_id, {"type": "status", "stage": "cancelado", "message": "Cancelado por el usuario"})
                                return
                            try:
                                if not entry.is_file(follow_symlinks=False):
                                    continue
                                if entry.name.lower().endswith(".pdf"):
                                    total_pdfs_in_oi += 1
                                    serie0 = _series_from_filename(entry.name)
                                    if serie0:
                                        series_present.add(serie0)
                                        serie_files.setdefault(serie0, []).append(entry.name)
                                else:
                                    omitted_nonpdf += 1
                            except Exception:
                                continue
                except PermissionError as e:
                    detail = f"Sin permisos para listar la carpeta origen. {type(e).__name__}: {e}"
                    _record_oi_error(audit, operation_id, oi_tag=oi_tag, code="LISTADO_PERMISOS", detail=detail)
                    _emit_progress(operation_id, i=i, total_ois=total_ois, oi_tag=oi_tag)
                    continue
                except Exception as e:
                    detail = f"No se pudo listar la carpeta origen. {type(e).__name__}: {e}"
                    _record_oi_error(audit, operation_id, oi_tag=oi_tag, code="LISTADO_ERROR", detail=detail)
                    _emit_progress(operation_id, i=i, total_ois=total_ois, oi_tag=oi_tag)
                    continue

                dup_series = {s: names for s, names in serie_files.items() if len(names) > 1}
                if dup_series:
                    for s, names in dup_series.items():
                        chosen = sorted(names, key=lambda x: x.lower())[0]
                        dup_primary[s] = chosen
                        audit["series_duplicadas"].append({"oi": oi_tag, "serie": s, "files": names})
                    _record_oi_warn(
                        operation_id,
                        oi_tag=oi_tag,
                        code="SERIE_DUPLICADA",
                        detail=f"Se detectaron {len(dup_series)} serie(s) duplicada(s) en la carpeta. Se copiará solo 1 PDF por serie.",

                    )
                
                if conforme_set is not None and conforme_set:
                    missing = sorted([s for s in conforme_set if s not in series_present], key=lambda x: x)
                    if missing:
                        audit["series_faltantes"].append({"oi": oi_tag, "count": len(missing), "series": missing[:50]})
                        faltantes_pdf = len(missing)
                        audit["archivos"]["pdf_omitidos_no_encontrado"] += len(missing)
                        for s in missing:
                            audit["faltantes_detalle"].append({"oi": oi_tag, "serie": s})
                        sample = ", ".join(missing[:6])
                        more = "" if len(missing) <= 6 else f" (+{len(missing) - 6} más)"
                        _record_oi_warn(
                            operation_id,
                            oi_tag=oi_tag,
                            code="SERIE_SIN_PDF",
                            detail=f"Faltan {len(missing)} serie(s) conforme(s) sin PDF en carpeta. Ej: {sample}{more}",
                        )

                # contabilidad (similar a modo actual)
                audit["archivos"]["pdf_detectados"] += total_pdfs_in_oi
                audit["archivos"]["archivos_no_pdf_omitidos"] += omitted_nonpdf

                copied_candidates = 0
                omitted_nc = 0

                for serie, names in serie_files.items():
                    # elegir primary determinístico
                    chosen_name = dup_primary.get(serie) or sorted(names, key=lambda x: x.lower())[0]

                    # filtros conforme/no_conforme
                    if conforme_set is not None:
                        if serie not in conforme_set:
                            omitted_nc += 1
                            continue
                    elif serie in no_conf_set:
                        omitted_nc += 1
                        continue
                    

                    # dedupe global
                    if serie in seen_global:
                        audit["archivos"]["pdf_omitidos_duplicados"] += 1
                        omitted_dup += 1
                        _mark_global_duplicate(serie, oi_tag)
                        continue

                    task = {
                        "serie": serie,
                        "oi": oi_tag,
                        "src": str(Path(src_folder) / chosen_name),
                        "file": chosen_name,
                    }
                    seen_global[serie] = task
                    tasks.append(task)
                    copied_candidates += 1

                audit["archivos"]["pdf_omitidos_no_conforme"] += omitted_nc
                audit["detalle_por_oi"].append({
                    "oi": oi_tag,
                    "origen_folder": str(src_folder),
                    "dest_folder": str(Path(ruta_destino)),
                    "pdf_detectados": total_pdfs_in_oi,
                    "pdf_copiados": copied_candidates,
                    "omitidos_no_conforme": omitted_nc,
                    "omitidos_duplicados": omitted_dup,
                    "no_pdf_omitidos": omitted_nonpdf,
                    "faltantes_pdf": faltantes_pdf,
                    "file_errors": 0,
                    "status": "OK",
                    "detail": "",
                })
                audit["ois_ok"] += 1
                _emit(operation_id, {"type": "oi_done", "oi": oi_tag, "copiados": copied_candidates, "omitidos_no_conforme": omitted_nc, "pdf_detectados": total_pdfs_in_oi})
                _emit_progress(operation_id, i=i, total_ois=total_ois, oi_tag=oi_tag, message=f"Escaneo {i}/{total_ois}")

            # ---- GASELAG ----
            if gaselag_keys:
                gaselag_folders = _find_gaselag_folders_in_origins(set(gaselag_keys), rutas_origen)
                base_count = len(oi_tags)
                for gi, serie_key in enumerate(gaselag_keys, start=1):
                    if cancel_token.is_cancelled():
                        _emit(operation_id, {"type": "status", "stage": "cancelado", "message": "Cancelado por el usuario"})
                        return
                    
                    i = base_count + gi
                    serie_sources = gaselag_series_map.get(serie_key, [])
                    serie_label = _gaselag_display_name(serie_sources[0]) if serie_sources else serie_key
                    oi_label = f"GASELAG-{serie_label}"
                    oi_key = "GASELAG"
                    
                    _emit(operation_id, {"type": "status", "stage": "oi", "oi": oi_label, "message": f"Escaneando {oi_label} ({i}/{total_ois})"})

                    folders = gaselag_folders.get(serie_key, [])
                    if len(folders) == 0:
                        audit["ois_faltantes"].append({"oi": oi_label, "detalle": "No se encontró carpeta de lote Gaselag en rutas origen."})
                        audit["detalle_por_oi"].append({
                            "oi": oi_label,
                            "origen_folder": "",
                            "dest_folder": "",
                            "pdf_detectados": 0,
                            "pdf_copiados": 0,
                            "omitidos_no_conforme": 0,
                            "omitidos_duplicados": 0,
                            "no_pdf_omitidos": 0,
                            "faltantes_pdf": 0,
                            "file_errors": 0,
                            "status": "FALTANTE",
                            "detail": "No se encontró carpeta en orígenes",
                        })
                        _emit(operation_id, {"type": "oi_warn", "oi": oi_label, "code": "GASELAG_SIN_CARPETA", "message": "No se encontró carpeta Gaselag para la serie en los orígenes."})
                        _emit_progress(operation_id, i=i, total_ois=total_ois, oi_tag=oi_label)
                        continue
                    if len(folders) > 1:
                        audit["ois_duplicadas"].append({"oi": oi_label, "carpetas": [str(p) for p in folders]})
                        audit["detalle_por_oi"].append({
                            "oi": oi_label,
                            "origen_folder": " | ".join(str(p) for p in folders),
                            "dest_folder": "",
                            "pdf_detectados": 0,
                            "pdf_copiados": 0,
                            "omitidos_no_conforme": 0,
                            "omitidos_duplicados": 0,
                            "no_pdf_omitidos": 0,
                            "faltantes_pdf": 0,
                            "file_errors": 0,
                            "status": "DUPLICADA",
                            "detail": "Múltiples carpetas en orígenes",
                        })
                        _emit(operation_id, {"type": "oi_warn", "oi": oi_label, "code": "GASELAG_CARPETA_DUPLICADA", "message": "Se encontraron múltiples carpetas Gaselag para la misma serie. No se copiará hasta corregir."})
                        _emit_progress(operation_id, i=i, total_ois=total_ois, oi_tag=oi_label)
                        continue

                    src_folder = folders[0]
                    no_conf_set = no_conf_map.get(oi_key, set())
                    raw_conforme_set = conformes_map.get(oi_key) if use_conforme_allowlist else None
                    conforme_set = _expand_conforme_set(raw_conforme_set) if raw_conforme_set is not None else None

                    serie_files: Dict[str, List[str]] = {}
                    series_present: Set[str] = set()
                    dup_primary: Dict[str, str] = {}
                    total_pdfs_in_oi = 0
                    omitted_nonpdf = 0
                    omitted_dup = 0
                    faltantes_pdf = 0

                    try:
                        with os.scandir(src_folder) as it:
                            for entry in it:
                                if cancel_token.is_cancelled():
                                    _emit(operation_id, {"type": "status", "stage": "cancelado", "message": "Cancelado por el usuario"})
                                    return
                                try:
                                    if not entry.is_file(follow_symlinks=False):
                                        continue
                                    if entry.name.lower().endswith(".pdf"):
                                        total_pdfs_in_oi += 1
                                        serie0 = _series_from_filename(entry.name)
                                        if serie0:
                                            series_present.add(serie0)
                                            serie_files.setdefault(serie0, []).append(entry.name)
                                    else:
                                        omitted_nonpdf += 1
                                except Exception:
                                    continue
                    except PermissionError as e:
                        detail = f"Sin permisos para listar la carpeta origen. {type(e).__name__}: {e}"
                        _record_oi_error(audit, operation_id, oi_tag=oi_label, code="LISTADO_PERMISOS", detail=detail)
                        _emit_progress(operation_id, i=i, total_ois=total_ois, oi_tag=oi_label)
                        continue
                    except Exception as e:
                        detail = f"No se pudo listar la carpeta origen. {type(e).__name__}: {e}"
                        _record_oi_error(audit, operation_id, oi_tag=oi_label, code="LISTADO_ERROR", detail=detail)
                        _emit_progress(operation_id, i=i, total_ois=total_ois, oi_tag=oi_label)
                        continue

                    dup_series = {s: names for s, names in serie_files.items() if len(names) > 1}
                    if dup_series:
                        for s, names in dup_series.items():
                            chosen = sorted(names, key=lambda x: x.lower())[0]
                            dup_primary[s] = chosen
                            audit["series_duplicadas"].append({"oi": oi_label, "serie": s, "files": names})
                        _record_oi_warn(
                            operation_id,
                            oi_tag=oi_label,
                            code="SERIE_DUPLICADA",
                            detail=f"Se detectaron {len(dup_series)} serie(s) duplicada(s) en la carpeta. Se copiará solo 1 PDF por serie.",
                        )

                    if conforme_set is not None and conforme_set:
                        missing = sorted([s for s in conforme_set if s not in series_present], key=lambda x: x)
                        if missing:
                            audit["series_faltantes"].append({"oi": oi_label, "count": len(missing), "series": missing[:50]})
                            faltantes_pdf = len(missing)
                            audit["archivos"]["pdf_omitidos_no_encontrado"] += len(missing)
                            for s in missing:
                                audit["faltantes_detalle"].append({"oi": oi_label, "serie": s})

                    audit["archivos"]["pdf_detectados"] += total_pdfs_in_oi
                    audit["archivos"]["archivos_no_pdf_omitidos"] += omitted_nonpdf

                    copied_candidates = 0
                    omitted_nc = 0

                    for serie, names in serie_files.items():
                        chosen_name = dup_primary.get(serie) or sorted(names, key=lambda x: x.lower())[0]

                        if conforme_set is not None:
                            if serie not in conforme_set:
                                omitted_nc += 1
                                continue
                        elif serie in no_conf_set:
                            omitted_nc += 1
                            continue

                        if serie in seen_global:
                            audit["archivos"]["pdf_omitidos_duplicados"] += 1
                            omitted_dup += 1
                            _mark_global_duplicate(serie, oi_label)
                            continue

                        task = {
                            "serie": serie,
                            "oi": oi_label,
                            "src": str(Path(src_folder) / chosen_name),
                            "file": chosen_name,
                        }
                        seen_global[serie] = task
                        tasks.append(task)
                        copied_candidates += 1

                    audit["archivos"]["pdf_omitidos_no_conforme"] += omitted_nc
                    audit["detalle_por_oi"].append({
                        "oi": oi_label,
                        "origen_folder": str(src_folder),
                        "dest_folder": str(Path(ruta_destino)),
                        "pdf_detectados": total_pdfs_in_oi,
                        "pdf_copiados": copied_candidates,
                        "omitidos_no_conforme": omitted_nc,
                        "omitidos_duplicados": omitted_dup,
                        "no_pdf_omitidos": omitted_nonpdf,
                        "faltantes_pdf": faltantes_pdf,
                        "file_errors": 0,
                        "status": "OK",
                        "detail": "",
                    })
                    audit["ois_ok"] += 1
                    _emit(operation_id, {"type": "oi_done", "oi": oi_label, "copiados": copied_candidates, "omitidos_no_conforme": omitted_nc, "pdf_detectados": total_pdfs_in_oi})
                    _emit_progress(operation_id, i=i, total_ois=total_ois, oi_tag=oi_label, message=f"Escaneo {i}/{total_ois}")

            if dup_global:
                out = []
                for serie, ois in dup_global.items():
                    clean = sorted([x for x in ois if x])
                    out.append({"serie": serie, "ois": clean})
                out.sort(key=lambda x: _serie_sort_key(x["serie"]))
                audit["series_duplicadas_globales"] = out

            if cancel_token.is_cancelled():
                _emit(operation_id, {"type": "status", "stage": "cancelado", "message": "Cancelado por el usuario"})
                return

            # Orden determinista por serie
            tasks.sort(key=lambda t: _serie_sort_key(str(t.get("serie") or "")))
            if not tasks:
                _emit(operation_id, {"type": "complete", "message": "No hay PDFs conformes para copiar.", "audit": audit, "percent": 100.0})
                return

            first_serie = str(tasks[0]["serie"])
            last_serie = str(tasks[-1]["serie"])
            consolidated_name = f"{first_serie}_AL_{last_serie}"
            dest_root = Path(ruta_destino) / consolidated_name

            if dest_root.exists():
                audit["destinos_duplicados"].append({"oi": "CONSOLIDADO", "destino": str(dest_root)})
                _emit(operation_id, {"type": "error", "message": f"La carpeta consolidada ya existe: {dest_root}"})
                return

            try:
                dest_root.mkdir(parents=True, exist_ok=False)
            except Exception as e:
                _emit(operation_id, {"type": "error", "message": f"No se pudo crear carpeta consolidada. {type(e).__name__}: {e}"})
                return

            copied = 0
            file_error_count = 0
            total = len(tasks)
            copied_dest_paths: List[Optional[Path]] = [None] * total
            EMIT_EVERY = 25

            def _dest_for_index(idx: int) -> Path:
                if group_size <= 0:
                    return dest_root
                g = idx // group_size
                start = g * group_size
                end = min(total - 1, (g + 1) * group_size - 1)
                s_ini = str(tasks[start]["serie"])
                s_fin = str(tasks[end]["serie"])
                sub = dest_root / f"{s_ini}_AL_{s_fin}"
                if not sub.exists():
                    try:
                        sub.mkdir(parents=True, exist_ok=True)
                    except Exception:
                        pass
                return sub

            for idx, t in enumerate(tasks):
                if cancel_token.is_cancelled():
                    _emit(operation_id, {"type": "status", "stage": "cancelado", "message": "Cancelado por el usuario"})
                    return

                src = str(t["src"])
                fname = str(t["file"])
                serie = str(t["serie"])
                oi_label = str(t["oi"])
                dest_folder = _dest_for_index(idx)
                dest_path = dest_folder / fname

                try:
                    ok = _copy2_atomic_with_retries(
                        src_path=src,
                        dest_path=dest_path,
                        cancel_token=cancel_token,
                        operation_id=operation_id,
                        audit_io=audit_io,
                        oi_label=oi_label,
                        serie=serie,
                        filename=fname,
                        max_attempts=io_max_attempts,
                        base_ms=io_base_ms,
                        max_ms=io_max_ms,
                        slow_ms=io_slow_ms,
                        verbose_events=verbose_events,
                    )
                    if ok:
                        copied += 1
                        copied_dest_paths[idx] = dest_path
                        if verbose_events:
                            _emit(operation_id, {"type": "file_ok", "oi": oi_label, "serie": serie, "file": fname})
                    else:
                        file_error_count += 1
                        _emit(operation_id, {"type": "file_error", "oi": oi_label, "serie": serie, "file": fname, "message": "Error copiando PDF (reintentos agotados o cancelado)."})
                except Exception as e:
                    file_error_count += 1
                    _emit(operation_id, {"type": "file_error", "oi": oi_label, "serie": serie, "file": fname, "message": f"Error copiando PDF. {type(e).__name__}: {e}"})

                if (idx + 1) % EMIT_EVERY == 0 or (idx + 1) == total:
                    pct = round(((idx + 1) / max(total, 1)) * 100.0, 2)
                    _emit(operation_id, {"type": "status", "stage": "copiando", "progress": pct, "percent": pct, "message": f"Copiados {idx + 1}/{total} PDFs"})

            audit["archivos"]["pdf_copiados"] += copied
            if file_error_count > 0:
                audit["ois_error"].append({"oi": "CONSOLIDADO", "code": "FILE_ERROR", "detail": f"{file_error_count} archivo(s) con error de copia."})

            # ===============================================================================
            # Opcional: generar PDFs consolidados (merge) en modo consolidate
            # - Si group_size == 0: 1 PDF global: <RANGO_GLOBAL>.pdf
            # - Si gropu_size > 0: 1 PDF por grupo: <RANGO_GRUPO>.pdf (en raíz de dest_root)
            # - (No se genera el PDF global cuando hay supcarpetas.)
            # ===============================================================================
            if generate_merged_pdfs:
                MAX_MERGE_PDFS = 3000
                try:
                    MAX_MERGE_PDFS = int(os.getenv("LOG02_MAX_MERGE_PDFS", "3000") or "3000")
                except Exception:
                    MAX_MERGE_PDFS = 3000
                merge_size = int(merge_group_size or 0)

                def _merge_pdf(output_pdf: Path, inputs: List[Path], label: str) -> Dict[str, Any]:
                    info: Dict[str, Any] = {"label": label, "output": str(output_pdf), "inputs": len(inputs), "created": False}
                    if not inputs:
                        info["skipped_reason"] = "Sin PDFs para consolidar."
                        return info
                    if output_pdf.exists():
                        info["skipped_reason"] = "El PDF consolidado ya existe (no se sobrescribe)."
                        _emit(
                            operation_id,
                            {
                                "type": "oi_warn",
                                "oi": "CONSOLIDADO",
                                "code": "MERGE_EXISTS",
                                "message": f"{label}: {info['skipped_reason']}",
                            },
                        )
                        return info
                    if len(inputs) > MAX_MERGE_PDFS:
                        info["skipped_reason"] = f"Excede el l\u00edmite de PDFs para consolidar ({len(inputs)} > {MAX_MERGE_PDFS})."
                        _emit(
                            operation_id,
                            {
                                "type": "oi_warn",
                                "oi": "CONSOLIDADO",
                                "code": "MERGE_LIMIT",
                                "message": f"{label}: {info['skipped_reason']}",
                            },
                        )
                        return info
                    
                    writer = PdfWriter()
                    skipped: List[str] = []
                    for p in inputs:
                        if cancel_token.is_cancelled():
                            info["skipped_reason"] = "Cancelado por el usuario."
                            return info
                        try:
                            reader = PdfReader(str(p))
                            for page in reader.pages:
                                writer.add_page(page)
                        except Exception as e:
                            skipped.append(str(p))
                            _emit(operation_id, {"type": "oi_warn", "oi": "CONSOLIDADO", "code": "MERGE_INPUT_ERROR", "message": f"{label}: no se pudo leer {p.name}. {type(e).__name__}: {e}"})
                            continue
                    try:
                        with open(output_pdf, "wb") as f:
                            writer.write(f)
                        info["created"] = True
                        info["skipped_inputs"] = skipped
                        _emit(operation_id, {"type": "status", "stage": "merge_pdf", "message": f"{label}: PDF consolidado creado ({output_pdf.name})."})
                    except Exception as e:
                        info["skipped_reason"] = f"No se pudo escribir PDF consolidado. {type(e).__name__}: {e}"
                        _emit(operation_id, {"type": "oi_warn", "oi": "CONSOLIDADO", "code": "MERGE_WRITE_ERROR", "message": f"{label}: {info['skipped_reason']}"})
                    return info
                
                copied_paths_all: List[Path] = [p for p in copied_dest_paths if isinstance(p, Path) and p.exists()]

                audit["pdfs_consolidados"] = {
                    "enabled": True,
                    "group_size": int(group_size or 0),
                    "merge_group_size": merge_size,
                    "max_merge_pdfs": MAX_MERGE_PDFS,
                    "global": None,
                    "groups": [],
                }

                if merge_size <= 0:
                    # PDF global en raíz de la carpeta consolidada
                    out_pdf = dest_root / f"{consolidated_name}.pdf"
                    _emit(
                        operation_id,
                        {
                            "type": "status",
                            "stage": "merge_pdf",
                            "progress": 0,
                            "percent": 0,
                            "message": "Consolidando PDF...",
                        },
                    )
                    audit["pdfs_consolidados"]["global"] = _merge_pdf(out_pdf, copied_paths_all, "GLOBAL")
                    _emit(
                        operation_id,
                        {
                            "type": "status",
                            "stage": "merge_pdf",
                            "progress": 100,
                            "percent": 100,
                            "message": "Consolidando PDF...",
                        },
                    )
                else:
                    # PDFs por grupo, ubicados en la raíz (no dentro de subcarpetas)
                    total_i = total
                    gcount = (total_i + merge_size - 1) // merge_size
                    _emit(
                        operation_id,
                        {
                            "type": "status",
                            "stage": "merge_pdf",
                            "progress": 0,
                            "percent": 0,
                            "message": "Consolidando PDF...",
                        },
                    )
                    for g in range(gcount):
                        start_i = g * merge_size
                        end_i = min(total_i - 1, (g + 1) * merge_size - 1)
                        s_ini = str(tasks[start_i]["serie"])
                        s_fin = str(tasks[end_i]["serie"])
                        label = f"GRUPO {g + 1}/{gcount}"
                        group_name = f"{s_ini}_AL_{s_fin}"
                        out_pdf = dest_root / f"{group_name}.pdf"
                        group_paths: List[Path] = []
                        for j in range(start_i, end_i + 1):
                            p = copied_dest_paths[j]
                            if isinstance(p, Path) and p.exists():
                                group_paths.append(p)
                        _emit(
                            operation_id,
                            {
                                "type": "status",
                                "stage": "merge_pdf",
                                "progress": round(((g) / max(gcount, 1)) * 100.0, 2),
                                "percent": round(((g) / max(gcount, 1)) * 100.0, 2),
                                "message": f"Consolidando PDF ({label})...",
                            },
                        )
                        audit["pdfs_consolidados"]["groups"].append(_merge_pdf(out_pdf, group_paths, label))
                        _emit(
                            operation_id,
                            {
                                "type": "status",
                                "stage": "merge_pdf",
                                "progress": round(((g + 1) / max(gcount, 1)) * 100.0, 2),
                                "percent": round(((g + 1) / max(gcount, 1)) * 100.0, 2),
                                "message": f"Consolidando PDF ({label})...",
                            },
                        )


            _emit(operation_id, {"type": "complete", "message": "Copiado finalizado.", "audit": audit, "percent": 100.0})
            return                        




        # 3) Procesar por OI (BASES)
        for i, oi_tag in enumerate(oi_tags, start=1):
            if cancel_token.is_cancelled():
                _emit(operation_id, {"type": "status", "stage": "cancelado", "message": "Cancelado por el usuario"})
                return
            
            _emit(operation_id, {"type": "status", "stage": "oi", "oi": oi_tag, "message": f"Buscando carpeta para {oi_tag} ({i}/{total_ois})"})

            folders = _find_oi_folders_in_origins(oi_tag, rutas_origen)
            if len(folders) == 0:
                audit["ois_faltantes"].append({"oi": oi_tag, "detalle": "No se encontró carpeta de lote en rutas origen."})
                audit["detalle_por_oi"].append({
                    "oi": oi_tag,
                    "origen_folder": "",
                    "dest_folder": "",
                    "pdf_detectados": 0,
                    "pdf_copiados": 0,
                    "omitidos_no_conforme": 0,
                    "omitidos_duplicados": 0,
                    "no_pdf_omitidos": 0,
                    "faltantes_pdf": 0,
                    "file_errors": 0,
                    "status": "FALTANTE",
                    "detail": "No se encontró carpeta en orígenes",
                })
                _emit(operation_id, {"type": "oi_warn", "oi": oi_tag, "code": "OI_SIN_CARPETA", "message": "No se encontró carpeta para la OI en los orígenes."})
                _emit_progress(operation_id, i=i, total_ois=total_ois, oi_tag=oi_tag)
                continue
            if len(folders) > 1:
                audit["ois_duplicadas"].append({"oi": oi_tag, "carpetas": [str(p) for p in folders]})
                audit["detalle_por_oi"].append({
                    "oi": oi_tag,
                    "origen_folder": " | ".join(str(p) for p in folders),
                    "dest_folder": "",
                    "pdf_detectados": 0,
                    "pdf_copiados": 0,
                    "omitidos_no_conforme": 0,
                    "omitidos_duplicados": 0,
                    "no_pdf_omitidos": 0,
                    "faltantes_pdf": 0,
                    "file_errors": 0,
                    "status": "DUPLICADA",
                    "detail": "Múltiples carpetas en orígenes",
                })
                _emit(operation_id, {"type": "oi_warn", "oi": oi_tag, "code": "OI_CARPETA_DUPLICADA", "message": "Se encontraron múltiples carpetas para la misma OI. No se copiará hasta corregir."})
                _emit_progress(operation_id, i=i, total_ois=total_ois, oi_tag=oi_tag)
                continue

            src_folder = folders[0]
            dest_folder = Path(ruta_destino) / src_folder.name
            if dest_folder.exists():
                audit["destinos_duplicados"].append({"oi": oi_tag, "destino": str(dest_folder)})
                audit["detalle_por_oi"].append({
                    "oi": oi_tag,
                    "origen_folder": str(src_folder),
                    "dest_folder": str(dest_folder),
                    "pdf_detectados": 0,
                    "pdf_copiados": 0,
                    "omitidos_no_conforme": 0,
                    "omitidos_duplicados": 0,
                    "no_pdf_omitidos": 0,
                    "faltantes_pdf": 0,
                    "file_errors": 0,
                    "status": "DESTINO_EXISTE",
                    "detail": "El destino ya existe",
                })
                _emit(operation_id, {"type": "oi_warn", "oi": oi_tag, "code": "DESTINO_DUPLICADO", "message": "La carpeta destino ya existe. No se copiará hasta corregir."})
                _emit_progress(operation_id, i=i, total_ois=total_ois, oi_tag=oi_tag)
                continue

            dest_created = False
            oi_ok = False
            file_error_count = 0
            copied = 0
            omitted_nc = 0
            detected_pdf = 0
            omitted_nonpdf = 0
            omitted_dup = 0
            faltantes_pdf = 0
            total_pdfs_in_oi = 0

            try:
                try:
                    dest_folder.mkdir(parents=True, exist_ok=False)
                    dest_created = True
                except FileExistsError:
                    audit["destinos_duplicados"].append({"oi": oi_tag, "destino": str(dest_folder)})
                    audit["detalle_por_oi"].append({
                        "oi": oi_tag,
                        "origen_folder": str(src_folder),
                        "dest_folder": str(dest_folder),
                        "pdf_detectados": 0,
                        "pdf_copiados": 0,
                        "omitidos_no_conforme": 0,
                        "omitidos_duplicados": 0,
                        "no_pdf_omitidos": 0,
                        "faltantes_pdf": 0,
                        "file_errors": 0,
                        "status": "DESTINO_EXISTE",
                        "detail": "El destino ya existe",
                    })
                    _emit(operation_id, {"type": "oi_warn", "oi": oi_tag, "code": "DESTINO_DUPLICADO", "message": "La carpeta destino ya existe. No se copiará hasta corregir."})
                    _emit_progress(operation_id, i=i, total_ois=total_ois, oi_tag=oi_tag)
                    continue
                except PermissionError as e:
                    detail = f"Sin permisos para crear la carpeta destino. {type(e).__name__}: {e}"
                    _record_oi_error(audit, operation_id, oi_tag=oi_tag, code="DESTINO_PERMISOS", detail=detail)
                    _emit_progress(operation_id, i=i, total_ois=total_ois, oi_tag=oi_tag)
                    continue
                except Exception as e:
                    detail = f"No se pudo crear la carpeta destino. {type(e).__name__}: {e}"
                    _record_oi_error(audit, operation_id, oi_tag=oi_tag, code="DESTINO_NO_CREADO", detail=detail)
                    _emit_progress(operation_id, i=i, total_ois=total_ois, oi_tag=oi_tag)
                    continue

                no_conf_set = no_conf_map.get(oi_tag, set())
                raw_conforme_set = conformes_map.get(oi_tag) if use_conforme_allowlist else None
                conforme_set = _expand_conforme_set(raw_conforme_set) if raw_conforme_set is not None else None

                # tracking por OI (duplicados/faltantes)
                serie_files: dict[str, list[str]] = {}
                series_present: set[str] = set()
                dup_primary: dict[str, str] = {}


                # 1) Primera pasada: contar PDFs (para poder emitir progreso fino)
                try:
                    with os.scandir(src_folder) as it:
                        for entry in it:
                            if cancel_token.is_cancelled():
                                _emit(operation_id, {"type": "status", "stage": "cancelado", "message": "Cancelado por el usuario"})
                                return
                            try:
                                if not entry.is_file(follow_symlinks=False):
                                    continue
                                if entry.name.lower().endswith(".pdf"):
                                    total_pdfs_in_oi += 1
                                    # contruir mapa serie -> archivos y set de series presentes
                                    serie0 = _series_from_filename(entry.name)
                                    if serie0:
                                        series_present.add(serie0)
                                        serie_files.setdefault(serie0, []).append(entry.name)
                            except Exception:
                                continue
                except PermissionError as e:
                    detail = f"Sin permisos para listar la carpeta origen. {type(e).__name__}: {e}"
                    _record_oi_error(audit, operation_id, oi_tag=oi_tag, code="LISTADO_PERMISOS", detail=detail)
                    _emit_progress(operation_id, i=i, total_ois=total_ois, oi_tag=oi_tag)
                    continue
                except Exception as e:
                    detail = f"No se pudo listar la carpeta origen. {type(e).__name__}: {e}"
                    _record_oi_error(audit, operation_id, oi_tag=oi_tag, code="LISTADO_ERROR", detail=detail)
                    _emit_progress(operation_id, i=i, total_ois=total_ois, oi_tag=oi_tag)
                    continue

                # Detectar duplicados (misma serie repetida en la carpeta)
                dup_series = {s: names for s, names in serie_files.items() if len(names) > 1}
                if dup_series:
                    for s, names in dup_series.items():
                        chosen = sorted(names, key=lambda x: x.lower())[0] # determinístico
                        dup_primary[s] = chosen
                        audit["series_duplicadas"].append({"oi": oi_tag, "serie": s, "files": names})
                    _record_oi_warn(
                        operation_id,
                        oi_tag=oi_tag,
                        code="SERIE_DUPLICADA",
                        detail=f"Se detectaron {len(dup_series)} serie(s) duplicada(s) en la carpeta. Se copiará solo 1 PDF por serie.",
                    )

                # Detectar faltantes SOLO si hay allowlist (MANIFIESTO -> series_conforme)
                # Nota: si no hay allowlist, no hay "lista esperada" confiable para marcar faltantes.
                if conforme_set is not None and conforme_set:
                    missing = sorted([s for s in conforme_set if s not in series_present], key=lambda x: x)
                    if missing:
                        audit["series_faltantes"].append({"oi": oi_tag, "count": len(missing), "series": missing[:50]})
                        faltantes_pdf = len(missing)
                        audit["archivos"]["pdf_omitidos_no_encontrado"] += len(missing)
                        for s in missing:
                            audit["faltantes_detalle"].append({"oi": oi_tag, "serie": s})
                        sample = ", ".join(missing[:6])
                        more = "" if len(missing) <= 6 else f" (+{len(missing) - 6} más)"
                        _record_oi_warn(
                            operation_id,
                            oi_tag=oi_tag,
                            code="SERIE_SIN_PDF",
                            detail=f"Faltan {len(missing)} serie(s) conforme(s) sin PDF en carpeta. Ej: {sample}{more}",
                        )
                        

                # Emite primer tick dentro de la OI (si hay PDFs)
                if total_pdfs_in_oi > 0:
                    _emit_progress(
                        operation_id,
                        i=i,
                        total_ois=total_ois,
                        oi_tag=oi_tag,
                        processed_in_oi=0,
                        total_in_oi=total_pdfs_in_oi,
                        message=f"{oi_tag}: iniciando copiado 0/{total_pdfs_in_oi} PDFs • OI {i}/{total_ois}",
                    )

                try:
                    processed_in_oi = 0
                    EMIT_EVERY = 25
                    with os.scandir(src_folder) as it:
                        for entry in it:
                            if cancel_token.is_cancelled():
                                _emit(operation_id, {"type": "status", "stage": "cancelado", "message": "Cancelado por el usuario"})
                                return
                            try:
                                if not entry.is_file(follow_symlinks=False):
                                    continue
                                if not entry.name.lower().endswith(".pdf"):
                                    omitted_nonpdf += 1
                                    continue
                            except Exception:
                                continue

                            detected_pdf += 1
                            processed_in_oi += 1
                            serie = _series_from_filename(entry.name)
                            if not serie:
                                omitted_nonpdf += 1
                                if total_pdfs_in_oi > 0 and (processed_in_oi % EMIT_EVERY == 0 or processed_in_oi == total_pdfs_in_oi):
                                    _emit_progress(operation_id, i=i, total_ois=total_ois, oi_tag=oi_tag, processed_in_oi=processed_in_oi, total_in_oi=total_pdfs_in_oi)
                                continue

                            # Omitir copias extra si la serie está duplicada (se copia solo el “primary”)
                            if serie in dup_primary and entry.name != dup_primary[serie]:
                                audit["archivos"]["pdf_omitidos_duplicados"] += 1
                                omitted_dup += 1
                                if verbose_events:
                                    _emit(operation_id, {"type": "file_skip", "oi": oi_tag, "serie": serie, "file": entry.name, "reason": "DUPLICADO"})
                                if total_pdfs_in_oi > 0 and (processed_in_oi % EMIT_EVERY == 0 or processed_in_oi == total_pdfs_in_oi):
                                    _emit_progress(operation_id, i=i, total_ois=total_ois, oi_tag=oi_tag, processed_in_oi=processed_in_oi, total_in_oi=total_pdfs_in_oi)
                                continue
                            
                            if conforme_set is not None:
                                if serie not in conforme_set:
                                    omitted_nc += 1
                                    if verbose_events:
                                        _emit(operation_id, {"type": "file_skip", "oi": oi_tag, "serie": serie, "reason": "NO_EN_MANIFIESTO"})
                                    if total_pdfs_in_oi > 0 and (processed_in_oi % EMIT_EVERY == 0 or processed_in_oi == total_pdfs_in_oi):
                                        _emit_progress(operation_id, i=i, total_ois=total_ois, oi_tag=oi_tag, processed_in_oi=processed_in_oi, total_in_oi=total_pdfs_in_oi)
                                    continue
                            elif serie in no_conf_set:
                                omitted_nc += 1
                                if verbose_events:
                                    _emit(operation_id, {"type": "file_skip", "oi": oi_tag, "serie": serie, "reason": "NO_CONFORME"})
                                if total_pdfs_in_oi > 0 and (processed_in_oi % EMIT_EVERY == 0 or processed_in_oi == total_pdfs_in_oi):
                                    _emit_progress(operation_id, i=i, total_ois=total_ois, oi_tag=oi_tag, processed_in_oi=processed_in_oi, total_in_oi=total_pdfs_in_oi)
                                continue

                            try:
                                dest_path = dest_folder / entry.name
                                ok = _copy2_atomic_with_retries(
                                    src_path=str(entry.path),
                                    dest_path=dest_path,
                                    cancel_token=cancel_token,
                                    operation_id=operation_id,
                                    audit_io=audit_io,
                                    oi_label=oi_tag,
                                    serie=serie,
                                    filename=entry.name,
                                    max_attempts=io_max_attempts,
                                    base_ms=io_base_ms,
                                    max_ms=io_max_ms,
                                    slow_ms=io_slow_ms,
                                    verbose_events=verbose_events,
                                )
                                if ok:
                                    copied += 1
                                    if verbose_events:
                                        _emit(operation_id, {"type": "file_ok", "oi": oi_tag, "serie": serie, "file": entry.name})
                                else:
                                    file_error_count += 1
                                    _emit(operation_id, {"type": "file_error", "oi": oi_tag, "serie": serie, "file": entry.name, "message": "Error copiando PDF (reintentos agotados o cancelado)."})

                            except Exception as e:
                                file_error_count += 1
                                _emit(operation_id, {"type": "file_error", "oi": oi_tag, "serie": serie, "file": entry.name, "message": f"Error copiando PDF. {type(e).__name__}: {e}"})
                                if total_pdfs_in_oi > 0 and (processed_in_oi % EMIT_EVERY == 0 or processed_in_oi == total_pdfs_in_oi):
                                    _emit_progress(operation_id, i=i, total_ois=total_ois, oi_tag=oi_tag, processed_in_oi=processed_in_oi, total_in_oi=total_pdfs_in_oi)
                                continue

                            if total_pdfs_in_oi > 0 and (processed_in_oi % EMIT_EVERY == 0 or processed_in_oi == total_pdfs_in_oi):
                                _emit_progress(operation_id, i=i, total_ois=total_ois, oi_tag=oi_tag, processed_in_oi=processed_in_oi, total_in_oi=total_pdfs_in_oi)
                except PermissionError as e:
                    detail = f"Sin permisos para leer la carpeta origen. {type(e).__name__}: {e}"
                    _record_oi_error(audit, operation_id, oi_tag=oi_tag, code="LECTURA_PERMISOS", detail=detail)
                    _emit_progress(operation_id, i=i, total_ois=total_ois, oi_tag=oi_tag)
                    continue
                except Exception as e:
                    detail = f"Error leyendo la carpeta origen. {type(e).__name__}: {e}"
                    _record_oi_error(audit, operation_id, oi_tag=oi_tag, code="LECTURA_ERROR", detail=detail)
                    _emit_progress(operation_id, i=i, total_ois=total_ois, oi_tag=oi_tag)
                    continue

                if file_error_count > 0:
                    detail = f"{file_error_count} archivo(s) con error de copia."
                    _record_oi_error(audit, operation_id, oi_tag=oi_tag, code="FILE_ERROR", detail=detail)
                else:
                    oi_ok = True

                audit["archivos"]["pdf_detectados"] += detected_pdf
                audit["archivos"]["pdf_copiados"] += copied
                audit["archivos"]["pdf_omitidos_no_conforme"] += omitted_nc
                audit["archivos"]["archivos_no_pdf_omitidos"] += omitted_nonpdf
                if oi_ok:
                    audit["ois_ok"] += 1

                audit["detalle_por_oi"].append({
                    "oi": oi_tag,
                    "origen_folder": str(src_folder),
                    "dest_folder": str(dest_folder),
                    "pdf_detectados": detected_pdf,
                    "pdf_copiados": copied,
                    "omitidos_no_conforme": omitted_nc,
                    "omitidos_duplicados": omitted_dup,
                    "no_pdf_omitidos": omitted_nonpdf,
                    "faltantes_pdf": faltantes_pdf,
                    "file_errors": file_error_count,
                    "status": "OK" if oi_ok else "ERROR",
                    "detail": "",
                })

                _emit(operation_id, {"type": "oi_done", "oi": oi_tag, "copiados": copied, "omitidos_no_conforme": omitted_nc, "pdf_detectados": detected_pdf})

                # Progreso por OI (cierra al valor exacto i/total_ois)
                _emit_progress(operation_id, i=i, total_ois=total_ois, oi_tag=oi_tag, message=f"Carpetas procesadas {i}/{total_ois}")
            finally:
                if dest_created and not oi_ok:
                    _cleanup_dest_folder(dest_folder)
            
        # 4) Procesar lotes GASELAG (match por serie BD_/CD_)
        if gaselag_keys:
            gaselag_folders = _find_gaselag_folders_in_origins(set(gaselag_keys), rutas_origen)
            base_count = len(oi_tags)
            for gi, serie_key in enumerate(gaselag_keys, start=1):
                if cancel_token.is_cancelled():
                    _emit(operation_id, {"type": "status", "stage": "cancelado", "message": "Cancelado por el usuario"})
                    return
                
                i = base_count + gi
                serie_sources = gaselag_series_map.get(serie_key, [])
                serie_label = _gaselag_display_name(serie_sources[0]) if serie_sources else serie_key
                oi_label = f"GASELAG:{serie_label}"
                oi_key = "GASELAG"

                _emit(operation_id, {"type": "status", "stage": "oi", "oi": oi_label, "message": f"Buscando carpeta Gaselag para {serie_label} ({i}/{total_ois})"})

                folders = gaselag_folders.get(serie_key, [])
                if len(folders) == 0:
                    audit["ois_faltantes"].append({"oi": oi_label, "detalle": "No se encontró carpeta de lote Gaselag en rutas origen."})
                    audit["detalle_por_oi"].append({
                        "oi": oi_label,
                        "origen_folder": "",
                        "dest_folder": "",
                        "pdf_detectados": 0,
                        "pdf_copiados": 0,
                        "omitidos_no_conforme": 0,
                        "omitidos_duplicados": 0,
                        "no_pdf_omitidos": 0,
                        "faltantes_pdf": 0,
                        "file_errors": 0,
                        "status": "FALTANTE",
                        "detail": "No se encontró carpeta en orígenes",
                    })
                    _emit(operation_id, {"type": "oi_warn", "oi": oi_label, "code": "GASELAG_SIN_CARPETA", "message": "No se encontró carpeta Gaselag para la serie en los orígenes."})
                    _emit_progress(operation_id, i=i, total_ois=total_ois, oi_tag=oi_label)
                    continue
                if len(folders) > 1:
                    audit["ois_duplicadas"].append({"oi": oi_label, "carpetas": [str(p) for p in folders]})
                    audit["detalle_por_oi"].append({
                        "oi": oi_label,
                        "origen_folder": " | ".join(str(p) for p in folders),
                        "dest_folder": "",
                        "pdf_detectados": 0,
                        "pdf_copiados": 0,
                        "omitidos_no_conforme": 0,
                        "omitidos_duplicados": 0,
                        "no_pdf_omitidos": 0,
                        "faltantes_pdf": 0,
                        "file_errors": 0,
                        "status": "DUPLICADA",
                        "detail": "Múltiples carpetas en orígenes",
                    })
                    _emit(operation_id, {"type": "oi_warn", "oi": oi_label, "code": "GASELAG_CARPETA_DUPLICADA", "message": "Se encontraron múltiples carpetas Gaselag para la misma serie. No se copiará hasta corregir."})
                    _emit_progress(operation_id, i=i, total_ois=total_ois, oi_tag=oi_label)
                    continue

                src_folder = folders[0]
                dest_folder = Path(ruta_destino) / src_folder.name
                if dest_folder.exists():
                    audit["destinos_duplicados"].append({"oi": oi_label, "destino": str(dest_folder)})
                    audit["detalle_por_oi"].append({
                        "oi": oi_label,
                        "origen_folder": str(src_folder),
                        "dest_folder": str(dest_folder),
                        "pdf_detectados": 0,
                        "pdf_copiados": 0,
                        "omitidos_no_conforme": 0,
                        "omitidos_duplicados": 0,
                        "no_pdf_omitidos": 0,
                        "faltantes_pdf": 0,
                        "file_errors": 0,
                        "status": "DESTINO_EXISTE",
                        "detail": "El destino ya existe",
                    })
                    _emit(operation_id, {"type": "oi_warn", "oi": oi_label, "code": "DESTINO_DUPLICADO", "message": "La carpeta destino ya existe. No se copiará hasta corregir."})
                    _emit_progress(operation_id, i=i, total_ois=total_ois, oi_tag=oi_label)
                    continue

                dest_created = False
                oi_ok = False
                file_error_count = 0
                copied = 0
                omitted_nc = 0
                detected_pdf = 0
                omitted_nonpdf = 0
                omitted_dup = 0
                faltantes_pdf = 0
                total_pdfs_in_oi = 0

                try:
                    try:
                        dest_folder.mkdir(parents=True, exist_ok=False)
                        dest_created = True
                    except FileExistsError:
                        audit["destinos_duplicados"].append({"oi": oi_label, "destino": str(dest_folder)})
                        audit["detalle_por_oi"].append({
                            "oi": oi_label,
                            "origen_folder": str(src_folder),
                            "dest_folder": str(dest_folder),
                            "pdf_detectados": 0,
                            "pdf_copiados": 0,
                            "omitidos_no_conforme": 0,
                            "omitidos_duplicados": 0,
                            "no_pdf_omitidos": 0,
                            "faltantes_pdf": 0,
                            "file_errors": 0,
                            "status": "DESTINO_EXISTE",
                            "detail": "El destino ya existe",
                        })
                        _emit(operation_id, {"type": "oi_warn", "oi": oi_label, "code": "DESTINO_DUPLICADO", "message": "La carpeta destino ya existe. No se copiará hasta corregir."})
                        _emit_progress(operation_id, i=i, total_ois=total_ois, oi_tag=oi_label)
                        continue
                    except PermissionError as e:
                        detail = f"Sin permisos para crear la carpeta destino. {type(e).__name__}: {e}"
                        _record_oi_error(audit, operation_id, oi_tag=oi_label, code="DESTINO_PERMISOS", detail=detail)
                        _emit_progress(operation_id, i=i, total_ois=total_ois, oi_tag=oi_label)
                        continue
                    except Exception as e:
                        detail = f"No se pudo crear la carpeta destino. {type(e).__name__}: {e}"
                        _record_oi_error(audit, operation_id, oi_tag=oi_label, code="DESTINO_ERROR", detail=detail)
                        _emit_progress(operation_id, i=i, total_ois=total_ois, oi_tag=oi_label)
                        continue

                    no_conf_set = no_conf_map.get(oi_key, set())
                    raw_conforme_set = conformes_map.get(oi_key) if use_conforme_allowlist else None
                    conforme_set = _expand_conforme_set(raw_conforme_set) if raw_conforme_set is not None else None

                    # Gaselag: duplicados y faltantes por serie (PDF individual)
                    serie_files: dict[str, list[str]] = {}
                    series_present: set[str] = set()
                    dup_primary: dict[str, str] = {}

                    # 1) Primera pasada: contar PDFs
                    try:
                        with os.scandir(src_folder) as it:
                            for entry in it:
                                if cancel_token.is_cancelled():
                                    _emit(operation_id, {"type": "status", "stage": "cancelado", "message": "Cancelado por el usuario"})
                                    return
                                try:
                                    if not entry.is_file(follow_symlinks=False):
                                        continue
                                    if entry.name.lower().endswith(".pdf"):
                                        total_pdfs_in_oi += 1
                                        serie0 = _series_from_filename(entry.name)
                                        if serie0:
                                            series_present.add(serie0)
                                            serie_files.setdefault(serie0, []).append(entry.name)
                                except Exception:
                                    continue
                    except PermissionError as e:
                        detail = f"Sin permisos para listar la carpeta origen. {type(e).__name__}: {e}"
                        _record_oi_error(audit, operation_id, oi_tag=oi_label, code="LISTADO_PERMISOS", detail=detail)
                        _emit_progress(operation_id, i=i, total_ois=total_ois, oi_tag=oi_label)
                        continue
                    except Exception as e:
                        detail = f"No se pudo listar la carpeta origen. {type(e).__name__}: {e}"
                        _record_oi_error(audit, operation_id, oi_tag=oi_label, code="LISTADO_ERROR", detail=detail)
                        _emit_progress(operation_id, i=i, total_ois=total_ois, oi_tag=oi_label)
                        continue

                    dup_series = {s: names for s, names in serie_files.items() if len(names) > 1}
                    if dup_series:
                        for s, names in dup_series.items():
                            chosen = sorted(names, key=lambda x: x.lower())[0]
                            dup_primary[s] = chosen
                            audit["series_duplicadas"].append({"oi": oi_label, "serie": s, "files": names})
                        _record_oi_warn(
                            operation_id,
                            oi_tag=oi_label,
                            code="SERIE_DUPLICADA",
                            detail=f"Se detectaron {len(dup_series)} serie(s) duplicada(s) en la carpeta. Se copiará solo 1 PDF por serie.",
                        )

                    # Faltantes: series esperadas (conformes) sin PDF presente en carpeta OI
                    if conforme_set is not None and conforme_set:
                        missing = sorted([s for s in conforme_set if s not in series_present], key=lambda x: x)
                        if missing:
                            audit["series_faltantes"].append({"oi": oi_label, "count": len(missing), "series": missing[:50]})
                            faltantes_pdf = len(missing)
                            audit["archivos"]["pdf_omitidos_no_encontrado"] += len(missing)
                            for s in missing:
                                audit["faltantes_detalle"].append({"oi": oi_label, "serie": s})
                            sample = ", ".join(missing[:6])
                            more = "" if len(missing) <= 6 else f" (+{len(missing) - 6} m?s)"
                            _record_oi_warn(
                                operation_id,
                                oi_tag=oi_label,
                                code="SERIE_SIN_PDF",
                                detail=f"Faltan {len(missing)} serie(s) conforme(s) sin PDF en carpeta. Ej: {sample}{more}",
                            )

                    if total_pdfs_in_oi > 0:
                        _emit_progress(
                            operation_id,
                            i=i,
                            total_ois=total_ois,
                            oi_tag=oi_label,
                            processed_in_oi=0,
                            total_in_oi=total_pdfs_in_oi,
                            message=f"{oi_label}: iniciando copiado 0/{total_pdfs_in_oi} PDFs • OI {i}/{total_ois}",
                        )
                    try:
                        processed_in_oi = 0
                        EMIT_EVERY = 25
                        with os.scandir(src_folder) as it:
                            for entry in it:
                                if cancel_token.is_cancelled():
                                    _emit(operation_id, {"type": "status", "stage": "cancelado", "message": "Cancelado por el usuario"})
                                    return
                                try:
                                    if not entry.is_file(follow_symlinks=False):
                                        continue
                                    if not entry.name.lower().endswith(".pdf"):
                                        omitted_nonpdf += 1
                                        continue
                                except Exception:
                                    continue

                                detected_pdf += 1
                                processed_in_oi += 1
                                serie = _series_from_filename(entry.name)
                                if not serie:
                                    omitted_nonpdf += 1
                                    if total_pdfs_in_oi > 0 and (processed_in_oi % EMIT_EVERY == 0 or processed_in_oi == total_pdfs_in_oi):
                                        _emit_progress(operation_id, i=i, total_ois=total_ois, oi_tag=oi_label, processed_in_oi=processed_in_oi, total_in_oi=total_pdfs_in_oi)
                                        continue

                                # Omitir copiar extra por duplicado
                                if serie in dup_primary and entry.name != dup_primary[serie]:
                                    audit["archivos"]["pdf_omitidos_duplicados"] += 1
                                    omitted_dup += 1
                                    if verbose_events:
                                        _emit(operation_id, {"type": "file_skip", "oi": oi_label, "serie": serie, "file": entry.name, "reason": "DUPLICADO"})
                                    if total_pdfs_in_oi > 0 and (processed_in_oi % EMIT_EVERY == 0 or processed_in_oi == total_pdfs_in_oi):
                                        _emit_progress(operation_id, i=i, total_ois=total_ois, oi_tag=oi_label, processed_in_oi=processed_in_oi, total_in_oi=total_pdfs_in_oi)
                                    continue
                                    

                                if conforme_set is not None:
                                    if serie not in conforme_set:
                                        omitted_nc += 1
                                        if verbose_events:
                                            _emit(operation_id, {"type": "file_skip", "oi": oi_label, "serie": serie, "reason": "NO_EN_MANIFIESTO"})
                                        if total_pdfs_in_oi > 0 and (processed_in_oi % EMIT_EVERY == 0 or processed_in_oi == total_pdfs_in_oi):
                                            _emit_progress(operation_id, i=i, total_ois=total_ois, oi_tag=oi_label, processed_in_oi=processed_in_oi, total_in_oi=total_pdfs_in_oi)
                                        continue
                                elif serie in no_conf_set:
                                    omitted_nc += 1
                                    if verbose_events:
                                        _emit(operation_id, {"type": "file_skip", "oi": oi_label, "serie": serie, "reason": "NO_CONFORME"})
                                    if total_pdfs_in_oi > 0 and (processed_in_oi % EMIT_EVERY == 0 or processed_in_oi == total_pdfs_in_oi):
                                        _emit_progress(operation_id, i=i, total_ois=total_ois, oi_tag=oi_label, processed_in_oi=processed_in_oi, total_in_oi=total_pdfs_in_oi)
                                    continue

                                try:
                                    dest_path = dest_folder / entry.name
                                    ok = _copy2_atomic_with_retries(
                                        src_path=str(entry.path),
                                        dest_path=dest_path,
                                        cancel_token=cancel_token,
                                        operation_id=operation_id,
                                        audit_io=audit_io,
                                        oi_label=oi_label,
                                        serie=serie,
                                        filename=entry.name,
                                        max_attempts=io_max_attempts,
                                        base_ms=io_base_ms,
                                        max_ms=io_max_ms,
                                        slow_ms=io_slow_ms,
                                        verbose_events=verbose_events,
                                    )
                                    if ok:
                                        copied += 1
                                        if verbose_events:
                                            _emit(operation_id, {"type": "file_ok", "oi": oi_label, "serie": serie, "file": entry.name})
                                    else:
                                        file_error_count += 1
                                        _emit(operation_id, {"type": "file_error", "oi": oi_label, "serie": serie, "file": entry.name, "message": "Error copiando PDF (reintentos agotados o cancelado)."})
                                except Exception as e:
                                    file_error_count += 1
                                    _emit(operation_id, {"type": "file_error", "oi": oi_label, "serie": serie, "file": entry.name, "message": f"Error copiando PDF. {type(e).__name__}: {e}"})
                                    if total_pdfs_in_oi > 0 and (processed_in_oi % EMIT_EVERY == 0 or processed_in_oi == total_pdfs_in_oi):
                                        _emit_progress(operation_id, i=i, total_ois=total_ois, oi_tag=oi_label, processed_in_oi=processed_in_oi, total_in_oi=total_pdfs_in_oi)
                                    continue

                                if total_pdfs_in_oi > 0 and (processed_in_oi % EMIT_EVERY == 0 or processed_in_oi == total_pdfs_in_oi):
                                    _emit_progress(operation_id, i=i, total_ois=total_ois, oi_tag=oi_label, processed_in_oi=processed_in_oi, total_in_oi=total_pdfs_in_oi)
                    except PermissionError as e:
                        detail = f"Sin permisos para leer la carpeta origen. {type(e).__name__}: {e}"
                        _record_oi_error(audit, operation_id, oi_tag=oi_label, code="LECTURA_PERMISOS", detail=detail)
                        _emit_progress(operation_id, i=i, total_ois=total_ois, oi_tag=oi_label)
                        continue
                    except Exception as e:
                        detail = f"Error leyendo la carpeta origen. {type(e).__name__}: {e}"
                        _record_oi_error(audit, operation_id, oi_tag=oi_label, code="LECTURA_ERROR", detail=detail)
                        _emit_progress(operation_id, i=i, total_ois=total_ois, oi_tag=oi_label)
                        continue

                    if file_error_count > 0:
                        detail = f"{file_error_count} archivo(s) con error de copia."
                        _record_oi_error(audit, operation_id, oi_tag=oi_label, code="FILE_ERROR", detail=detail)
                    else:
                        oi_ok = True

                    audit["archivos"]["pdf_detectados"] += detected_pdf
                    audit["archivos"]["pdf_copiados"] += copied
                    audit["archivos"]["pdf_omitidos_no_conforme"] += omitted_nc
                    audit["archivos"]["archivos_no_pdf_omitidos"] += omitted_nonpdf
                    if oi_ok:
                        audit["ois_ok"] += 1

                    audit["detalle_por_oi"].append({
                        "oi": oi_label,
                        "origen_folder": str(src_folder),
                        "dest_folder": str(dest_folder),
                        "pdf_detectados": detected_pdf,
                        "pdf_copiados": copied,
                        "omitidos_no_conforme": omitted_nc,
                        "omitidos_duplicados": omitted_dup,
                        "no_pdf_omitidos": omitted_nonpdf,
                        "faltantes_pdf": faltantes_pdf,
                        "file_errors": file_error_count,
                        "status": "OK" if oi_ok else "ERROR",
                        "detail": "",
                    })

                    _emit(operation_id, {"type": "oi_done", "oi": oi_label, "copiados": copied, "omitidos_no_conforme": omitted_nc, "pdf_detectados": detected_pdf})
                    _emit_progress(operation_id, i=i, total_ois=total_ois, oi_tag=oi_label, message=f"Carpetas procesadas {i}/{total_ois}")
                finally:
                    if dest_created and not oi_ok:
                        _cleanup_dest_folder(dest_folder)

        _emit(operation_id, {"type": "complete", "message": "Copiado finalizado.", "audit": audit, "percent": 100.0})
    except Exception as e:
        _emit(operation_id, {"type": "error", "message": f"Fallo en copiado: {type(e).__name__}: {e}"})
    finally:
        try:
            progress_manager.finish(operation_id)
        except Exception:
            pass
        try:
            cancel_manager.remove(operation_id)
        except Exception:
            pass


@router.post("/copiar-conformes/start", response_model=Log02CopyConformesStartResponse)
def log02_copiar_conformes_start(payload: Log02CopyConformesStartRequest) -> Log02CopyConformesStartResponse:
    rutas_origen = [_clean_path(x) for x in (payload.rutas_origen or []) if _clean_path(x)]
    if not rutas_origen:
        raise HTTPException(status_code=400, detail="Debe ingresar al menos una ruta de origen.")
    if len(rutas_origen) > 20:
        rutas_origen = rutas_origen[:20]

    ruta_destino = _clean_path(payload.ruta_destino)
    if not ruta_destino:
        raise HTTPException(status_code=400, detail="Debe ingresar una ruta de destino.")
    
    # Validaciones rápida de accesos (mismas reglas que S2-T07)
    origen_checks = [_check_read_dir(x) for x in rutas_origen]
    if not all(o.existe and o.es_directorio and o.lectura for o in origen_checks):
        raise HTTPException(status_code=400, detail="Una o más rutas de origen no son accesibles (lectura).")
    dest_check = _check_dest_dir(ruta_destino)
    if not (dest_check.existe and dest_check.es_directorio and dest_check.lectura and dest_check.escritura):
        raise HTTPException(status_code=400, detail="La ruta destino no es accesible (lectura/escritura).")

    # Enforce allowlist si existe
    roots_abs = _allowed_roots_abs()
    for r in rutas_origen:
        _ensure_within_allowed_or_400(r, roots_abs, "Origen")
    _ensure_within_allowed_or_400(ruta_destino, roots_abs, "Destino")

    # Validación de modo salida
    output_mode = (payload.output_mode or "keep_structure").strip().lower()
    if output_mode not in ("keep_structure", "consolidate"):
        raise HTTPException(status_code=400, detail="output_mode inválido. Use keep_structure o consolidate.")
    try:
        group_size = int(payload.group_size or 0)
    except Exception:
        group_size = 0
    if group_size < 0:
        group_size = 0

    generate_merged_pdfs = bool(getattr(payload, "generate_merged_pdfs", False))
    merge_group_raw = getattr(payload, "merge_group_size", None)
    try:
        if merge_group_raw is None:
            merge_group_size = int(group_size or 0)
        else:
            merge_group_size = int(merge_group_raw or 0)
    except Exception:
        merge_group_size = int(group_size or 0)
    if merge_group_size < 0:
        merge_group_size = 0


    operation_id = str(uuid.uuid4())
    cancel_token = cancel_manager.create(operation_id)

    # Inicializa canal y lanza hilo
    progress_manager.ensure(operation_id)
    th = threading.Thread(
        target=_copy_conformes_worker,
        kwargs={
            "operation_id": operation_id,
            "cancel_token": cancel_token,
            "run_id": int(payload.run_id),
            "rutas_origen": rutas_origen,
            "ruta_destino": ruta_destino,
            "output_mode": output_mode,
            "group_size": group_size,
            "merge_group_size": merge_group_size,
            "generate_merged_pdfs": generate_merged_pdfs,
        },
        daemon=True,
    )
    th.start()
    return Log02CopyConformesStartResponse(operation_id=operation_id)


def _ndjson_stream(operation_id: str):
    sub = progress_manager.subscribe_existing(operation_id)
    if sub is None:
        # si no existe aún, lo creamos para poder devolver algo claro
        channel, history = progress_manager.subscribe(operation_id)
    else:
        channel, history = sub

    try:
        # 1) Primer chunk inmediato para que el navegador "abra" la respuesta (evita headers provisionales)
        # Incluimos padding para atravesar posibles buffers (gzip/proxy) que esperan mínimo tamaño.
        yield progress_manager.encode_event(
            {
                "type": "hello",
                "operation_id": operation_id,
                "ts": time.time(),
                "pad": " " * 2048,
            }
        )
        # enviar historial primero
        for ev in history:
            yield progress_manager.encode_event(ev)

        # Empujón inicial para que el navegador "abra" el stream (y evitar buffering por chunks pequeños)
        yield b"\n"

        # luego stream en vivo
        while True:
            try:
                item = channel.queue.get(timeout=1.0)
            except queue.Empty:
                # heartbeat: fuerza flush/chunks en Chrome y mantiene viva la conexión
                yield progress_manager.encode_event({"type": "ping", "ts": time.time()})
                continue
            
            if item is PM_SENTINEL:
                break
            if isinstance(item, dict):
                yield progress_manager.encode_event(item)

            try:
                item = channel.queue.get(timeout=1.0)
            except queue.Empty:
                # Heartbeat: mantiene viva la conexión y fuerza flush continuo.
                yield progress_manager.encode_event({"type": "ping", "ts": time.time()})
                continue

            if item is PM_SENTINEL:
                break
            if isinstance(item, dict):
                yield progress_manager.encode_event(item)
    finally:
        try:
            progress_manager.unsubscribe(operation_id)
        except Exception:
            pass


@router.get("/copiar-conformes/progress/{operation_id}")
def log02_copiar_conformes_progress(operation_id: str):
    # Headers anti-buffering: ayudan si hay proxies/middlewares (p.ej. GZip) que agrupan chunks.
    headers = {
        "Cache-Control": "no-cache, no-store, must-revalidate, no-transform",
        "Pragma": "no-cache",
        "X-Accel-Buffering": "no",
        "Connection": "keep-alive",
        # Si existe GZipMiddleware u otro compresor, esto suele evitar que bufferice el stream
        "Content-Encoding": "identity",
    }
    return StreamingResponse(
        _ndjson_stream(operation_id),
        media_type="application/x-ndjson; charset=utf-8",
        headers=headers,
    )


@router.get("/copiar-conformes/poll/{operation_id}", response_model=Log02CopyConformesPollResponse)
def log02_copiar_conformes_poll(operation_id: str, cursor: int = -1) -> Log02CopyConformesPollResponse:
    channel, events, cursor_next = progress_manager.get_events_since(operation_id, cursor)
    done = channel.closed
    summary = None
    if events:
        for ev in reversed(events):
            if ev.get("type") == "complete":
                summary = ev.get("audit")
                break
    if summary is None and done and channel.history:
        for ev in reversed(channel.history):
            if ev.get("type") == "complete":
                summary = ev.get("audit")
                break
    return Log02CopyConformesPollResponse(cursor_next=cursor_next, events=events, done=done, summary=summary)


@router.post("/copiar-conformes/cancel/{operation_id}")
def log02_copiar_conformes_cancel(operation_id: str) -> Dict[str, Any]:
    ok = cancel_manager.cancel(operation_id)
    return {"ok": bool(ok)}

@router.get("/copiar-conformes/reporte/{operation_id}")
def log02_copiar_conformes_reporte(
    operation_id: str,
    format_: str = Query("xlsx", description="Formato de descarga: xlsx|csv", alias="format"),
) -> Any:
    audit = _get_complete_audit(operation_id)
    if audit is None:
        raise HTTPException(status_code=404, detail="Operación no encontrada o sin auditoría.")
    fmt = (format_ or "xlsx").strip().lower()
    if fmt not in ("xlsx", "csv"):
        raise HTTPException(status_code=400, detail="Formato inválido. Use xlsx o csv.")
    if fmt == "csv":
        data = _build_report_csv(audit)
        headers = {"Content-Disposition": f'attachment; filename="LOG02_AUDITORIA_{operation_id}.csv"'}
        return StreamingResponse(io.BytesIO(data), media_type="text/csv; charset=utf-8", headers=headers)
    data = _build_report_xlsx(audit)
    headers = {"Content-Disposition": f'attachment; filename="LOG02_AUDITORIA_{operation_id}.xlsx"'}
    return StreamingResponse(
        io.BytesIO(data),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )
