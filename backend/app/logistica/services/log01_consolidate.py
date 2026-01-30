from __future__ import annotations

import json
import logging
import os
import re
import time
import unicodedata
from copy import copy
from dataclasses import dataclass
from datetime import date, datetime
from io import BytesIO
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple, cast, Literal
from typing import TypedDict

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell, MergedCell
from openpyxl.styles import Alignment, Font
from openpyxl.utils.datetime import WINDOWS_EPOCH, from_excel
from openpyxl.worksheet.worksheet import Worksheet

from app.core.settings import get_settings
from app.oi_tools.services.progress_manager import progress_manager, _SENTINEL as SENTINEL
from app.oi_tools.services.cancel_manager import CancelToken

logger = logging.getLogger(__name__)


class Log01Cancelled(Exception):
    pass


@dataclass
class Log01InputFile:
    name: str
    data: Optional[bytes] = None
    path: Optional[str] = None


@dataclass
class Log01ProcessResult:
    xlsx_bytes: bytes
    out_name: str
    summary: Dict[str, Any]
    no_conforme_json: bytes
    manifest_json: bytes
    no_conforme_filename: str
    manifest_filename: str


# ----------------------------
# Utilitarios
# ----------------------------

_BASES_FILENAME_RE = re.compile(r"Base Comercial\s+OI-(\d{4})-(\d{4})", re.IGNORECASE)


def _parse_oi_parts_from_filename(filename: str) -> tuple[int, int]:
    m = _BASES_FILENAME_RE.search(filename or "")
    if not m:
        raise ValueError(
            "El nombre del archivo de Base Comercial debe incluir el patron Base Comercial OI-####-YYYY (ej: Base Comercial OI-0123-2025.xlsx)."
        )
    return int(m.group(1)), int(m.group(2))


def _parse_oi_number_from_filename(filename: str) -> int:
    oi_num, _oi_year = _parse_oi_parts_from_filename(filename)
    return oi_num


def _norm_str(v: Any) -> str:
    if v is None:
        return ""
    return str(v).strip()


def _norm_header(s: Any) -> str:
    """
    Normalización obligatoria (cabeceras):
    - None -> ""
    - str + strip
    - lower
    - quitar diacríticos (NFKD)
    - colapsar espacios
    - reemplazar . : ; , por espacio (consistente)
    """
    if s is None:
        return ""
    s = str(s).strip().lower()
    s = unicodedata.normalize("NFKD", s)
    s = "".join(ch for ch in s if not unicodedata.combining(ch))
    # Puntuación frecuente como separador
    s = re.sub(r"[.:;,]+", " ", s)
    s = re.sub(r"[\s\t\r\n]+", " ", s).strip()
    return s


def _classify_file_error(msg: str) -> str:
    msg = (msg or "").lower()
    if "invalid_oi_filename" in msg:
        return "INVALID_OI_FILENAME"
    if "missing_headers" in msg or "cabeceras requeridas" in msg:
        return "MISSING_HEADERS"
    if "archivo vacío" in msg:
        return "EMPTY_FILE"
    return "FILE_INVALID"


def _parse_oi_tag_from_filename(filename: str) -> Optional[str]:
    m = _BASES_FILENAME_RE.search(filename or "")
    if not m:
        return None
    return f"OI-{m.group(1)}-{m.group(2)}"


def _split_error_code_detail(raw: str) -> tuple[str | None, str | None]:
    if "·" not in raw:
        return None, raw
    code, detail = raw.split("·", 1)
    code = code.strip() or None
    detail = detail.strip() or None
    return code, detail


def _normalize_estado_literal(v: Any) -> Optional[str]:
    """
    Normaliza y valida el estado SOLO por texto literal:
    - "CONFORME"
    - "NO CONFORME"
    Sin reglas numéricas. Valores numéricos se rechazan.
    Regla operativa (Opción A): si el valor no normaliza, el registro se ignora.
    Tolera variaciones comunes: espacios dobles, guiones, puntuación y diacríticos.
     """
    if v is None:
        return None
    # Rechazar números (0/1/2...) explícitamente: no hay reglas numéricas en LOG-01
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        return None
    s = str(v).strip()
    if not s:
        return None

    # quitar diacríticos´
    s = unicodedata.normalize("NFKD", s)
    s = "".join(ch for ch in s if not unicodedata.combining(ch))

    s = s.upper()
    s = s.replace("-", " ")
    # eliminar puntuación y símbolos (mantener solo letras y espacios)
    s = re.sub(r"[^A-Z\s]+", " ", s)
    s = re.sub(r"\s+", " ", s).strip()

    if s == "CONFORME":
        return "CONFORME"
    if s == "NO CONFORME":
        return "NO CONFORME"
    return None


def _natural_key(s: Any) -> list:
    s = str(s or "")
    # split por dígitos para ordenar series de manera humana
    return [int(t) if t.isdigit() else t.lower() for t in re.split(r"(\d+)", s)]


def _oi_compare_key(oi_year: Optional[int], oi_num: Optional[int]) -> tuple[int, int]:
    return (int(oi_year or 0), int(oi_num or 0))


def _find_item_header_cell(ws: Worksheet) -> Optional[tuple[int, int]]:
    # busca "Item" normalizado, en un rango razonable
    r_max = min(ws.max_row or 1, 200)
    c_max = min(ws.max_column or 1, 80)
    for r in range(1, r_max + 1):
        for c in range(1, c_max + 1):
            if _norm_header(ws.cell(row=r, column=c).value) == "item":
                return (r, c)
    return None


@dataclass
class SerieInfo:
    oi_num: int
    oi_year: int
    estado: str  # CONFORME / NO CONFORME
    values: Dict[str, Any]


_OUTPUT_KEYS = [
    "medidor",
    "q3",
    "error_q3",
    "q2",
    "error_q2",
    "q1",
    "error_q1",
    "estado_pe",
    "fecha",
    "certificado",
    "estado",
    "precinto",
    "banco_numero",
    "certificado_banco",
    "organismo",
]

# Cabeceras requeridas BASES (por nombre lógico) y aliases aceptados (normalizados)
_REQUIRED_INPUT_KEYS_BASES = [
    "medidor",
    "q3",
    "error_q3",
    "q2",
    "error_q2",
    "q1",
    "error_q1",
    "estado_pe",
    "fecha",
    "certificado",
    "estado",
    "precinto",
    "banco_numero",
    "certificado_banco",
    "organismo",
]

_INPUT_HEADER_ALIASES_BASES: Dict[str, List[str]] = {
    # Bases comerciales (cerrado)
    "medidor": ["serie del medidor"],
    "q3": ["q3 (litros/hora)"],
    "error_q3": ["error q3 (%)"],
    "q2": ["q2 (litros/hora)"],
    "error_q2": ["error q2 (%)"],
    "q1": ["q1 (litros/hora)"],
    "error_q1": ["error q1 (%)"],
    "estado_pe": ["ensayo de presion estatica"],
    "fecha": ["fecha de ejecucion"],
    "certificado": ["numero de certificado"],
    "estado": ["estado"],
    "precinto": ["numero de serie del precinto de verificacion inicial"],
    "banco_numero": ["numero de banco de ensayo"],
    "certificado_banco": ["numero de certificado del banco de pruebas"],
    "organismo": ["organismo de inspeccion"],
}

_REQUIRED_INPUT_KEYS_GASELAG_V1 = [
    # organismo es fijo OI-066 (NO viene del archivo)
    "medidor",
    "precinto",
    "fecha",
    "banco_numero",
    "certificado",
    "estado_pe",
    "q3",
    "error_q3",
    "q2",
    "error_q2",
    "q1",
    "error_q1",
    "estado",
    "certificado_banco",
]

_REQUIRED_INPUT_KEYS_GASELAG_V2 = [
    "medidor",
    "q3",
    "error_q3",
    "q2",
    "error_q2",
    "q1",
    "error_q1",
    "estado_pe",
    "fecha",
    "certificado",
    "estado",
    "precinto",
    "banco_numero",
    "certificado_banco",
    "organismo",
]

_INPUT_HEADER_ALIASES_GASELAG: Dict[str, List[str]] = {
    # GASELAG (cerrado)
    "medidor": ["nro serie", "nro. serie", "numero serie", "número serie", "número de serie"],
    "precinto": ["precinto"],
    "fecha": ["fecha de ensayo presion estatica"],
    "banco_numero": ["banco ensayo errores de indicacion"],
    "certificado": ["cert verificacion inicial", "cert. verificacion inicial"],
    "estado_pe": ["resultado de p estatica", "resultado de p. estatica", "resultado de p. estática"],
    "q3": ["q3 (l/h)", "q3 (l h)"],
    "error_q3": ["error q3 (%)", "error q3 (%) "],
    "q2": ["q2 (l/h)", "q2 (l h)"],
    "error_q2": ["error q2 (%)", "error q2 (%) "],
    "q1": ["q1 (l/h)", "q1 (l h)"],
    "error_q1": ["error q1 (%)", "error q1 (%) "],
    "estado": ["conclusion", "conclusión"],
    "certificado_banco": ["certificado banco", "certificado  banco"],
}



def _emit(operation_id: Optional[str], ev: Dict[str, Any]) -> None:
    if not operation_id:
        logger.warning("LOG01 emit skipped: operation_id None")
        return
    progress_manager.emit(operation_id, ev)
    logger.debug("LOG01 emit operation_id=%s ev=%s", operation_id, ev)


def _copy_cell_style(src: Cell, dst: Cell) -> None:
    dst.font = src.font.copy()
    dst.fill = src.fill.copy()
    dst.border = src.border.copy()
    dst.alignment = src.alignment.copy()
    dst.number_format = src.number_format
    dst.protection = src.protection.copy()
    dst.comment = None


def _writable_cell(ws: Worksheet, row: int, col: int) -> Cell:
    return cast(Cell, ws.cell(row=row, column=col))


def _apply_output_format(ws_out: Worksheet, row: int, col: int, key: str) -> None:
    cell = ws_out.cell(row=row, column=col)
    # Formatos mínimos consistentes
    cell.font = Font(name="Arial", size=8)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Formato fecha dd/mm/yyyy si aplica (la plantilla puede tenerlo, pero aseguramos consistencia)
    if key == "fecha":
        cell.number_format = "dd/mm/yyyy"


def _normalize_output_date(d: Any) -> Optional[date]:
    if d is None or d is SENTINEL:
        return None
    if isinstance(d, date) and not isinstance(d, datetime):
        return d
    if isinstance(d, datetime):
        return d.date()
    # strings dd/mm/yyyy | yyyy/mm/dd | etc
    s = str(d).strip()
    if not s:
        return None
    # intentar dd/mm/yyyy
    for fmt in ("%d/%m/%Y", "%Y/%m/%d", "%Y-%m-%d", "%d-%m-%Y", "%d.%m.%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except Exception:
            pass
    return None


def _normalize_input_date(v: Any, epoch: date) -> Optional[date]:
    if v is None:
        return None
    if isinstance(v, date) and not isinstance(v, datetime):
        return v
    if isinstance(v, datetime):
        return v.date()

    # openpyxl serial date
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        try:
            dt = from_excel(v, epoch=epoch)
            if isinstance(dt, datetime):
                return dt.date()
            if isinstance(dt, date):
                return dt
        except Exception:
            return None

    s = str(v).strip()
    if not s:
        return None
    for fmt in ("%d/%m/%Y", "%Y/%m/%d", "%Y-%m-%d", "%d-%m-%Y", "%d.%m.%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except Exception:
            pass
    return None


def _read_input_bytes(item: Log01InputFile) -> bytes:
    if item.data is not None:
        return item.data
    if item.path:
        with open(item.path, "rb") as f:
            return f.read()
    return b""


def process_log01_files(
    file_items: List[Log01InputFile],
    operation_id: Optional[str],
    output_filename: Optional[str],
    cancel_token: Optional["CancelToken"],
    source: Literal["AUTO","BASES", "GASELAG"] = "AUTO",
) -> Log01ProcessResult:
    cancel_emitted = False

    def _raise_cancelled() -> None:
        nonlocal cancel_emitted
        if cancel_token and cancel_token.is_cancelled():
            if not cancel_emitted:
                _emit(
                    operation_id,
                    {"type": "status", "stage": "cancelled", "message": "Cancelado por el usuario"},
                )
                cancel_emitted = True
            raise Log01Cancelled()

    _emit(operation_id, {"type": "status", "stage": "received", "message": "Archivos recibidos", "progress": 0})

    # 1) Consolidar serie -> (estado final por OI mayor; CONFORME prevalece si existe)
    series: Dict[str, SerieInfo] = {}
    series_meta: Dict[str, Dict[str, Any]] = {}
    total_files = len(file_items)
    ok_files = 0
    bad_files = 0
    rows_total_read = 0

    # Auditoría "de origen" (por archivo/OI), NO depende del dedupe
    audit_by_oi: List[Dict[str, Any]] = []
    files_rejected: List[Dict[str, Any]] = []
    input_conformes_total = 0
    input_no_conformes_total = 0

    def _new_source_bucket() ->  Dict[str, Any]:
        return {
            "files_total": 0,
            "files_ok": 0,
            "files_error": 0,
            "rows_read": 0,
            "conformes": 0,
            "no_conformes": 0,
            "rows_ignored_invalid_estado": 0, 
        }
    
    # Auditoría agregada por tipo de origen
    by_source: Dict[str, Dict[str, Any]] = {
        "BASES": _new_source_bucket(),
        "GASELAG": _new_source_bucket(),
    }

    for idx, item in enumerate(file_items, start=1):
        _raise_cancelled()

        fname = item.name or f"archivo_{idx}.xlsx"
        bases_filename = bool(_BASES_FILENAME_RE.search(fname or ""))
        _emit(
            operation_id,
            {
                "type": "status",
                "stage": "file",
                "message": f"Procesando {fname}",
                "file": fname,
                "index": idx,
                "total": total_files,
                "progress": float((idx - 1) * 100 / max(total_files, 1)),
            },
        )

        oi_num: Optional[int] = None
        oi_year: Optional[int] = None
        oi_tag: Optional[str] = None
        source_type: str = "AUTO"
        try:
            data = _read_input_bytes(item)
            if not data:
                raise ValueError("El archivo está vacío o no se pudo leer.")

            wb = load_workbook(filename=BytesIO(data), data_only=True, read_only=True)
            ws = wb.worksheets[0]

            item_pos = _find_item_header_cell(ws)
            if not item_pos:
                raise ValueError(
                    "No se encontró la cabecera 'Item' (normalizada) en la hoja."
                )
            header_row, _item_col = item_pos
            header_row_idx: int = header_row

            input_header_map: Dict[str, int] = {}
            for c in range(1, ws.max_column + 1):
                name = _norm_header(ws.cell(row=header_row, column=c).value)
                if name and name not in input_header_map:
                    input_header_map[name] = c

            def _peek_col_value(col: Optional[int], max_scan: int = 25) -> Optional[str]:
                if not col:
                    return None
                max_row = min(ws.max_row or 0, header_row_idx + max_scan)
                for row in ws.iter_rows(
                    min_row=header_row_idx + 1,
                    max_row=max_row,
                    min_col=col,
                    max_col=col,
                    values_only=True,
                ):
                    v = row[0] if row else None
                    if v is None:
                        continue
                    if isinstance(v, str) and not v.strip():
                        continue
                    return _norm_str(v)
                return None

            # AUTO / BASES / GASELAG: se resuelve por archivo en source_type

            def _make_find_input_col(aliases: Dict[str, List[str]]):
                def find_input_col(key: str) -> Optional[int]:
                    col = input_header_map.get(_norm_header(key))
                    if col:
                        return col
                    for alias in aliases.get(key, []):
                        col2 = input_header_map.get(_norm_header(alias))
                        if col2:
                            return col2
                    return None
                return find_input_col

            find_bases = _make_find_input_col(_INPUT_HEADER_ALIASES_BASES)
            missing_bases = [k for k in _REQUIRED_INPUT_KEYS_BASES if not find_bases(k)]

            find_gaselag = _make_find_input_col(_INPUT_HEADER_ALIASES_GASELAG)
            missing_gaselag_v2 = [k for k in _REQUIRED_INPUT_KEYS_GASELAG_V2 if not find_gaselag(k)]
            missing_gaselag_v1 = [k for k in _REQUIRED_INPUT_KEYS_GASELAG_V1 if not find_gaselag(k)]

            gaselag_variant: str | None = None

            if source == "BASES":
                if missing_bases:
                    raise ValueError("Faltan cabeceras requeridas (BASES): " + ", ".join(missing_bases))
                source_type = "BASES"
                find_input_col = find_bases
            elif source == "GASELAG":
                if not missing_gaselag_v2:
                    source_type = "GASELAG"
                    gaselag_variant = "V2"
                    find_input_col = find_gaselag
                elif not missing_gaselag_v1:
                    source_type = "GASELAG"
                    gaselag_variant = "V1"
                    find_input_col = find_gaselag
                else:
                    # reporta el más cercano (útil en soporte)
                    miss = missing_gaselag_v2 if len(missing_gaselag_v2) <= len(missing_gaselag_v1) else missing_gaselag_v1
                    raise ValueError("Faltan cabeceras requeridas (GASELAG): " + ", ".join(miss))
            else:
                # AUTO: detectar por cabeceras + nombre Base Comercial + organismo
                if bases_filename:
                    source_type = "BASES"
                    find_input_col = find_bases
                elif (not missing_bases) and (not missing_gaselag_v2):
                    org_col = find_gaselag("organismo") or find_bases("organismo")
                    org_val = _peek_col_value(org_col)
                    org_norm = (org_val or "").strip().upper()
                    if org_norm == "OI-066":
                        source_type = "GASELAG"
                        gaselag_variant = "V2"
                        find_input_col = find_gaselag
                    elif org_norm == "OI-040":
                        source_type = "BASES"
                        find_input_col = find_bases
                    else:
                        source_type = "GASELAG"
                        gaselag_variant = "V2"
                        find_input_col = find_gaselag
                elif not missing_gaselag_v2:
                    source_type = "GASELAG"
                    gaselag_variant = "V2"
                    find_input_col = find_gaselag
                elif not missing_gaselag_v1:
                    source_type = "GASELAG"
                    gaselag_variant = "V1"
                    find_input_col = find_gaselag
                elif not missing_bases:
                    source_type = "BASES"
                    find_input_col = find_bases
                else:
                    # Reportar el set "más cercano" para debugging
                    if len(missing_bases) <= min(len(missing_gaselag_v2), len(missing_gaselag_v1)):
                        raise ValueError("Faltan cabeceras requeridas (BASES): " + ", ".join(missing_bases))
                    miss = missing_gaselag_v2 if len(missing_gaselag_v2) <= len(missing_gaselag_v1) else missing_gaselag_v1
                    raise ValueError("Faltan cabeceras requeridas (GASELAG): " + ", ".join(miss))

            # Validación por nombre SOLO para BASES
            if source_type == "BASES":
                oi_num, oi_year = _parse_oi_parts_from_filename(fname)
                oi_tag = _parse_oi_tag_from_filename(fname)
            else:
                oi_num = 0
                oi_year = 0
                oi_tag = "GASELAG"

            if source_type == "BASES":
                required_keys = _REQUIRED_INPUT_KEYS_BASES
            else:
                required_keys = (
                    _REQUIRED_INPUT_KEYS_GASELAG_V2
                    if gaselag_variant == "V2"
                    else _REQUIRED_INPUT_KEYS_GASELAG_V1
                )

            file_conformes = 0
            file_no_conformes = 0
            file_no_conforme_series: list[str] = []
            extracted = 0
            ignored_invalid_estado = 0
            invalid_estado_exmples: list[str] = []

            # Construcción de columnas solo de las requeridas del modo
            col_by_key = {key: find_input_col(key) for key in required_keys}
            missing = [key for key, col in col_by_key.items() if not col]
            if missing:
                raise ValueError(
                    "Faltan cabeceras requeridas: " + ", ".join(missing)
                )

            serie_col = col_by_key["medidor"]
            estado_col = col_by_key["estado"]
            relevant_cols = [col for col in col_by_key.values() if col]
            max_col = max(relevant_cols) if relevant_cols else 1

            def _row_value(row: tuple[Any, ...], col: Optional[int]) -> Any:
                if not col:
                    return None
                idx = col - 1
                if idx < 0 or idx >= len(row):
                    return None
                return row[idx]

            def _is_blank_row(row: tuple[Any, ...], cols: list[int]) -> bool:
                for c in cols:
                    v = _row_value(row, c)
                    if v is None:
                        continue
                    if isinstance(v, str) and not v.strip():
                        continue
                    return False
                return True

            # lectura desde la fila siguiente al header "Item", hasta fila completamente vacia
            start_row = header_row + 1
            for r, row in enumerate(
                ws.iter_rows(
                    min_row=start_row,
                    max_row=ws.max_row,
                    min_col=1,
                    max_col=max_col,
                    values_only=True,
                ),
                start=start_row,
            ):
                _raise_cancelled()

                if _is_blank_row(row, relevant_cols):
                    if extracted == 0:
                        continue
                    break

                serie = _norm_str(_row_value(row, serie_col))
                if not serie:
                    continue

                raw_estado = _row_value(row, estado_col)
                estado = _normalize_estado_literal(raw_estado)
                if not estado:
                    ignored_invalid_estado += 1
                    if raw_estado is not None and str(raw_estado).strip():
                        if len(invalid_estado_exmples) < 5:
                            invalid_estado_exmples.append(str(raw_estado)[:80])
                    continue
                    

                if estado == "CONFORME":
                    file_conformes += 1
                else:
                    file_no_conformes += 1
                    file_no_conforme_series.append(serie)

                # construir values para salida
                row_values: Dict[str, Any] = {}
                # inicializar todos los campos esperados en salida
                for key in _OUTPUT_KEYS:
                    if key == "estado":
                        row_values[key] = estado
                        continue
                    if source_type == "GASELAG" and key == "organismo":
                        col_org = find_input_col("organismo") # Opcional en V1, requerido en V2
                        if col_org:
                            row_values[key] = _row_value(row, col_org)
                        else:
                            row_values[key] = "OI-066"
                        continue
                    col = col_by_key.get(key)
                    val = _row_value(row, col)
                    if key == "fecha":
                        row_values[key] = _normalize_input_date(val, epoch=wb.epoch)
                    else:
                        row_values[key] = val

                extracted += 1
                prev = series.get(serie)
                if source_type == "BASES":
                    # Dedupe por serie:
                    # - Se queda con la OI mayor dentro del mismo estado.
                    # - Si existe CONFORME en alguna OI, prevalece sobre NO CONFORME.
                    oi_key = _oi_compare_key(oi_year, oi_num)
                    meta = series_meta.setdefault(
                        serie,
                        {
                            "has_conforme": False,
                            "has_no_conforme": False,
                            "latest_key": None,
                            "latest_estado": None,
                            "latest_oi_year": None,
                            "latest_oi_num": None,
                            "best_conforme_key": None,
                            "best_conforme_oi_year": None,
                            "best_conforme_oi_num": None,
                            "best_no_conforme_key": None,
                            "best_no_conforme_oi_year": None,
                            "best_no_conforme_oi_num": None,
                        },
                    )

                    # track latest overall (para auditoría)
                    if meta["latest_key"] is None or oi_key > meta["latest_key"]:
                        meta["latest_key"] = oi_key
                        meta["latest_estado"] = estado
                        meta["latest_oi_year"] = oi_year or 0
                        meta["latest_oi_num"] = oi_num or 0

                    if estado == "CONFORME":
                        meta["has_conforme"] = True
                        if meta["best_conforme_key"] is None or oi_key > meta["best_conforme_key"]:
                            meta["best_conforme_key"] = oi_key
                            meta["best_conforme_oi_year"] = oi_year or 0
                            meta["best_conforme_oi_num"] = oi_num or 0
                            series[serie] = SerieInfo(
                                oi_num=oi_num or 0,
                                oi_year=oi_year or 0,
                                estado=estado,
                                values=row_values,
                            )
                    else:
                        meta["has_no_conforme"] = True
                        if meta["best_no_conforme_key"] is None or oi_key > meta["best_no_conforme_key"]:
                            meta["best_no_conforme_key"] = oi_key
                            meta["best_no_conforme_oi_year"] = oi_year or 0
                            meta["best_no_conforme_oi_num"] = oi_num or 0
                            # Solo se usa NO CONFORME si no existe CONFORME.
                            if not meta["has_conforme"]:
                                series[serie] = SerieInfo(
                                    oi_num=oi_num or 0,
                                    oi_year=oi_year or 0,
                                    estado=estado,
                                    values=row_values,
                                )
                else:
                    # GASELAG: no hay OI real; mantenemos oi_num=0 para no inventar OI-XXXX
                    if prev is None or (prev.oi_num == 0 and prev.oi_year == 0):
                        # baseline: solo registrar si no existe un registro BASES (oi_num>0)
                        series[serie] = SerieInfo(
                            oi_num=0,
                            oi_year=0,
                            estado=estado,
                            values=row_values,
                        )

            file_no_conforme_series = sorted(set(file_no_conforme_series), key=_natural_key)
            rows_total_read += extracted
            if source_type not in ("BASES", "GASELAG"):
                raise ValueError("No se pudo clasificar el archivo por cabeceras.")
            ok_files += 1

            # Agregado por tipo
            b= by_source.setdefault(source_type, _new_source_bucket())
            b["files_ok"] += 1
            b["rows_read"] += extracted
            b["conformes"] += file_conformes
            b["no_conformes"] += file_no_conformes
            b["rows_ignored_invalid_estado"] += ignored_invalid_estado

            input_conformes_total += file_conformes
            input_no_conformes_total += file_no_conformes
            audit_by_oi.append(
                {
                    "filename": fname,
                    "oi_num": oi_num if source_type == "BASES" else 0,
                    "oi_year": oi_year if source_type == "BASES" else 0,
                    "oi_tag": oi_tag,
                    "source": source_type,
                    "status": "OK",
                    "rows_read": extracted,
                    "conformes": file_conformes,
                    "no_conformes": file_no_conformes,
                    "rows_ignored_invalid_estado": ignored_invalid_estado,
                    "invalid_estado_examples": invalid_estado_exmples,
                    "series_no_conforme_origen": file_no_conforme_series,
                    "error": None,
                }
            )
            _emit(
                operation_id,
                {
                    "type": "status",
                    "stage": "file_done",
                    "message": f"{fname} | Leidos: {extracted} | Conformes: {file_conformes} | No conformes: {file_no_conformes} | Ignorados(estado): {ignored_invalid_estado}",
                    "file": fname,
                    "rows_read": extracted,
                    "conformes": file_conformes,
                    "no_conformes": file_no_conformes,
                    "rows_ignored_invalid_estado": ignored_invalid_estado,
                },
            )

        except Exception as e:
            bad_files += 1
            raw = str(e) or "Error no especificado"
            err_code, err_detail = _split_error_code_detail(raw)
            err_code = err_code or _classify_file_error(raw)
            if not err_detail:
                err_detail = "Error no especificado"
            # En errores debemos registrar el tipo REAL del archivo si ya fue detectado
            # Si aún no se detectó, mantener AUTO e intentar inferir por el texto
            err_source = source_type or "AUTO"
            if err_code == "INVALID_OI_FILENAME":
                err_source = "BASES"
            elif err_source == "AUTO":
                up = raw.upper()
                if "(BASES)" in up:
                    err_source = "BASES"
                elif "(GASELAG)" in up:
                    err_source = "GASELAG"
                else:
                    # Heurística por filename: si parece Base Comercial (OI-####-YYYY) => BASES
                    if _parse_oi_tag_from_filename(fname):
                        err_source = "BASES"
                    else:
                        err_source = "AUTO"
            
            # Si es BASES y el error NO es por filename inválido, intenta extraer oi_num/oi_tag
            if err_source == "BASES" and err_code != "INVALID_OI_FILENAME":
                if (
                    not isinstance(oi_num, int)
                    or oi_num <= 0
                    or not isinstance(oi_year, int)
                    or oi_year <= 0
                ):
                    try:
                        oi_num, oi_year = _parse_oi_parts_from_filename(fname)
                    except Exception:
                        pass
                if not isinstance(oi_tag, str) or not oi_tag:
                    oi_tag = _parse_oi_tag_from_filename(fname)
            elif err_source == "GASELAG":
                oi_num = 0
                oi_year = 0
                oi_tag = "GASELAG"

            # Agregado por tipo
            if err_source != "AUTO":
                b= by_source.setdefault(err_source, _new_source_bucket())
                b["files_error"] += 1

            audit_by_oi.append(
                {
                    "filename": fname,
                    "oi_num": (oi_num if err_source == "BASES" else 0),
                    "oi_year": (oi_year if err_source == "BASES" else 0),
                    "oi_tag": oi_tag,
                    "source": err_source,
                    "status": "ERROR",
                    "rows_read": 0,
                    "conformes": 0,
                    "no_conformes": 0,
                    "error": raw,
                }
            )
            files_rejected.append(
                {
                    "filename": fname,
                    "oi_num": (oi_num if err_source == "BASES" else 0),
                    "oi_year": (oi_year if err_source == "BASES" else 0),
                    "oi_tag": oi_tag,
                    "code": err_code,
                    "detail": err_detail,
                    "source": err_source,
                }
            )
            _emit(
                operation_id,
                {
                    "type": "status",
                    "stage": "file_error",
                    "message": f"{fname} | ERROR: {err_code} | {err_detail}",
                    "file": fname,
                    "code": err_code,
                    "detail": err_detail,
                },
            )

    _raise_cancelled()

    # 2) Resultado final: solo CONFORMES (orden natural por serie)
    conformes = [s for s, info in series.items() if info.estado == "CONFORME"]
    conformes.sort(key=_natural_key)

    series_total_dedup = len(series)
    series_conformes = len(conformes)
    series_no_conformes_final = series_total_dedup - series_conformes
    series_duplicates_eliminated = max(rows_total_read - series_total_dedup, 0)

    # Auditoría: series con NO CONFORME más reciente que una CONFORME
    conflict_series: List[Dict[str, Any]] = []
    for serie, meta in series_meta.items():
        if (
            meta.get("has_conforme")
            and meta.get("has_no_conforme")
            and meta.get("latest_estado") == "NO CONFORME"
        ):
            conflict_series.append(
                {
                    "serie": serie,
                    "latest_oi_year": meta.get("latest_oi_year"),
                    "latest_oi_num": meta.get("latest_oi_num"),
                    "latest_estado": meta.get("latest_estado"),
                    "best_conforme_oi_year": meta.get("best_conforme_oi_year"),
                    "best_conforme_oi_num": meta.get("best_conforme_oi_num"),
                    "best_no_conforme_oi_year": meta.get("best_no_conforme_oi_year"),
                    "best_no_conforme_oi_num": meta.get("best_no_conforme_oi_num"),
                }
            )
    conflict_series.sort(key=lambda x: _natural_key(x.get("serie") or ""))

    # 3) Render a plantilla LOG01 (fila 2+, item desde 1)
    st = get_settings()
    template_path = getattr(st, "log01_template_abs_path", None)
    if not template_path:
        template_path = str((st.data_dir / "templates" / "logistica" / "LOG01_PLANTILLA_SALIDA.xlsx").resolve())

    wb_out = load_workbook(template_path)
    ws_out = next((w for w in wb_out.worksheets if w.title.strip().upper() == "BD"), wb_out.worksheets[0])

    # --- Render por cabeceras (robusto ante cambios de plantilla) ---
    # 1) Encontrar fila de cabecera: buscamos "item" en las primeras 30 filas
    header_row = None
    max_scan_rows = min(30, ws_out.max_row or 30)
    max_scan_cols = min(ws_out.max_column or 50, 80)
    for r in range(1, max_scan_rows + 1):
        for c in range(1, max_scan_cols + 1):
            if _norm_header(ws_out.cell(row=r, column=c).value) == "item":
                header_row = r
                break
        if header_row:
            break
    if not header_row:
        raise ValueError("No se encontró la cabecera 'Item' en la hoja BD de la plantilla.")

    data_start_row = header_row + 1

    # 2) Construir mapa cabecera->columna según plantilla
    header_map: Dict[str, int] = {}
    for c in range(1, (ws_out.max_column or max_scan_cols) + 1):
        h = _norm_header(ws_out.cell(row=header_row, column=c).value)
        if h and h not in header_map:
            header_map[h] = c

    # Aliases mínimos de cabecera en plantilla (por si usan nombres distintos)
    # Clave: key interno; valores: posibles headers en plantilla
    template_aliases: Dict[str, List[str]] = {
        "item": ["item"],
        "medidor": ["medidor", "serie", "nro serie", "nro. serie", "nro de serie", "numero de serie", "número de serie"],
        "q3": ["q3"],
        "error_q3": ["error q3", "error q3 (%)", "error q3 %"],
        "q2": ["q2"],
        "error_q2": ["error q2", "error q2 (%)", "error q2 %"],
        "q1": ["q1"],
        "error_q1": ["error q1", "error q1 (%)", "error q1 %"],
        "estado_pe": ["estado pe", "ensayo de presion estatica", "resultado p estatica", "resultado de p estatica"],        
        "fecha": ["fecha", "fecha de ejecucion", "fecha de ejecución"],
        "certificado": ["certificado", "numero de certificado", "número de certificado"],
        "estado": ["estado", "conclusion", "conclusión"],
        "precinto": ["precinto"],
        "banco_numero": ["banco numero", "numero de banco", "número de banco", "numero de banco de ensayo"],
        "certificado_banco": ["certificado banco", "numero de certificado del banco", "número de certificado del banco"],
        "organismo": ["organismo", "organismo de inspeccion", "organismo de inspección"],
    }

    def _find_out_col(key: str) -> Optional[int]:
        col = header_map.get(_norm_header(key))
        if col:
            return col
        for alias in template_aliases.get(key, []):
            col = header_map.get(_norm_header(alias))
            if col:
                return col
        return None

    col_item = _find_out_col("item") or 1
    out_cols: Dict[str, int] = {}
    missing_out: List[str] = []
    for key in _OUTPUT_KEYS:
        col = _find_out_col(key)
        if col:
            out_cols[key] = col
        else:
            missing_out.append(key)
    # En plantilla corporativa podrían existir columnas extra; pero si faltan claves críticas, fallar claro
    critical = {"medidor", "estado", "q3", "q2", "q1"}
    if critical.intersection(set(missing_out)):
        raise ValueError(
            "La plantilla BD no contiene cabeceras requeridas: " + ", ".join(sorted(critical.intersection(set(missing_out))))
        )

    # 3) Limpiar filas de datos previas (solo en columnas relevantes)
    # Usar fila modelo = primera fila de datos (data_start_row) para copiar estilo
    model_row = data_start_row
    # determinar hasta dónde limpiar (si hay contenido anterior)
    max_clear_row = ws_out.max_row or model_row
    cols_to_clear = {col_item, *out_cols.values()}
    for r in range(data_start_row, max_clear_row + 1):
        for c in cols_to_clear:
            ws_out.cell(row=r, column=c).value = None

    # 4) Escribir datos
    out_row = data_start_row
    item_counter = 1
    for i, serie in enumerate(conformes, start=1):
        _raise_cancelled()
        info = series[serie]
        vals = info.values

        # copiar estilo desde model_row a la nueva fila (si la fila ya existe o se expande)
        if out_row != model_row:
            for c in cols_to_clear:
                src = ws_out.cell(row=model_row, column=c)
                dst = ws_out.cell(row=out_row, column=c)
                dst._style = src._style
                dst.number_format = src.number_format

        # item
        cell_item = _writable_cell(ws_out, out_row, col_item)
        cell_item.value = item_counter
        _apply_output_format(ws_out, out_row, col_item, "item")

        for key, col in out_cols.items():
            v = vals.get(key)
            if key == "fecha":
                v = _normalize_output_date(v)
                cell = _writable_cell(ws_out, out_row, col)
                cell.value = v
                # asegurar formato dd/mm/yyyy si la celda existe
                try:
                    cell.number_format = "dd/mm/yyyy"
                except Exception:
                    pass
            else:
                if isinstance(v, float) and abs(v) < 1e-9:
                    v = 0.0
                if key in ("q3", "q2", "q1") and isinstance(v, (int, float)) and not isinstance(v, bool):
                    v = round(float(v), 2)
                elif key in ("error_q3", "error_q2", "error_q1") and isinstance(v, (int, float)) and not isinstance(v, bool):
                    v = round(float(v), 1)
                cell = _writable_cell(ws_out, out_row, col)
                cell.value = v
            _apply_output_format(ws_out, out_row, col, key)

        out_row += 1
        item_counter += 1

        # progreso
        if i % 50 == 0 or i == len(conformes):
            _emit(
                operation_id,
                {
                    "type": "status",
                    "stage": "render",
                    "message": f"Escribiendo plantilla ({i}/{len(conformes)})",
                    "progress": 70.0 + (30.0 * i / max(len(conformes), 1)),
                },
            )

    # serializar xlsx
    out_buf = BytesIO()
    wb_out.save(out_buf)
    xlsx_bytes = out_buf.getvalue()

    # nombre sugerido
    if output_filename and output_filename.strip():
        out_name = output_filename.strip()
    else:
        if conformes:
            out_name = f"BD_{conformes[0]}_AL_{conformes[-1]}.xlsx"
        else:
            out_name = "BD_SIN_CONFORMES.xlsx"

    summary = {
        "files_total": total_files,
        "files_ok": ok_files,
        "files_error": bad_files,
        "rows_total_read": rows_total_read,
        # Rango de series (post-dedupe, SOLO CONFORMES)
        "serie_ini": conformes[0] if conformes else None,
        "serie_fin": conformes[-1] if conformes else None,
        "series_duplicates_eliminated": series_duplicates_eliminated,
        "series_total_dedup": series_total_dedup,
        "series_conformes": series_conformes,
        "series_no_conformes_final": series_no_conformes_final,
        # Auditoría "de origen"
        "audit_by_oi": audit_by_oi,
        "files_rejected": files_rejected,
        "totals_input": {
            "rows_read": rows_total_read,
            "conformes": input_conformes_total,
            "no_conformes": input_no_conformes_total,
        },
        # Agregado por tipo
        "by_source": by_source,
        # Detalle técnico
        "detail": {
            "series_duplicates_eliminated": series_duplicates_eliminated,
            "series_conflict_no_conforme_post_conforme": {
                "count": len(conflict_series),
                "items": conflict_series[:100],
            },
        },
    }

    # Completar files_total por tipo (ok + error)
    for _k, _v in by_source.items():
        _v["files_total"] = int(_v.get("files_ok", 0)) + int(_v.get("files_error", 0))

    _emit(
        operation_id,
        {"type": "complete", "message": "Consolidación completada", "percent": 100.0, "result": summary},
    )

    stem = Path(out_name).stem
    no_conforme_filename = f"{stem}_NO_CONFORME_FINAL.json"
    manifest_filename = f"{stem}_MANIFIESTO.json"

    # Mapear (oi_year, oi_num) -> oi_tag desde auditoría (primer match)
    oi_key_to_tag: Dict[tuple[int, int], str] = {}
    for a in audit_by_oi:
        oi = a.get("oi_num")
        oi_year = a.get("oi_year")
        tag = a.get("oi_tag")
        if (
            isinstance(oi, int)
            and isinstance(oi_year, int)
            and oi > 0
            and oi_year > 0
            and isinstance(tag, str)
            and tag
        ):
            oi_key_to_tag.setdefault((oi_year, oi), tag)

    def _tag_for(oi_year: int, oi_num: int) -> str:
        if oi_num == 0:
            return "GASELAG"
        tag = oi_key_to_tag.get((oi_year, oi_num))
        if tag:
            return tag
        if oi_year:
            return f"OI-{oi_num:04d}-{oi_year:04d}"
        return f"OI-{oi_num:04d}"

    # NO CONFORME final (post-dedupe)
    no_conforme_series = [s for s, info in series.items() if info.estado == "NO CONFORME"]
    no_conforme_series.sort(key=_natural_key)

    no_conforme_payload = {
        "operation_id": operation_id,
        "generated_at": datetime.utcnow().isoformat(timespec="seconds") + "Z",
        "total_no_conforme_final": len(no_conforme_series),
        "items": [
            {
                "oi": _tag_for(series[s].oi_year, series[s].oi_num),
                "oi_num": series[s].oi_num,
                "oi_year": series[s].oi_year,
                "serie": s,
            }
            for s in no_conforme_series
        ],
    }
    no_conforme_json = json.dumps(no_conforme_payload, ensure_ascii=False, indent=2).encode("utf-8")

    # Manifiesto por OI (listas para LOG-02)
    by_oi: Dict[tuple[int, int], Dict[str, Any]] = {}
    for serie, info in series.items():
        oi_key = (info.oi_year, info.oi_num)
        bucket = by_oi.setdefault(
            oi_key,
            {
                "oi": _tag_for(info.oi_year, info.oi_num),
                "oi_num": info.oi_num,
                "oi_year": info.oi_year,
                "series_no_conforme": [],
                "series_conforme": [],
            },
        )
        if info.estado == "NO CONFORME":
            bucket["series_no_conforme"].append(serie)
        else:
            bucket["series_conforme"].append(serie)

    for bucket in by_oi.values():
        bucket["series_no_conforme"].sort(key=_natural_key)
        bucket["total_no_conforme"] = len(bucket["series_no_conforme"])
        bucket["series_conforme"].sort(key=_natural_key)
        bucket["total_conforme"] = len(bucket["series_conforme"])
        bucket["total_series"] = bucket["total_no_conforme"] + bucket["total_conforme"]

    by_oi_origen_map: Dict[str, Dict[str, Any]] = {}
    for a in audit_by_oi:
        if a.get("status") != "OK":
            continue
        series_list = a.get("series_no_conforme_origen") or []
        if not isinstance(series_list, list):
            continue
        oi_num = a.get("oi_num")
        oi_year = a.get("oi_year")
        oi_tag = a.get("oi_tag")
        if not isinstance(oi_tag, str) or not oi_tag:
            if isinstance(oi_num, int):
                oi_tag = _tag_for(oi_year if isinstance(oi_year, int) else 0, oi_num)
            else:
                continue
        bucket = by_oi_origen_map.setdefault(
            oi_tag,
            {
                "oi": oi_tag,
                "oi_num": oi_num if isinstance(oi_num, int) else None,
                "oi_year": oi_year if isinstance(oi_year, int) else None,
                "source_files": [],
                "series_no_conforme": set(),
            },
        )
        if isinstance(oi_num, int) and bucket.get("oi_num") is None:
            bucket["oi_num"] = oi_num
        if isinstance(oi_year, int) and bucket.get("oi_year") is None:
            bucket["oi_year"] = oi_year
        fname = a.get("filename")
        if isinstance(fname, str) and fname:
            bucket["source_files"].append(fname)
        for serie in series_list:
            if isinstance(serie, str) and serie:
                bucket["series_no_conforme"].add(serie)

    by_oi_origen: list[Dict[str, Any]] = []
    origin_no_conformes_total = 0
    for bucket in by_oi_origen_map.values():
        series_sorted = sorted(bucket["series_no_conforme"], key=_natural_key)
        bucket["series_no_conforme"] = series_sorted
        bucket["source_files"] = sorted(set(bucket["source_files"]))
        bucket["total_no_conforme"] = len(series_sorted)
        origin_no_conformes_total += len(series_sorted)
        by_oi_origen.append(bucket)
    by_oi_origen.sort(
        key=lambda x: (
            x.get("oi_year") or 0,
            x.get("oi_num") or 0,
            x.get("oi") or "",
        )
    )

    manifest_payload = {
        "operation_id": operation_id,
        "generated_at": datetime.utcnow().isoformat(timespec="seconds") + "Z",
        "output_excel": out_name,
        "source": source,
        "totals": {
            "files_total": total_files,
            "files_ok": ok_files,
            "files_error": bad_files,
            "rows_read": rows_total_read,
            "series_total_dedup": series_total_dedup,
            "series_conformes": series_conformes,
            "series_no_conformes_final": len(no_conforme_series),
            "series_duplicates_eliminated": series_duplicates_eliminated,
            "files_rejected": len(files_rejected),
            "origin_no_conformes_total": origin_no_conformes_total,
            "final_no_conformes_total": len(no_conforme_series),
        },
        "by_oi_origen": by_oi_origen,
        "by_oi": sorted(
            by_oi.values(),
            key=lambda x: (x.get("oi_year") or 0, x.get("oi_num") or 0, x.get("oi") or ""),
        ),
    }
    manifest_json = json.dumps(manifest_payload, ensure_ascii=False, indent=2).encode("utf-8")

    return Log01ProcessResult(
        xlsx_bytes=xlsx_bytes,
        out_name=out_name,
        summary=summary,
        no_conforme_json=no_conforme_json,
        manifest_json=manifest_json,
        no_conforme_filename=no_conforme_filename,
        manifest_filename=manifest_filename,
    )
